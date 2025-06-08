#!/usr/bin/env python
# coding: utf-8

"""
Refactored Within- and Between-Subjects Tests Analysis

This module analyzes experimental data across four waves plus LLM simulations,
computing various statistical tests and generating an Excel report.
"""

import pandas as pd
import numpy as np
from scipy.stats import ttest_ind, ttest_1samp, chi2_contingency, f_oneway, binomtest, pearsonr
from statsmodels.stats.proportion import proportion_confint, proportions_ztest
import statsmodels.formula.api as smf
from linearmodels.panel import PanelOLS
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import argparse
import os
from typing import Dict, List, Tuple, Optional, Union, Any
from abc import ABC, abstractmethod
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# Data Loading and Preprocessing
# ============================================================================

class DataLoader:
    """Handles loading and preprocessing of experimental data."""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.label_files = {
            "wave1": os.path.join("data", "wave_csv", "wave_1_labels_anonymized.csv"),
            "wave2": os.path.join("data", "wave_csv", "wave_2_labels_anonymized.csv"),
            "wave3": os.path.join("data", "wave_csv", "wave_3_labels_anonymized.csv"),
            "wave4": os.path.join("data", "wave_csv", "wave_4_labels_anonymized.csv"),
            "LLM": os.path.join(output_dir, "csv_comparison/csv_formatted_label/responses_llm_imputed_label_formatted.csv"),
        }
        self._setup_pandas_alias()
        
    def _setup_pandas_alias(self):
        """Set up pandas read_csv alias to handle TWIN_ID compatibility."""
        _orig_read_csv = pd.read_csv
        
        def read_csv_alias(*args, **kwargs):
            df = _orig_read_csv(*args, **kwargs)
            if "TWIN_ID" in df.columns:
                df["PROLIFIC_PID"] = df["TWIN_ID"].astype(str)
            return df
            
        pd.read_csv = read_csv_alias
    
    def valid_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filter to valid completed responses."""
        condition1 = (df["PROLIFIC_PID"].astype(str).str.strip() != "")
        condition2 = (df["Finished"].astype(str).str.strip().str.upper() == "TRUE")
        return df[condition1 & condition2].copy()
    
    def get_common_ids(self) -> set:
        """Get IDs of respondents who completed all 4 waves."""
        common_ids = None
        for wave, file in self.label_files.items():
            if wave == "LLM":
                continue
            df = pd.read_csv(file, low_memory=False)
            df = self.valid_rows(df)
            ids = set(df["PROLIFIC_PID"].astype(str).str.strip())
            if common_ids is None:
                common_ids = ids
            else:
                common_ids = common_ids.intersection(ids)
        # Number of respondents who finished all 4 waves: {len(common_ids)}
        return common_ids
    
    def load_wave_data(self, wave: str, common_ids: set) -> pd.DataFrame:
        """Load and filter data for a specific wave."""
        df = pd.read_csv(self.label_files[wave], low_memory=False)
        
        if wave == "LLM":
            # LLM data uses TWIN_ID
            df = df[df["TWIN_ID"].astype(str).str.strip() != ""]
            df["TWIN_ID"] = df["TWIN_ID"].astype(str).str.strip()
            df = df[df["TWIN_ID"].str.isnumeric()]
        else:
            # Human data uses PROLIFIC_PID
            df = self.valid_rows(df)
            df = df[df["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]
            df["PROLIFIC_PID"] = df["PROLIFIC_PID"].astype(str).str.strip()
        
        df.columns = [col.strip() for col in df.columns]
        return df

# ============================================================================
# Excel Writer Utilities
# ============================================================================

class ExcelWriter:
    """Handles writing results to Excel with proper formatting."""
    
    def __init__(self, filename: str):
        self.filename = filename
        
    def get_unique_sheet_name(self, base_name: str) -> str:
        """Generate a unique sheet name by appending suffix if needed."""
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet_name = base_name
            suffix = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_v{suffix}"
                suffix += 1
            wb.close()
            return sheet_name
        except FileNotFoundError:
            # File doesn't exist yet, so use base name
            return base_name
    
    def write_results(self, sheet_name: str, results: List[Tuple[str, pd.DataFrame]], 
                     header_note: str = None, mode: str = 'a'):
        """Write multiple result tables to a sheet with proper spacing."""
        # Check if file exists and adjust mode accordingly
        file_exists = os.path.exists(self.filename)
        actual_mode = 'w' if not file_exists else mode
        
        # Configure writer parameters based on mode
        writer_kwargs = {'engine': 'openpyxl', 'mode': actual_mode}
        if actual_mode == 'a':
            writer_kwargs['if_sheet_exists'] = 'overlay'
            
        with pd.ExcelWriter(self.filename, **writer_kwargs) as writer:
            if sheet_name not in writer.book.sheetnames:
                ws = writer.book.create_sheet(sheet_name)
            else:
                ws = writer.sheets[sheet_name]
            
            startrow = 1
            
            # Write header note if provided
            if header_note:
                ws.cell(row=startrow, column=1, value=header_note)
                startrow += 3
            
            # Write each result table with title
            for title, df in results:
                if title:
                    ws.cell(row=startrow, column=1, value=title).font = Font(bold=True, size=12)
                    startrow += 2
                
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
                startrow += len(df) + 3

# ============================================================================
# Base Analysis Classes
# ============================================================================

class BaseAnalysis(ABC):
    """Abstract base class for all analyses."""
    
    def __init__(self, excel_writer: ExcelWriter):
        self.excel_writer = excel_writer
        
    @abstractmethod
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run the analysis on the provided data."""
        pass
    
    def compute_descriptive_stats(self, series: pd.Series, name: str = None) -> Dict:
        """Compute standard descriptive statistics."""
        stats = {
            "N": series.shape[0],
            "Mean": round(series.mean(), 2),
            "Median": round(series.median(), 2),
            "SD": round(series.std(), 2),
            "Min": series.min(),
            "Max": series.max()
        }
        if name:
            stats["Variable"] = name
        return stats

# ============================================================================
# Specific Analysis Classes
# ============================================================================

class BaseRateAnalysis(BaseAnalysis):
    """Base rate neglect analysis (Kahneman & Tversky, 1973)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1-Base Rate Neglect"
        self.wave1_cols = ["Form B - 70 eng _1", "Q156_1"]
        self.wave4_cols = ["Form A _1", "Q156_1"]
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run base rate analysis for all waves."""
        results = []
        
        # Wave 1 analysis
        w1_results = self._analyze_wave(data["wave1"], self.wave1_cols, "Wave 1")
        results.extend(w1_results)
        
        # LLM analysis
        llm_results = self._analyze_wave(data["LLM"], self.wave4_cols, "LLM")
        results.extend(llm_results)
        
        # Wave 4 analysis
        w4_results = self._analyze_wave(data["wave4"], self.wave4_cols, "Wave 4")
        results.extend(w4_results)
        
        header_note = (
            "Analysis: Two-sample t-test comparing participants' responses across the two base rate conditions.\n"
            "Each participant responded to one question only. This test evaluates whether the average responses differ significantly."
        )
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, cols: List[str], wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze base rate for a single wave."""
        results = []
        
        # Convert to numeric
        for col in cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Compute descriptive stats
        desc_stats_list = []
        for col in cols:
            data = df[col].dropna()
            desc_stats_list.append(self.compute_descriptive_stats(data, col))
        desc_stats_df = pd.DataFrame(desc_stats_list)
        
        # Two-sample t-test
        data1 = df[cols[0]].dropna()
        data2 = df[cols[1]].dropna()
        
        if len(data1) > 1 and len(data2) > 1:
            t_stat, p_val = ttest_ind(data1, data2, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
        
        ttest_df = pd.DataFrame([{
            "Comparison": f"{cols[0]} vs. {cols[1]}",
            "Mean 1": round(data1.mean(), 2),
            "Mean 2": round(data2.mean(), 2),
            "t-statistic": round(t_stat, 3),
            "p-value": round(p_val, 3),
            "N1": len(data1),
            "N2": len(data2)
        }])
        
        results.append((f"{wave_label} - Descriptive Statistics", desc_stats_df))
        results.append((f"{wave_label} - T-Test Results", ttest_df))
        
        return results

class OutcomeBiasAnalysis(BaseAnalysis):
    """Outcome bias analysis (Baron & Hershey, 1988)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1-Outcome Bias"
        self.outcome_cols = ["Q161", "Q162"]
        self.outcome_mapping = {
            "Clearly correct, an excellent decision": 3,
            "Correct, all things considered": 2,
            "Correct, but the opposite would be reasonable too": 1,
            "The decision and its opposite are equally good": 0,
            "Incorrect, but not unreasonable": -1,
            "Incorrect, all things considered": -2,
            "Incorrect, a very bad decision": -3
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run outcome bias analysis for all waves."""
        results = []
        
        # Analyze each wave
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df, wave_label)
            results.extend(wave_results)
        
        # Add response distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        header_note = (
            "Analysis: Outcome Bias (Baron & Hershey, 1988).\n"
            "Responses for Q161 and Q162 were originally text-based Likert ratings (ranging from -3 to +3).\n"
            "Per the paper, we define outcome bias as the mean rating in the positive outcome condition (Q161_conv) minus\n"
            "the mean rating in the negative outcome condition (Q162_conv). A positive and statistically significant\n"
            "difference supports the presence of outcome bias."
        )
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze outcome bias for a single wave."""
        results = []
        
        # Map responses
        df["Q161_conv"] = df["Q161"].str.strip().map(self.outcome_mapping)
        df["Q162_conv"] = df["Q162"].str.strip().map(self.outcome_mapping)
        
        # Compute descriptive stats
        desc_stats_list = []
        for col in ["Q161_conv", "Q162_conv"]:
            data = df[col].dropna()
            desc_stats_list.append(self.compute_descriptive_stats(data, col))
        desc_stats_df = pd.DataFrame(desc_stats_list)
        
        # Compute outcome bias
        group_pos = df["Q161_conv"].dropna()
        group_neg = df["Q162_conv"].dropna()
        
        outcome_bias = group_pos.mean() - group_neg.mean()
        
        if len(group_pos) > 1 and len(group_neg) > 1:
            t_stat, p_val = ttest_ind(group_pos, group_neg, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
        
        ttest_df = pd.DataFrame([{
            "Outcome Bias (Mean Difference)": round(outcome_bias, 3),
            "t-statistic": round(t_stat, 3),
            "p-value": round(p_val, 3)
        }])
        
        results.append((f"{wave_label} - Descriptive Statistics", desc_stats_df))
        results.append((f"{wave_label} - T-Test Results", ttest_df))
        
        return results
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions across waves."""
        dist_list = []
        
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            # Map responses
            df["Q161_conv"] = df["Q161"].str.strip().map(self.outcome_mapping)
            df["Q162_conv"] = df["Q162"].str.strip().map(self.outcome_mapping)
            
            for q in ["Q161_conv", "Q162_conv"]:
                dist = df[q].value_counts(dropna=False).sort_index()
                dist_df = pd.DataFrame({
                    "Wave": wave_label,
                    "Question": q,
                    "Converted Response": dist.index,
                    "Count": dist.values
                })
                dist_list.append(dist_df)
        
        all_dist = pd.concat(dist_list, ignore_index=True)
        return [("Converted Response Distributions", all_dist)]

class FalseConsensusAnalysis(BaseAnalysis):
    """False consensus analysis (Furnas & LaPira, 2024)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1-False Consensus"
        self.self_cols = [f"False Cons. self _{i}" for i in range(1, 11)]
        self.others_cols = [f"False cons. others _{i}" for i in [1,2,3,4,5,6,7,10,11,12]]
        self.self_mapping = {
            "strongly support": 2,
            "somewhat support": 1,
            "neither oppose nor support": 0,
            "somewhat oppose": -1,
            "strongly oppose": -2
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run false consensus analysis for all waves."""
        results = []
        models = {}  # Store models for fit summary
        long_dfs = {}  # Store long dataframes for distributions
        wave_dfs = {}  # Store original dataframes for raw distributions
        
        # Analyze each wave
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results, model, long_df = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
            if model is not None:
                models[wave_label] = model
            if long_df is not None:
                long_dfs[wave_label] = long_df
            wave_dfs[wave_label] = df  # Store original df
        
        # Add distributions
        dist_results = self._get_distributions(long_dfs)
        results.extend(dist_results)
        
        # Add model fit summary
        fit_results = self._get_model_fit_summary(models)
        results.extend(fit_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze false consensus for a single wave."""
        results = []
        model = None
        long_df = None
        
        # Map self responses
        for col in self.self_cols:
            df[col + "_num"] = df[col].astype(str).str.lower().str.strip().map(self.self_mapping)
        
        # Convert others columns to numeric
        for col in self.others_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Compute public support
        public_support = {}
        for i in range(1, 11):
            self_col = f"False Cons. self _{i}_num"
            valid_responses = df[self_col].dropna()
            n_total = valid_responses.shape[0]
            n_support = valid_responses[(valid_responses == 1) | (valid_responses == 2)].shape[0]
            support_pct = (n_support / n_total) * 100 if n_total > 0 else np.nan
            public_support[i] = support_pct
        
        public_support_df = pd.DataFrame([
            {"Item": i, "Policy": f"Policy {i}", f"Support_{wave_label}": v}
            for i, v in public_support.items()
        ])
        
        # Build long-format data
        long_data = []
        for i in range(1, 11):
            self_col = f"False Cons. self _{i}_num"
            others_col = f"False cons. others _{i}" if i <= 7 else f"False cons. others _{i+2}"
            
            if wave_label == "LLM":
                id_col = "TWIN_ID"
            else:
                id_col = "PROLIFIC_PID"
                
            df_item = df[[id_col, self_col, others_col]].dropna()
            if not df_item.empty:
                df_item = df_item.rename(columns={
                    id_col: "PROLIFIC_PID",
                    self_col: "Self",
                    others_col: "Predicted"
                })
                df_item["Policy"] = f"Policy {i}"
                df_item["Actual"] = public_support[i]
                df_item["Misperception"] = df_item["Predicted"] - df_item["Actual"]
                df_item["Item"] = i
                long_data.append(df_item)
        
        if long_data:
            long_df = pd.concat(long_data, ignore_index=True)
            long_df = long_df.dropna(subset=['Misperception'])
            long_df = long_df[np.isfinite(long_df['Misperception'])]
            
            if not long_df.empty:
                # Run PanelOLS regression for all waves
                regression_results, model = self._run_panel_regression(long_df)
                if regression_results is not None:
                    results.append((f"{wave_label} - Two-Way Fixed Effects Regression", regression_results))
        
        results.append((f"{wave_label} - Public Support", public_support_df))
        
        return results, model, long_df
    
    def _run_panel_regression(self, long_df: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], Optional[Any]]:
        """Run PanelOLS regression for large fixed effects."""
        long_df = long_df.set_index(['PROLIFIC_PID', 'Item'])
        
        # Create dummy variables
        long_df["Self_cat"] = pd.Categorical(long_df["Self"], categories=[-2, -1, 0, 1, 2])
        long_df["Self_cat_str"] = long_df["Self_cat"].astype(str)
        
        self_dummies = pd.get_dummies(long_df['Self_cat_str'], prefix='Self_cat_str')
        if 'Self_cat_str_0' in self_dummies.columns:
            self_dummies = self_dummies.drop('Self_cat_str_0', axis=1)
        
        policy_dummies = pd.get_dummies(long_df['Policy'], prefix='Policy')
        policy_dummies = policy_dummies.drop(policy_dummies.columns[0], axis=1)
        
        X = pd.concat([self_dummies, policy_dummies], axis=1)
        y = long_df['Misperception']
        
        try:
            model = PanelOLS(y, X, entity_effects=True).fit()
            
            # Extract Self_cat_str effects
            self_effects_list = []
            for col in self_dummies.columns:
                if col.startswith('Self_cat_str_'):
                    self_rating = float(col.replace('Self_cat_str_', ''))
                    if col in model.params.index:
                        self_effects_list.append({
                            "Self_Rating": self_rating,
                            "Coefficient": model.params[col],
                            "Std_Error": model.std_errors[col],
                            "t_value": model.tstats[col],
                            "p_value": model.pvalues[col],
                            "CI_lower": model.conf_int().loc[col, 'lower'],
                            "CI_upper": model.conf_int().loc[col, 'upper']
                        })
            
            self_effects_df = pd.DataFrame(self_effects_list).sort_values('Self_Rating').reset_index(drop=True)
            return self_effects_df, model
            
        except Exception as e:
            # Error in PanelOLS regression: {e}
            return None, None
    
    def _get_distributions(self, long_dfs: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions across waves."""
        results = []
        
        if not long_dfs:
            return results
        
        # Distribution of Self Ratings - matching original order
        self_dist_list = []
        for wave_label in ["Wave 1", "LLM", "Wave 4"]:  # Fixed order
            if wave_label in long_dfs and long_dfs[wave_label] is not None and not long_dfs[wave_label].empty:
                df = long_dfs[wave_label]
                self_dist = df["Self"].value_counts(dropna=False).sort_index().reset_index()
                self_dist.columns = ["Self Rating", "Count"]
                self_dist.insert(0, "Wave", wave_label)
                self_dist_list.append(self_dist)
        
        if self_dist_list:
            combined_self_dist = pd.concat(self_dist_list, ignore_index=True)
            results.append(("Distribution of Self Ratings", combined_self_dist))
        
        # Distribution of Actual Public Support (%) - matching original order
        actual_dist_list = []
        for wave_label in ["Wave 1", "LLM", "Wave 4"]:  # Fixed order
            if wave_label in long_dfs and long_dfs[wave_label] is not None and not long_dfs[wave_label].empty:
                df = long_dfs[wave_label]
                actual_dist = df["Actual"].round().value_counts(dropna=False).sort_index().reset_index()
                actual_dist.columns = ["Actual (%)", "Count"]
                actual_dist.insert(0, "Wave", wave_label)
                actual_dist_list.append(actual_dist)
        
        if actual_dist_list:
            combined_actual_dist = pd.concat(actual_dist_list, ignore_index=True)
            results.append(("Distribution of Actual Public Support (%)", combined_actual_dist))
        
        return results
    
    def _get_model_fit_summary(self, models: Dict[str, Any]) -> List[Tuple[str, pd.DataFrame]]:
        """Get model fit summary for all waves."""
        results = []
        
        # Create a summary table with model fit statistics
        fit_summary_list = []
        
        for wave_label, model in models.items():
            if model is not None:
                try:
                    # Extract statistics from PanelOLS model
                    fit_summary_list.append({
                        "Wave": wave_label,
                        "R-squared": round(model.rsquared, 3) if hasattr(model, 'rsquared') else np.nan,
                        "Adj. R-squared": round(model.rsquared_adj, 3) if hasattr(model, 'rsquared_adj') else np.nan,
                        "F-statistic": round(model.f_statistic.stat, 3) if hasattr(model, 'f_statistic') and hasattr(model.f_statistic, 'stat') else np.nan,
                        "p-value (F)": round(model.f_statistic.pval, 3) if hasattr(model, 'f_statistic') and hasattr(model.f_statistic, 'pval') else np.nan,
                        "n_obs": int(model.nobs) if hasattr(model, 'nobs') else np.nan
                    })
                except Exception as e:
                    # Error extracting fit statistics for {wave_label}: {e}
                    pass
                    # Fallback to basic info
                    fit_summary_list.append({
                        "Wave": wave_label,
                        "R-squared": np.nan,
                        "Adj. R-squared": np.nan,
                        "F-statistic": np.nan,
                        "p-value (F)": np.nan,
                        "n_obs": np.nan
                    })
        
        if fit_summary_list:
            model_fit_summary = pd.DataFrame(fit_summary_list)
            results.append(("Model Fit Summary (TWFE)", model_fit_summary))
        
        return results
    

class SunkCostAnalysis(BaseAnalysis):
    """Sunk cost fallacy analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1-Sunk Cost"
        self.sunk_cost_cols = ["Q181", "Q182"]
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run sunk cost analysis for all waves."""
        results = []
        
        header_note = (
            "Analysis: Sunk Cost Fallacy.\n"
            "Participants in the no-sunk-cost condition (Q181) and sunk-cost condition (Q182) provided\n"
            "responses on a scale of 0-20.\n\n"
            "We expect that participants in the no-sunk-cost condition will show a higher score\n"
            "(indicating a stronger preference for convenience) than those in the sunk-cost condition.\n"
            "A significant independent samples t-test would support the presence of a sunk cost effect."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze sunk cost for a single wave."""
        import math
        results = []
        
        # Convert to numeric
        for col in self.sunk_cost_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Compute descriptive stats
        desc_stats_list = []
        for col in self.sunk_cost_cols:
            data = df[col].dropna()
            desc_stats_list.append({
                "Condition": col,
                "N": data.shape[0],
                "Mean": round(data.mean(), 2),
                "SD": round(data.std(), 2),
                "Median": round(data.median(), 2),
                "Min": data.min(),
                "Max": data.max()
            })
        desc_stats_df = pd.DataFrame(desc_stats_list)
        
        # Perform t-test
        group_no_sunk = df["Q181"].dropna()
        group_sunk = df["Q182"].dropna()
        
        if len(group_no_sunk) > 1 and len(group_sunk) > 1:
            t_stat, p_val = ttest_ind(group_no_sunk, group_sunk, equal_var=False)
            
            # Cohen's d
            n1, n2 = len(group_no_sunk), len(group_sunk)
            s1, s2 = group_no_sunk.std(), group_sunk.std()
            pooled_sd = math.sqrt(((n1 - 1) * s1**2 + (n2 - 1) * s2**2) / (n1 + n2 - 2))
            cohens_d = (group_no_sunk.mean() - group_sunk.mean()) / pooled_sd
        else:
            t_stat, p_val, cohens_d = np.nan, np.nan, np.nan
        
        ttest_df = pd.DataFrame([{
            "No-sunk-cost Mean": round(group_no_sunk.mean(), 2),
            "Sunk-cost Mean": round(group_sunk.mean(), 2),
            "t-statistic": round(t_stat, 3),
            "p-value": round(p_val, 3),
            "Cohen's d": round(cohens_d, 3)
        }])
        
        results.append((f"{wave_label} - Descriptive Statistics", desc_stats_df))
        results.append((f"{wave_label} - T-Test Results", ttest_df))
        
        return results

class AllaisProblemAnalysis(BaseAnalysis):
    """Allais problem analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1-Allais Problem"
        self.allais_cols = ["Q192", "Q193"]
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run Allais problem analysis for all waves."""
        results = []
        
        header_note = (
            "Analysis: Allais Problem.\n\n"
            "Participants were randomly assigned to one of two forms of the Allais problem.\n"
            "Form 1 (Q192) presents a choice between:\n"
            "  Option A: One million dollars for sure\n"
            "  Option B: 89% chance of one million dollars, 10% chance of five million dollars, 1% chance of nothing\n\n"
            "Form 2 (Q193) presents a choice between:\n"
            "  Option C: 11% chance of one million dollars, 89% chance of nothing\n"
            "  Option D: 10% chance of five million dollars, 90% chance of nothing\n\n"
            "Responses were mapped to options A, B for Form 1 and C, D for Form 2.\n"
            "For completeness, we perform binomial tests for each form against a null of 0.5. \n"
            "Stanovich & West expect most participants to choose Option A in Form 1 and Option D in Form 2."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze Allais problem for a single wave."""
        results = []
        
        # Map responses
        df["Q192_mapped"] = df["Q192"].apply(self._map_allais_form1)
        df["Q193_mapped"] = df["Q193"].apply(self._map_allais_form2)
        
        # Frequency tables
        freq_list = []
        for col, form_name in [("Q192_mapped", "Form 1 (Q192)"), ("Q193_mapped", "Form 2 (Q193)")]:
            counts = df[col].value_counts(dropna=True).sort_index()
            percents = df[col].value_counts(normalize=True, dropna=True).sort_index() * 100
            freq_df = pd.DataFrame({
                "Option": counts.index,
                "Count": counts.values,
                "Percent": percents.round(1).values
            })
            freq_df["Form"] = form_name
            freq_list.append(freq_df)
        
        if freq_list:
            freq_combined = pd.concat(freq_list, ignore_index=True)
            freq_combined = freq_combined[["Form", "Option", "Count", "Percent"]]
        else:
            freq_combined = pd.DataFrame()
        
        # Binomial tests
        binom_results = []
        
        # Form 1: Test if Option A > 0.5
        form1_data = df["Q192_mapped"].dropna()
        if len(form1_data) > 0:
            n_form1 = len(form1_data)
            n_A = (form1_data == "A").sum()
            result_form1 = binomtest(n_A, n_form1, p=0.5, alternative='greater')
            binom_results.append({
                "Form": f"Form 1 (Q192) - {wave_label}",
                "n_total": n_form1,
                "n_target": n_A,
                "Proportion_target": round(n_A/n_form1, 3),
                "p-value": round(result_form1.pvalue, 3)
            })
        
        # Form 2: Test if Option D > 0.5
        form2_data = df["Q193_mapped"].dropna()
        if len(form2_data) > 0:
            n_form2 = len(form2_data)
            n_D = (form2_data == "D").sum()
            result_form2 = binomtest(n_D, n_form2, p=0.5, alternative='greater')
            binom_results.append({
                "Form": f"Form 2 (Q193) - {wave_label}",
                "n_total": n_form2,
                "n_target": n_D,
                "Proportion_target": round(n_D/n_form2, 3),
                "p-value": round(result_form2.pvalue, 3)
            })
        
        binom_df = pd.DataFrame(binom_results) if binom_results else pd.DataFrame()
        
        if not freq_combined.empty:
            results.append((f"{wave_label} - Frequency Table", freq_combined))
        if not binom_df.empty:
            results.append((f"{wave_label} - Binomial Tests", binom_df))
        
        return results
    
    def _map_allais_form1(self, resp):
        """Map Form 1 responses."""
        if pd.isna(resp):
            return np.nan
        text = resp.lower()
        if "for sure" in text or "100" in text:
            return "A"
        elif "89%" in text and "10%" in text and "1%" in text:
            return "B"
        else:
            return "Unknown"
    
    def _map_allais_form2(self, resp):
        """Map Form 2 responses."""
        if pd.isna(resp):
            return np.nan
        text = resp.lower()
        if "11%" in text and "89%" in text:
            return "C"
        elif "10%" in text and "90%" in text:
            return "D"
        else:
            return "Unknown"

class NonseparabilityAnalysis(BaseAnalysis):
    """Nonseparability of risks and benefits analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W1 Nonseparability"
        self.tech_names = {
            1: "Bicycles",
            2: "Alcoholic beverages", 
            3: "Chemical plants",
            4: "Pesticides"
        }
        self.bene_mapping = {
            "not at all beneficial": 1,
            "low benefit": 2,
            "slightly beneficial": 3,
            "neutral": 4,
            "moderately beneficial": 5,
            "very beneficial": 6,
            "extremely beneficial": 7
        }
        self.risk_mapping = {
            "not at all risky": 1,
            "low risk": 2,
            "slightly risky": 3,
            "neutral": 4,
            "moderately risky": 5,
            "very risky": 6,
            "extremely risky": 7
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run nonseparability analysis for all waves."""
        results = []
        
        # Analyze each wave
        for wave_label, df in [("Wave 1", data["wave1"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze nonseparability for a single wave."""
        results = []
        
        # Detect columns
        bene_cols = sorted([col for col in df.columns if col.lower().startswith("nonseparabilty bene")])
        risk_cols = sorted([col for col in df.columns if col.lower().startswith("nonseparability ris")])
        
        # Map responses
        for col in bene_cols:
            df[col + "_num"] = df[col].apply(lambda x: self._robust_map_response(x, self.bene_mapping))
        for col in risk_cols:
            df[col + "_num"] = df[col].apply(lambda x: self._robust_map_response(x, self.risk_mapping))
        
        bene_num_cols = [col + "_num" for col in bene_cols]
        risk_num_cols = [col + "_num" for col in risk_cols]
        
        # Analyze each technology
        results_list = []
        for i in range(min(len(bene_num_cols), len(risk_num_cols))):
            bene_col = bene_num_cols[i]
            risk_col = risk_num_cols[i]
            df_item = df[[bene_col, risk_col]].dropna()
            n = df_item.shape[0]
            
            if n > 1:
                r, p = pearsonr(df_item[bene_col], df_item[risk_col])
                t_stat = r * np.sqrt((n-2) / (1-r**2)) if n > 2 else np.nan
            else:
                r, p, t_stat = np.nan, np.nan, np.nan
                
            results_list.append({
                "Item": i+1,
                "Technology": self.tech_names.get(i+1, f"Item {i+1}"),
                "n": n,
                "Mean Benefit": round(df_item[bene_col].mean(), 2) if n > 0 else np.nan,
                "Mean Risk": round(df_item[risk_col].mean(), 2) if n > 0 else np.nan,
                "Correlation": round(r, 3) if not np.isnan(r) else np.nan,
                "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
                "p-value": round(p, 3) if not np.isnan(p) else np.nan
            })
        
        results_df = pd.DataFrame(results_list)
        results.append((f"{wave_label} Results", results_df))
        
        return results
    
    def _robust_map_response(self, resp, mapping):
        """Map response to numeric value."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        for key, val in mapping.items():
            if key in text:
                return float(val)
        return np.nan

# ============================================================================
# Wave 2 Analysis Classes
# ============================================================================

class FramingAnalysis(BaseAnalysis):
    """Framing problem analysis (Tversky & Kahneman, 1981)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Framing"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run framing analysis for relevant waves."""
        results = []
        
        header_note = (
            "Analysis: Framing Problem \n\n"
            "Participants were randomly assigned to one of two versions of the framing problem:\n\n"
            "Positive Frame (Q157):\n"
            "  - If Program A is adopted, 200 people will be saved.\n"
            "  - If Program B is adopted, there is a 1/3 probability that 600 people will be saved, and a 2/3 probability that no people will be saved.\n\n"
            "Negative Frame (Q158):\n"
            "  - If Program A is adopted, 400 people will die.\n"
            "  - If Program B is adopted, there is a 1/3 probability that nobody will die, and a 2/3 probability that 600 people will die.\n\n"
            "Responses were provided on a 6-point Likert scale:\n"
            "  1 = strongly favor Program A, 2 = slightly favor Program A, 3 = favor Program A,\n"
            "  4 = favor Program B, 5 = slightly favor Program B, 6 = strongly favor Program B.\n\n"
            "Expectation from Tversky & Kahneman is that in the loss frame, participants will be more inclined to choose Program B -> higher scores.\n\n"
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze framing for a single wave."""
        results = []
        
        # Map responses
        df["Q157_mapped"] = df["Q157"].apply(self._map_framing_response)
        df["Q158_mapped"] = df["Q158"].apply(self._map_framing_response)
        
        # Separate groups
        group_gain = df.loc[df["Q157_mapped"].notna(), "Q157_mapped"]
        group_loss = df.loc[df["Q158_mapped"].notna(), "Q158_mapped"]
        
        # Descriptive stats
        summary_df = pd.DataFrame({
            "Group": ["Gain Frame (Q157)", "Loss Frame (Q158)"],
            "n": [group_gain.shape[0], group_loss.shape[0]],
            "Mean": [round(group_gain.mean(), 2), round(group_loss.mean(), 2)],
            "SD": [round(group_gain.std(ddof=1), 2), round(group_loss.std(ddof=1), 2)]
        })
        
        # T-test
        if group_gain.shape[0] > 1 and group_loss.shape[0] > 1:
            t_stat, p_val = ttest_ind(group_gain, group_loss, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
            
        ttest_df = pd.DataFrame({
            "t-statistic": [round(t_stat, 3)],
            "p-value": [round(p_val, 3)]
        })
        
        results.append((f"{wave_label} Summary", summary_df))
        results.append((f"{wave_label} T-Test Results", ttest_df))
        
        return results
    
    def _map_framing_response(self, resp):
        """Map framing response to 6-point scale."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        
        if "program a" in text:
            if "strongly" in text:
                return 1
            elif "slightly" in text:
                return 3
            else:
                return 2
        elif "program b" in text:
            if "strongly" in text:
                return 6
            elif "slightly" in text:
                return 4
            else:
                return 5
        else:
            return np.nan
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions."""
        dist_list = []
        
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            df["Q157_mapped"] = df["Q157"].apply(self._map_framing_response)
            df["Q158_mapped"] = df["Q158"].apply(self._map_framing_response)
            
            for q in ["Q157_mapped", "Q158_mapped"]:
                counts = df[q].value_counts(dropna=False).sort_index()
                for val, count in counts.items():
                    dist_list.append({
                        "Wave": wave_label,
                        "Question": q,
                        "Mapped Value": val,
                        "Count": count
                    })
        
        dist_df = pd.DataFrame(dist_list)
        return [("Distribution of Mapped Responses", dist_df)]

class LindaProblemAnalysis(BaseAnalysis):
    """Conjunction fallacy / Linda problem analysis (Tversky & Kahneman, 1983)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Linda"
        self.linda_mapping = {
            "extremely improbable": 1,
            "very improbable": 2,
            "somewhat probable": 3,
            "moderately probable": 4,
            "very probable": 5,
            "extremely probable": 6
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run Linda problem analysis."""
        results = []
        
        header_note = (
            "Analysis: Linda Problem \n\n"
            "Participants read a description of Linda and then completed one of two sets of questions.\n\n"
            "Condition 1 (Q159): Subjects answered three questions, with the third asking:\n"
            "  'It is ___ that Linda is a bank teller.'\n\n"
            "Condition 2 (Q160): Subjects answered three questions, with the third asking:\n"
            "  'It is ___ that Linda is a bank teller and is active in the feminist movement.'\n\n"
            "Responses were provided on a 6-point Likert scale with text options such as:\n"
            "  'Extremely improbable', 'Very improbable', 'Somewhat probable', 'Moderately probable', 'Very probable', 'Extremely probable'.\n"
            "These were mapped numerically (1 = Extremely improbable, ..., 6 = Extremely probable).\n\n"
            "If subjects exhibit the conjunction fallacy, the mean probability rating for the conjunction\n"
            "condition (Q160_3) will be higher than for the single event condition (Q159_3).\n"
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze Linda problem for a single wave."""
        results = []
        
        # Map responses
        df["Linda_Single"] = df["Q159_3"].apply(lambda x: self._robust_map_linda(x, self.linda_mapping))
        df["Linda_Conjunction"] = df["Q160_3"].apply(lambda x: self._robust_map_linda(x, self.linda_mapping))
        
        # Separate groups
        group_single = df.loc[df["Linda_Single"].notna(), "Linda_Single"]
        group_conj = df.loc[df["Linda_Conjunction"].notna(), "Linda_Conjunction"]
        
        # Descriptive stats
        summary_df = pd.DataFrame({
            "Group": ["Single Event (Q159_3)", "Conjunction (Q160_3)"],
            "n": [group_single.shape[0], group_conj.shape[0]],
            "Mean": [round(group_single.mean(), 2), round(group_conj.mean(), 2)],
            "SD": [round(group_single.std(ddof=1), 2), round(group_conj.std(ddof=1), 2)],
            "Median": [round(group_single.median(), 2), round(group_conj.median(), 2)],
            "Min": [group_single.min(), group_conj.min()],
            "Max": [group_single.max(), group_conj.max()]
        })
        
        # T-test
        if group_single.shape[0] > 1 and group_conj.shape[0] > 1:
            t_stat, p_val = ttest_ind(group_single, group_conj, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
            
        ttest_df = pd.DataFrame({
            "t-statistic": [round(t_stat, 3)],
            "p-value": [round(p_val, 3)]
        })
        
        results.append((f"{wave_label} Summary", summary_df))
        results.append((f"{wave_label} T-Test Results", ttest_df))
        
        return results
    
    def _robust_map_linda(self, resp, mapping):
        """Map Linda response to numeric value."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        for key, val in mapping.items():
            if key in text:
                return float(val)
        return np.nan
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions."""
        dist_list = []
        
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            df["Linda_Single"] = df["Q159_3"].apply(lambda x: self._robust_map_linda(x, self.linda_mapping))
            df["Linda_Conjunction"] = df["Q160_3"].apply(lambda x: self._robust_map_linda(x, self.linda_mapping))
            
            for q in ["Linda_Single", "Linda_Conjunction"]:
                counts = df[q].value_counts(dropna=False).sort_index()
                for val, count in counts.items():
                    dist_list.append({
                        "Wave": wave_label,
                        "Question": q,
                        "Mapped Value": val,
                        "Count": count
                    })
        
        dist_df = pd.DataFrame(dist_list)
        return [("Distribution of Mapped Responses", dist_df)]

class AnchoringAnalysis(BaseAnalysis):
    """Anchoring and adjustment analysis."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Anchoring"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run anchoring analysis."""
        results = []
        
        header_note = (
            "Analysis: Anchoring Effects \n\n"
            "Participants were assigned to either a low or high anchor condition in two domains:\n\n"
            "African Countries:\n"
            "  - Low Anchor: Q163 asks about whether there are more or fewer than 12 African countries in the UN, then Q164 asks for a numeric estimate.\n"
            "  - High Anchor: Q165 asks about whether there are more or fewer than 65 African countries in the UN, then Q166 asks for a numeric estimate.\n\n"
            "Trees:\n"
            "  - Low Anchor: Q167 asks if the tallest redwood is more or less than 85 feet tall, then Q168 asks for a numeric estimate.\n"
            "  - High Anchor: Q169 asks if the tallest redwood is more or less than 1000 feet tall, then Q170 asks for a numeric estimate.\n\n"
            "For each domain, descriptive statistics (n, mean, SD, median, min, max) of the numeric estimates are reported by anchor condition,\n"
            "and an independent samples t-test compares the high vs. low anchor groups.\n"
            "In both domains, we expect subjects anchored high will provide higher estimates than the subjects anchored low."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze anchoring for a single wave."""
        results = []
        
        # Convert to numeric
        for col in ["Q164", "Q166", "Q168", "Q170"]:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # African countries analysis
        africa_results = self._analyze_domain(
            df, "African Countries", 
            low_col="Q164", high_col="Q166",
            low_anchor="12", high_anchor="65"
        )
        
        # Tree analysis
        tree_results = self._analyze_domain(
            df, "Trees",
            low_col="Q168", high_col="Q170", 
            low_anchor="85 feet", high_anchor="1000 feet"
        )
        
        # Combine results
        combined_summary = pd.concat([
            africa_results["summary"],
            pd.DataFrame([[""]*africa_results["summary"].shape[1]], columns=africa_results["summary"].columns),
            tree_results["summary"]
        ], ignore_index=True)
        
        combined_ttest = pd.concat([africa_results["ttest"], tree_results["ttest"]], ignore_index=True)
        
        results.append((f"{wave_label} Summary Statistics", combined_summary))
        results.append((f"{wave_label} T-Test Results", combined_ttest))
        
        return results
    
    def _analyze_domain(self, df: pd.DataFrame, domain: str, low_col: str, high_col: str, 
                       low_anchor: str, high_anchor: str) -> Dict:
        """Analyze one anchoring domain."""
        low_data = df.loc[df[low_col].notna(), low_col]
        high_data = df.loc[df[high_col].notna(), high_col]
        
        summary = pd.DataFrame({
            "Domain": [domain, domain],
            "Anchor": [f"Low ({low_anchor})", f"High ({high_anchor})"],
            "n": [low_data.shape[0], high_data.shape[0]],
            "Mean": [round(low_data.mean(), 2), round(high_data.mean(), 2)],
            "SD": [round(low_data.std(ddof=1), 2), round(high_data.std(ddof=1), 2)],
            "Median": [round(low_data.median(), 2), round(high_data.median(), 2)],
            "Min": [low_data.min(), high_data.min()],
            "Max": [low_data.max(), high_data.max()]
        })
        
        if low_data.shape[0] > 1 and high_data.shape[0] > 1:
            t_stat, p_val = ttest_ind(high_data, low_data, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
        
        ttest = pd.DataFrame({
            "Domain": [domain],
            "t-statistic": [round(t_stat, 3)],
            "p-value": [round(p_val, 3)]
        })
        
        return {"summary": summary, "ttest": ttest}

# ============================================================================
# Additional Wave 2 Analysis Classes
# ============================================================================

class RelativeSavingsAnalysis(BaseAnalysis):
    """Absolute vs Relative Savings analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Relative Savings"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run relative savings analysis."""
        results = []
        
        header_note = (
            "Analysis: Relative vs. Absolute Savings \n\n"
            "Two scenarios were presented to participants:\n\n"
            "1. Calculator scenario (Q183):\n"
            "   - You go to purchase a calculator for $30. The salesperson informs you that the calculator is on sale for $20 at the other branch (a $10 saving).\n\n"
            "2. Jacket scenario (Q184):\n"
            "   - You go to purchase a jacket for $250. The salesperson informs you that the jacket is on sale for $240 at the other branch (a $10 saving).\n\n"
            "The absolute saving is the same (at $10), but we expect people will more likely drive for the calculator than for the jacket.\n\n"
            "Chi-square test was used to compare binary categorical choices."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze relative savings for a single wave."""
        results = []
        
        # Map responses
        df["Q183_mapped"] = df["Q183"].apply(self._map_yes_no)
        df["Q184_mapped"] = df["Q184"].apply(self._map_yes_no)
        
        # Ensure each subject appears in only one condition
        mask = ((df["Q183_mapped"].notna()) ^ (df["Q184_mapped"].notna()))
        df_clean = df[mask]
        
        # Define groups
        group_calc = df_clean.loc[df_clean["Q183_mapped"].notna(), "Q183_mapped"]
        group_jacket = df_clean.loc[df_clean["Q184_mapped"].notna(), "Q184_mapped"]
        
        # Frequency tables
        freq_list = []
        for series, cond in [(group_calc, "Calculator (Q183)"), (group_jacket, "Jacket (Q184)")]:
            counts = series.value_counts(dropna=True).sort_index()
            percents = series.value_counts(normalize=True, dropna=True).sort_index() * 100
            freq_df = pd.DataFrame({
                "Condition": cond,
                "Response": counts.index,
                "Count": counts.values,
                "Percent": percents.round(1).values
            })
            freq_list.append(freq_df)
        
        if freq_list:
            freq_combined = pd.concat(freq_list, ignore_index=True)
        else:
            freq_combined = pd.DataFrame()
        
        # Chi-square test
        yes_calc = (group_calc == "yes").sum()
        no_calc = (group_calc == "no").sum()
        yes_jacket = (group_jacket == "yes").sum()
        no_jacket = (group_jacket == "no").sum()
        
        contingency = np.array([[yes_calc, no_calc], [yes_jacket, no_jacket]])
        
        if contingency.min() > 0:
            chi2, p_chi2, dof, expected = chi2_contingency(contingency)
            chi2_df = pd.DataFrame({
                "Chi2": [round(chi2, 3)],
                "Degrees of Freedom": [dof],
                "p-value": [round(p_chi2, 3)]
            })
        else:
            chi2_df = pd.DataFrame()
        
        # Summary table
        desc_df = pd.DataFrame({
            "Condition": ["Calculator (Q183)", "Jacket (Q184)"],
            "n": [group_calc.shape[0], group_jacket.shape[0]],
            "Yes (%)": [round(yes_calc/(yes_calc+no_calc)*100, 1) if (yes_calc+no_calc) > 0 else np.nan,
                       round(yes_jacket/(yes_jacket+no_jacket)*100, 1) if (yes_jacket+no_jacket) > 0 else np.nan],
            "No (%)": [round(no_calc/(yes_calc+no_calc)*100, 1) if (yes_calc+no_calc) > 0 else np.nan,
                      round(no_jacket/(yes_jacket+no_jacket)*100, 1) if (yes_jacket+no_jacket) > 0 else np.nan]
        })
        
        results.append((f"{wave_label} - Descriptive Statistics", desc_df))
        if not freq_combined.empty:
            results.append((f"{wave_label} - Frequency Table", freq_combined))
        if not chi2_df.empty:
            results.append((f"{wave_label} - Chi-Square Test", chi2_df))
        
        return results
    
    def _map_yes_no(self, resp):
        """Map yes/no responses."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).strip().lower()
        if "yes" in text:
            return "yes"
        elif "no" in text:
            return "no"
        else:
            return np.nan

class MysideBiasAnalysis(BaseAnalysis):
    """Myside bias analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Myside"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run myside bias analysis."""
        results = []
        
        header_note = (
            "Analysis: Myside Bias \n\n"
            "Two versions of the question were administered between subjects:\n\n"
            "Version 1 (Q194): Ford Explorer ban question in Germany.\n"
            "Version 2 (Q195): German car ban question in U.S.\n\n"
            "Responses were on a 6-point Likert scale:\n"
            "1 = Definitely No, 2 = No, 3 = Probably No, 4 = Probably Yes, 5 = Yes, 6 = Definitely Yes.\n\n"
            "Independent samples t-test compares the mean responses between conditions."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze myside bias for a single wave."""
        results = []
        
        # Map responses
        df["Q194_mapped"] = df["Q194"].apply(self._map_likert_scale)
        df["Q195_mapped"] = df["Q195"].apply(self._map_likert_scale)
        
        # Sort so rows with non-missing Q194 come first
        df = df.sort_values(by="Q194_mapped", na_position="last")
        # Then drop duplicates, keeping the non-missing response if available
        if wave_label == "LLM":
            df = df.drop_duplicates(subset="TWIN_ID", keep="first")
        else:
            df = df.drop_duplicates(subset="PROLIFIC_PID", keep="first")
        
        # Ensure each subject appears in only one condition
        mask = (df["Q194_mapped"].notna()) ^ (df["Q195_mapped"].notna())
        df = df[mask]
        
        # Define groups
        group_germany = df.loc[df["Q194_mapped"].notna(), "Q194_mapped"]
        group_us = df.loc[df["Q195_mapped"].notna(), "Q195_mapped"]
        
        # Descriptive statistics
        desc_df = pd.DataFrame({
            "Condition": ["Ford (Q194)", "German (Q195)"],
            "n": [group_germany.shape[0], group_us.shape[0]],
            "Mean": [round(group_germany.mean(), 2) if group_germany.shape[0] > 0 else np.nan,
                    round(group_us.mean(), 2) if group_us.shape[0] > 0 else np.nan],
            "SD": [round(group_germany.std(ddof=1), 2) if group_germany.shape[0] > 1 else np.nan,
                  round(group_us.std(ddof=1), 2) if group_us.shape[0] > 1 else np.nan],
            "Median": [round(group_germany.median(), 2) if group_germany.shape[0] > 0 else np.nan,
                      round(group_us.median(), 2) if group_us.shape[0] > 0 else np.nan],
            "Min": [group_germany.min() if group_germany.shape[0] > 0 else np.nan,
                   group_us.min() if group_us.shape[0] > 0 else np.nan],
            "Max": [group_germany.max() if group_germany.shape[0] > 0 else np.nan,
                   group_us.max() if group_us.shape[0] > 0 else np.nan]
        })
        
        # T-test
        if group_germany.shape[0] > 1 and group_us.shape[0] > 1:
            t_stat, p_val = ttest_ind(group_germany, group_us, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
            
        ttest_df = pd.DataFrame({
            "Statistic": ["t-statistic", "p-value"],
            "Value": [round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
                     round(p_val, 3) if not np.isnan(p_val) else np.nan]
        })
        
        results.append((f"{wave_label} - Descriptive Statistics", desc_df))
        results.append((f"{wave_label} - T-Test Results", ttest_df))
        
        return results
    
    def _map_likert_scale(self, resp):
        """Map Likert scale responses."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).strip().lower()
        if "definitely no" in text:
            return 1
        elif "probably no" in text:
            return 3
        elif text == "no":
            return 2
        elif "probably yes" in text:
            return 4
        elif "definitely yes" in text:
            return 6
        elif text == "yes":
            return 5
        else:
            return np.nan
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions."""
        dist_list = []
        
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            for q in ["Q194", "Q195"]:
                df[f"{q}_mapped"] = df[q].apply(self._map_likert_scale)
                dist = df[f"{q}_mapped"].value_counts(dropna=False).sort_index()
                for val, count in dist.items():
                    dist_list.append({
                        "Wave": wave_label,
                        "Question": q,
                        "Converted Response": val,
                        "Count": count
                    })
        
        if dist_list:
            all_dist = pd.DataFrame(dist_list)
            return [("Converted Response Distributions", all_dist)]
        else:
            return []

class OmissionBiasAnalysis(BaseAnalysis):
    """Omission bias analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W2-Omission"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run omission bias analysis."""
        results = []
        
        header_note = (
            "Analysis: Omission Bias \n\n"
            "Scenario:\n"
            "  There will be a deadly flu in your area next winter. Your doctor says you have a 10% chance of dying from the flu.\n"
            "  A new flu vaccine is available that prevents you from catching the flu but carries a 5% risk of death from a weaker flu virus.\n\n"
            "Participants answered the question on a 4-point Likert scale:\n"
            "  'I would definitely not...', 'I would probably not...', 'I would probably...', 'I would definitely...'\n\n"
            "Per Stanovich & West, we then bin subjects into two categories:\n"
            "  - Low Vaccination: responses 1 or 2\n"
            "  - High Vaccination: responses 3 or 4\n\n"
            "We compute 95% confidence intervals for the proportion of responses in each group."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 2", data["wave2"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze omission bias for a single wave."""
        results = []
        
        # Determine column name based on what exists in the dataframe
        if "Omission bias " in df.columns:
            omission_col = "Omission bias "
        elif "Omission bias" in df.columns:
            omission_col = "Omission bias"
        else:
            # Return empty results if column not found
            # Warning: Omission bias column not found in {wave_label} data
            pass
            return results
        
        # Map responses
        df["Omission_mapped"] = df[omission_col].apply(self._map_omission)
        df["Omission_Group"] = df["Omission_mapped"].apply(self._assign_group)
        
        # Add temporary column to prioritize rows with valid responses
        df["has_response"] = df["Omission_mapped"].notna().astype(int)
        
        # Sort so valid (non-NaN) responses are first
        df = df.sort_values(by="Omission_mapped", na_position="last")
        
        # Drop duplicates — this will keep the non-NaN version if it exists
        if wave_label == "LLM":
            df = df.drop_duplicates(subset="TWIN_ID", keep="first")
        else:
            df = df.drop_duplicates(subset="PROLIFIC_PID", keep="first")
        
        # Now filter to keep only those with valid responses
        df = df[df["Omission_mapped"].notna()]
        
        # Frequency table
        freq_df = df["Omission_Group"].value_counts(dropna=True).sort_index().reset_index()
        freq_df.columns = ["Group", "Count"]
        freq_df["Percent"] = (freq_df["Count"] / freq_df["Count"].sum() * 100).round(1)
        
        # Compute confidence intervals
        n_total = df.shape[0]
        n_low = df[df["Omission_Group"] == "Low Vaccination"].shape[0]
        n_high = df[df["Omission_Group"] == "High Vaccination"].shape[0]
        
        low_prop = n_low / n_total if n_total > 0 else 0
        high_prop = n_high / n_total if n_total > 0 else 0
        
        if n_total > 0:
            low_ci_low, low_ci_high = proportion_confint(count=n_low, nobs=n_total, alpha=0.05, method="wilson")
            high_ci_low, high_ci_high = proportion_confint(count=n_high, nobs=n_total, alpha=0.05, method="wilson")
        else:
            low_ci_low, low_ci_high = 0, 0
            high_ci_low, high_ci_high = 0, 0
        
        ci_df = pd.DataFrame({
            "Group": ["Low Vaccination", "High Vaccination"],
            "Proportion": [round(low_prop, 3), round(high_prop, 3)],
            "95% CI Lower": [round(low_ci_low, 3), round(high_ci_low, 3)],
            "95% CI Upper": [round(low_ci_high, 3), round(high_ci_high, 3)],
            "n": [n_low, n_high]
        })
        
        results.append((f"{wave_label} - Frequency Table", freq_df))
        results.append((f"{wave_label} - Confidence Intervals", ci_df))
        
        return results
    
    def _map_omission(self, resp):
        """Map omission responses to numeric scale."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        if "definitely not" in text:
            return 1
        elif "probably not" in text:
            return 2
        elif "definitely" in text:
            return 4
        elif "probably" in text:
            return 3
        else:
            return np.nan
    
    def _assign_group(self, x):
        """Assign to low or high vaccination group."""
        if pd.isna(x):
            return np.nan
        if x in [1, 2]:
            return "Low Vaccination"
        elif x in [3, 4]:
            return "High Vaccination"
        else:
            return np.nan

# ============================================================================
# Wave 3 Analysis Classes
# ============================================================================

class LessIsMoreAnalysis(BaseAnalysis):
    """Less is More analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W3-Less is More"
        self.formA_cols = ["Q171", "Q174", "Q177"]
        self.formB_cols = ["Q172", "Q175", "Q178"]
        self.formC_cols = ["Q173", "Q176", "Q179"]
        self.question_mappings = {
            "Gamble": {"Form A": "Q171_num", "Form B": "Q172_num", "Form C": "Q173_num"},
            "Proportion Dominance 1": {"Form A": "Q174_num", "Form B": "Q175_num", "Form C": "Q176_num"},
            "Proportion Dominance 2": {"Form A": "Q177_num", "Form B": "Q178_num", "Form C": "Q179_num"}
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run less is more analysis."""
        results = []
        
        # Analyze each wave
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze less is more for a single wave."""
        results = []
        
        # Map responses
        for col in self.formA_cols + self.formB_cols + self.formC_cols:
            df[col + "_num"] = df[col].apply(self._map_less_is_more)
        
        # Assign forms
        df["Form_Assignment"] = df.apply(self._assign_form, axis=1)
        df_valid = df[df["Form_Assignment"].notna()]
        
        # Analyze each question
        for question, mapping in self.question_mappings.items():
            scores = self._get_question_scores(df_valid, mapping)
            
            # Descriptive stats
            desc_df = pd.DataFrame({
                "Form": list(scores.keys()),
                "n": [s.size for s in scores.values()],
                "Mean Score": [round(s.mean(), 2) for s in scores.values()],
                "SD": [round(s.std(ddof=1), 2) for s in scores.values()],
                "Median": [round(s.median(), 2) for s in scores.values()],
                "Min": [s.min() for s in scores.values()],
                "Max": [s.max() for s in scores.values()]
            })
            
            # ANOVA
            if sum(len(s) > 0 for s in scores.values()) >= 2:
                available_scores = [s for s in scores.values() if len(s) > 0]
                F_stat, p_val = f_oneway(*available_scores)
                k = len(available_scores)
                n = sum([len(s) for s in available_scores])
                df_between = k - 1
                df_within = n - k
            else:
                F_stat, p_val, df_between, df_within = np.nan, np.nan, np.nan, np.nan
            
            # Include df columns for LLM wave
            if wave_label == "LLM":
                anova_df = pd.DataFrame({
                    "F-statistic": [round(F_stat, 3) if not np.isnan(F_stat) else np.nan],
                    "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan],
                    "df_between": [df_between],
                    "df_within": [df_within]
                })
            else:
                anova_df = pd.DataFrame({
                    "F-statistic": [round(F_stat, 3) if not np.isnan(F_stat) else np.nan],
                    "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan]
                })
            
            results.append((f"Analysis: {question} ({wave_label})", desc_df))
            results.append((f"{question} - ANOVA", anova_df))
        
        return results
    
    def _map_less_is_more(self, resp):
        """Map less is more responses."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).strip().lower()
        
        mapping = {
            "strongly disagree": 1,
            "disagree strongly": 1,
            "very unlikely": 1,
            "disagree a little": 2,
            "moderately disagree": 2,
            "somewhat unlikely": 2,
            "neither agree nor disagree": 3,
            "neutral": 3,
            "slightly disagree": 3,
            "agree a little": 4,
            "slightly agree": 4,
            "somewhat likely": 4,
            "agree strongly": 5,
            "moderately agree": 5,
            "strongly agree": 6,
            "very likely": 6,
            "extremely likely": 6,
            "likely": 5,
        }
        
        for key, val in mapping.items():
            if key in text:
                return val
        return np.nan
    
    def _assign_form(self, row):
        """Assign form based on responses."""
        if not pd.isna(row["Q171_num"]):
            return "Form A"
        elif not pd.isna(row["Q172_num"]):
            return "Form B"
        elif not pd.isna(row["Q173_num"]):
            return "Form C"
        else:
            return np.nan
    
    def _get_question_scores(self, df, form_cols):
        """Get scores for each form."""
        return {
            "Form A": df.loc[df["Form_Assignment"] == "Form A", form_cols["Form A"]],
            "Form B": df.loc[df["Form_Assignment"] == "Form B", form_cols["Form B"]],
            "Form C": df.loc[df["Form_Assignment"] == "Form C", form_cols["Form C"]]
        }
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions."""
        results = []
        
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            dist_list = []
            
            # Map responses
            for col in self.formA_cols + self.formB_cols + self.formC_cols:
                df[col + "_num"] = df[col].apply(self._map_less_is_more)
            
            df["Form_Assignment"] = df.apply(self._assign_form, axis=1)
            df_valid = df[df["Form_Assignment"].notna()]
            
            # Collect distributions
            for form, cols in [("Form A", ["Q171_num", "Q174_num", "Q177_num"]),
                             ("Form B", ["Q172_num", "Q175_num", "Q178_num"]),
                             ("Form C", ["Q173_num", "Q176_num", "Q179_num"])]:
                for col in cols:
                    subset = df_valid[df_valid["Form_Assignment"] == form]
                    counts = subset[col].value_counts(dropna=False).sort_index()
                    total = subset[col].notna().sum()
                    
                    for val, count in counts.items():
                        dist_list.append({
                            "Form": form,
                            "Question": col,
                            "Response": val,
                            "Count": count
                        })
                    
                    # Add non-missing total
                    dist_list.append({
                        "Form": form,
                        "Question": col,
                        "Response": "Non-missing total",
                        "Count": total
                    })
            
            dist_df = pd.DataFrame(dist_list) if dist_list else pd.DataFrame()
            if not dist_df.empty:
                results.append((f"Response Distributions for {wave_label} (Q171–Q179)", dist_df))
        
        return results

class ThalerProblemAnalysis(BaseAnalysis):
    """WTA/WTP Thaler problem analysis."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W3-Thalerproblem"
        self.thaler_mapping = {
            "$5,000,000 or more": 10,
            "$1,000,000": 9,
            "$500,000": 8,
            "$250,000": 7,
            "$100,000": 6,
            "$50,000": 5,
            "$10,000": 4,
            "$1,000": 3,
            "$100": 2,
            "$10": 1
        }
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run Thaler problem analysis."""
        results = []
        
        header_note = (
            "Analysis: Thaler Problem\n\n"
            "Three conditions (each subject sees one):\n"
            "  - WTP-certainty (Q189)\n"
            "  - WTA-certainty (Q190)\n"
            "  - WTP-noncertainty (Q191)\n\n"
            "Text responses (e.g. '$10,000', '$5,000,000 or more') are mapped to 1–10 as in the paper:\n"
            "  1 = $10, 2 = $100, 3 = $1,000, 4 = $10,000, 5 = $50,000,\n"
            "  6 = $100,000, 7 = $250,000, 8 = $500,000, 9 = $1,000,000, 10 = $5,000,000 or more\n\n"
            "We replicate the original two t-tests:\n"
            "  1) WTA-certainty vs. WTP-certainty\n"
            "  2) WTP-certainty vs. WTP-noncertainty\n\n"
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get distributions
        dist_results = self._get_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze Thaler problem for a single wave."""
        results = []
        
        # Map responses
        df["WTP_cert_mapped"] = df["Q189"].apply(self._map_thaler_response)
        df["WTA_cert_mapped"] = df["Q190"].apply(self._map_thaler_response)
        df["WTP_noncert_mapped"] = df["Q191"].apply(self._map_thaler_response)
        
        # Filter valid respondents
        mask = (
            (df["WTP_cert_mapped"].notna() & df["WTA_cert_mapped"].isna() & df["WTP_noncert_mapped"].isna()) |
            (df["WTP_cert_mapped"].isna() & df["WTA_cert_mapped"].notna() & df["WTP_noncert_mapped"].isna()) |
            (df["WTP_cert_mapped"].isna() & df["WTA_cert_mapped"].isna() & df["WTP_noncert_mapped"].notna())
        )
        df = df[mask]
        
        # Groups
        group_wtp_cert = df.loc[df["WTP_cert_mapped"].notna(), "WTP_cert_mapped"]
        group_wta_cert = df.loc[df["WTA_cert_mapped"].notna(), "WTA_cert_mapped"]
        group_wtp_noncert = df.loc[df["WTP_noncert_mapped"].notna(), "WTP_noncert_mapped"]
        
        # Descriptive statistics
        desc_df = pd.DataFrame({
            "Condition": ["WTP-certainty (Q189)", "WTA-certainty (Q190)", "WTP-noncertainty (Q191)"],
            "n": [group_wtp_cert.shape[0], group_wta_cert.shape[0], group_wtp_noncert.shape[0]],
            "Mean": [round(group_wtp_cert.mean(), 2), round(group_wta_cert.mean(), 2), 
                    round(group_wtp_noncert.mean(), 2)],
            "SD": [round(group_wtp_cert.std(ddof=1), 2), round(group_wta_cert.std(ddof=1), 2),
                  round(group_wtp_noncert.std(ddof=1), 2)]
        })
        
        # T-tests
        ttest_results = []
        
        # WTA vs WTP
        if group_wta_cert.shape[0] > 1 and group_wtp_cert.shape[0] > 1:
            t_stat, p_val = ttest_ind(group_wta_cert, group_wtp_cert, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
            
        ttest_results.append({
            "Comparison": "WTA-certainty vs. WTP-certainty",
            "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
            "p-value": round(p_val, 3) if not np.isnan(p_val) else np.nan
        })
        
        # WTP cert vs noncert
        if group_wtp_cert.shape[0] > 1 and group_wtp_noncert.shape[0] > 1:
            t_stat, p_val = ttest_ind(group_wtp_cert, group_wtp_noncert, equal_var=False)
        else:
            t_stat, p_val = np.nan, np.nan
            
        ttest_results.append({
            "Comparison": "WTP-certainty vs. WTP-noncertainty",
            "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
            "p-value": round(p_val, 3) if not np.isnan(p_val) else np.nan
        })
        
        ttest_df = pd.DataFrame(ttest_results)
        
        # Combine descriptive stats and t-tests into one table with blank row
        blank_row = pd.DataFrame([["","","",""]], columns=desc_df.columns)
        combined_df = pd.concat([desc_df, blank_row, ttest_df], ignore_index=True)
        results.append((f"{wave_label} - Analysis", combined_df))
        
        return results
    
    def _map_thaler_response(self, resp):
        """Map Thaler responses to 1-10 scale."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        for key, val in self.thaler_mapping.items():
            short_key = key.lower().replace(",", "").replace(" ", "")
            short_text = text.replace(",", "").replace(" ", "")
            if short_key in short_text:
                return val
        return np.nan
    
    def _get_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get response distributions in wide format."""
        distribution_tables = []
        
        for wave_label, df, prefix in [("Wave 3", data["wave3"], "W3"), 
                                      ("LLM", data["LLM"], "LLM"),
                                      ("Wave 4", data["wave4"], "W4")]:
            # Map responses
            df["WTP_cert_mapped"] = df["Q189"].apply(self._map_thaler_response)
            df["WTA_cert_mapped"] = df["Q190"].apply(self._map_thaler_response)
            df["WTP_noncert_mapped"] = df["Q191"].apply(self._map_thaler_response)
            
            # Filter valid respondents
            mask = (
                (df["WTP_cert_mapped"].notna() & df["WTA_cert_mapped"].isna() & df["WTP_noncert_mapped"].isna()) |
                (df["WTP_cert_mapped"].isna() & df["WTA_cert_mapped"].notna() & df["WTP_noncert_mapped"].isna()) |
                (df["WTP_cert_mapped"].isna() & df["WTA_cert_mapped"].isna() & df["WTP_noncert_mapped"].notna())
            )
            df = df[mask]
            
            # Get distributions for each condition
            for cond, mapped_col in [("WTP-certainty", "WTP_cert_mapped"), 
                                    ("WTA-certainty", "WTA_cert_mapped"),
                                    ("WTP-noncertainty", "WTP_noncert_mapped")]:
                series = df[mapped_col].dropna()
                if len(series) > 0:
                    dist = series.value_counts().sort_index().to_frame()
                    dist.columns = [f"{prefix} {cond}"]
                    distribution_tables.append(dist)
        
        # Combine all distributions into wide format
        if distribution_tables:
            distributions_final = pd.concat(distribution_tables, axis=1).fillna(0).astype(int)
            return [("Response Distributions", distributions_final)]
        else:
            return []

class ProbabilityMatchingAnalysis(BaseAnalysis):
    """Probability Matching vs Maximizing analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W3-PMatchvsMax"
        self.card_cols = [f"Q198_{i}" for i in range(1, 11)]
        self.dice_cols = [f"Q203_{i}" for i in range(1, 7)]
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run probability matching analysis."""
        results = []
        
        header_note = (
            "Analysis: Probability Matching vs. Maximizing \n\n"
            "Card Problem (Q198_1..Q198_10):\n"
            "  - Distribution: 7 '1', 3 '2'. MATCH if exactly 7 '1' & 3 '2', MAX if all '1', OTHER otherwise.\n"
            "Dice Problem (Q203_1..Q203_6):\n"
            "  - Distribution: 4 'red', 2 'green'. MATCH if exactly 4 'red' & 2 'green', MAX if all 'red', OTHER otherwise.\n\n"
            "Frequencies are shown below for each strategy in each task.\n"
            "Stanovich & West observed that, even though the MAX strategy is normative, a majority of subjects used MATCH or OTHER."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        # Get raw distributions
        dist_results = self._get_raw_distributions(data)
        results.extend(dist_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze probability matching for a single wave."""
        results = []
        
        # Fix numeric strings for LLM data
        if wave_label == "LLM":
            for col in self.card_cols:
                df[col] = df[col].apply(lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace('.', '').isdigit() else x)
        
        # Classify responses
        classifications = []
        for idx, row in df.iterrows():
            card_answers = [row.get(c, np.nan) for c in self.card_cols]
            dice_answers = [row.get(c, np.nan) for c in self.dice_cols]
            n_card = sum(not pd.isna(x) for x in card_answers)
            n_dice = sum(not pd.isna(x) for x in dice_answers)
            
            if n_card > 0 and n_dice > 0:
                classifications.append(("BOTH", np.nan))
            elif n_card == 10:
                strategy = self._classify_card_responses(card_answers)
                classifications.append(("Card", strategy))
            elif n_dice == 6:
                strategy = self._classify_dice_responses(dice_answers)
                classifications.append(("Dice", strategy))
            else:
                classifications.append(("None", np.nan))
        
        df["Task"] = [c[0] for c in classifications]
        df["Strategy"] = [c[1] for c in classifications]
        
        # Summarize frequencies
        strategy_counts = []
        for task in ["Card", "Dice"]:
            task_label = f"{task} Problem"
            strategies = df.loc[df["Task"] == task, "Strategy"].value_counts(dropna=True)
            for strategy in ["MATCH", "MAX", "OTHER"]:
                count = strategies.get(strategy, 0)
                strategy_counts.append({
                    "Task": task_label,
                    "Strategy": strategy,
                    "Count": count
                })
        
        counts_df = pd.DataFrame(strategy_counts)
        
        # Confidence intervals for MAX proportion
        ci_results = []
        for task in ["Card", "Dice"]:
            task_label = f"{task} Problem"
            task_data = df[df["Task"] == task]["Strategy"]
            n_total = len(task_data)
            n_max = (task_data == "MAX").sum()
            
            if n_total > 0:
                prop = n_max / n_total
                ci_low, ci_high = proportion_confint(count=n_max, nobs=n_total, alpha=0.05, method="wilson")
                ci_results.append({
                    "Task": task_label,
                    "Test": "MAX % with 95% CI",
                    "Proportion": round(prop, 3),
                    "95% CI Lower": round(ci_low, 3),
                    "95% CI Upper": round(ci_high, 3)
                })
        
        ci_df = pd.DataFrame(ci_results) if ci_results else pd.DataFrame()
        
        results.append((f"{wave_label} Analysis", counts_df))
        if not ci_df.empty:
            results.append((f"{wave_label} - Confidence Intervals", ci_df))
        
        return results
    
    def _classify_card_responses(self, responses):
        """Classify card responses as MATCH, MAX, or OTHER."""
        if len(responses) < 10 or any(pd.isna(x) for x in responses):
            return np.nan
        n1 = sum(str(x).strip() == "1" for x in responses)
        n2 = sum(str(x).strip() == "2" for x in responses)
        if n1 == 10:
            return "MAX"
        elif n1 == 7 and n2 == 3:
            return "MATCH"
        else:
            return "OTHER"
    
    def _classify_dice_responses(self, responses):
        """Classify dice responses as MATCH, MAX, or OTHER."""
        if len(responses) < 6 or any(pd.isna(x) for x in responses):
            return np.nan
        n_red = sum(str(x).strip().lower() == "red" for x in responses)
        n_green = sum(str(x).strip().lower() == "green" for x in responses)
        if n_red == 6:
            return "MAX"
        elif n_red == 4 and n_green == 2:
            return "MATCH"
        else:
            return "OTHER"
    
    def _get_raw_distributions(self, data: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
        """Get raw response distributions."""
        dist_list = []
        
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            # Classify to get Task assignment
            classifications = []
            for idx, row in df.iterrows():
                card_answers = [row.get(c, np.nan) for c in self.card_cols]
                dice_answers = [row.get(c, np.nan) for c in self.dice_cols]
                n_card = sum(not pd.isna(x) for x in card_answers)
                n_dice = sum(not pd.isna(x) for x in dice_answers)
                
                if n_card > 0 and n_dice > 0:
                    classifications.append("BOTH")
                elif n_card == 10:
                    classifications.append("Card")
                elif n_dice == 6:
                    classifications.append("Dice")
                else:
                    classifications.append("None")
            
            df["Task"] = classifications
            
            # Card responses
            card_subset = df[df["Task"] == "Card"]
            if not card_subset.empty:
                card_raw = card_subset[self.card_cols].stack().dropna().str.strip().str.lower()
                card_counts = card_raw.value_counts()
                for val, count in card_counts.items():
                    dist_list.append({
                        "Wave": wave_label,
                        "Task": "Card",
                        "Raw Response": val,
                        "Count": count
                    })
            
            # Dice responses
            dice_subset = df[df["Task"] == "Dice"]
            if not dice_subset.empty:
                dice_raw = dice_subset[self.dice_cols].stack().dropna().str.strip().str.lower()
                dice_counts = dice_raw.value_counts()
                for val, count in dice_counts.items():
                    dist_list.append({
                        "Wave": wave_label,
                        "Task": "Dice",
                        "Raw Response": val,
                        "Count": count
                    })
        
        dist_df = pd.DataFrame(dist_list) if dist_list else pd.DataFrame()
        return [("Response Distributions", dist_df)] if not dist_df.empty else []

class DenominatorNeglectAnalysis(BaseAnalysis):
    """Denominator neglect analysis (Stanovich & West, 2008)."""
    
    def __init__(self, excel_writer: ExcelWriter):
        super().__init__(excel_writer)
        self.sheet_name = "W3-DenomNeglect"
        
    def run(self, data: Dict[str, pd.DataFrame]) -> None:
        """Run denominator neglect analysis."""
        results = []
        
        header_note = (
            "Analysis: Denominator Neglect\n\n"
            "Participants chose between:\n"
            "  - A small tray (1 black, 9 white => 10% chance of winning $2),\n"
            "  - A large tray (8 black, 92 white => 8% chance).\n\n"
            "The small tray is the normative choice, but following Stanovich & West \n"
            "we note a sizable minority still picks the large tray."
        )
        
        # Analyze each wave
        for wave_label, df in [("Wave 3", data["wave3"]), ("LLM", data["LLM"]), ("Wave 4", data["wave4"])]:
            wave_results = self._analyze_wave(df.copy(), wave_label)
            results.extend(wave_results)
        
        sheet_name = self.excel_writer.get_unique_sheet_name(self.sheet_name)
        self.excel_writer.write_results(sheet_name, results, header_note)
        
    def _analyze_wave(self, df: pd.DataFrame, wave_label: str) -> List[Tuple[str, pd.DataFrame]]:
        """Analyze denominator neglect for a single wave."""
        results = []
        
        # Map responses
        df["DenomNeg_tray"] = df["Denominator neglect"].apply(self._classify_tray)
        
        # Frequency table
        freq_df = df["DenomNeg_tray"].value_counts(dropna=True).sort_index().reset_index()
        freq_df.columns = ["Response", "Count"]
        freq_df["Percent"] = (freq_df["Count"] / freq_df["Count"].sum() * 100).round(1)
        
        # Confidence interval for large tray
        n_total = df["DenomNeg_tray"].notna().sum()
        n_large = (df["DenomNeg_tray"] == "Large tray").sum()
        
        if n_total > 0:
            prop_large = n_large / n_total
            ci_low, ci_high = proportion_confint(count=n_large, nobs=n_total, alpha=0.05, method="wilson")
            
            summary_df = pd.DataFrame({
                "Measure": ["Proportion choosing Large Tray (%)", "95% CI Lower (%)", "95% CI Upper (%)"],
                "Value": [round(prop_large * 100, 2), round(ci_low * 100, 2), round(ci_high * 100, 2)]
            })
        else:
            summary_df = pd.DataFrame()
        
        results.append((f"{wave_label} - Frequency Table", freq_df))
        if not summary_df.empty:
            results.append((f"{wave_label} - Large Tray Choice Summary", summary_df))
        
        return results
    
    def _classify_tray(self, resp):
        """Classify tray choice."""
        if pd.isna(resp):
            return np.nan
        text = str(resp).lower().strip()
        if "small" in text:
            return "Small tray"
        elif "large" in text:
            return "Large tray"
        else:
            return np.nan

# ============================================================================
# Analysis Runner
# ============================================================================

class AnalysisRunner:
    """Orchestrates running all analyses."""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.data_loader = DataLoader(output_dir)
        
        # Set up output paths
        excel_output_dir = os.path.join(output_dir, "accuracy_evaluation")
        os.makedirs(excel_output_dir, exist_ok=True)
        self.output_filename = os.path.join(excel_output_dir, "within_subject_analysis.xlsx")
        
        self.excel_writer = ExcelWriter(self.output_filename)
        
        # Initialize all analyses
        self.analyses = [
            BaseRateAnalysis(self.excel_writer),
            OutcomeBiasAnalysis(self.excel_writer),
            FalseConsensusAnalysis(self.excel_writer),
            SunkCostAnalysis(self.excel_writer),
            AllaisProblemAnalysis(self.excel_writer),
            NonseparabilityAnalysis(self.excel_writer),
            FramingAnalysis(self.excel_writer),
            LindaProblemAnalysis(self.excel_writer),
            AnchoringAnalysis(self.excel_writer),
            RelativeSavingsAnalysis(self.excel_writer),
            MysideBiasAnalysis(self.excel_writer),
            OmissionBiasAnalysis(self.excel_writer),
            LessIsMoreAnalysis(self.excel_writer),
            ThalerProblemAnalysis(self.excel_writer),
            ProbabilityMatchingAnalysis(self.excel_writer),
            DenominatorNeglectAnalysis(self.excel_writer),
        ]
    
    def run(self):
        """Run all analyses."""
        # Delete existing output file if it exists to prevent duplicate sheets
        if os.path.exists(self.output_filename):
            os.remove(self.output_filename)
            # Removed existing file
        
        # Load data
        # Loading data
        common_ids = self.data_loader.get_common_ids()
        
        data = {
            "wave1": self.data_loader.load_wave_data("wave1", common_ids),
            "wave2": self.data_loader.load_wave_data("wave2", common_ids),
            "wave3": self.data_loader.load_wave_data("wave3", common_ids),
            "wave4": self.data_loader.load_wave_data("wave4", common_ids),
            "LLM": self.data_loader.load_wave_data("LLM", common_ids)
        }
        
        # Run each analysis
        # Running analyses
        for analysis in self.analyses:
            # Running {analysis.__class__.__name__}
            try:
                analysis.run(data)
            except Exception as e:
                # Log error but continue
                import logging
                logging.error(f"Error in {analysis.__class__.__name__}: {e}")
                continue
        
        # Analysis complete

# ============================================================================
# Main Entry Point
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Run within- and between-subjects analysis.")
    parser.add_argument("--config", help="Path to YAML configuration file")
    parser.add_argument("--output_dir", help="Directory containing the simulation trial outputs and where the analysis Excel file will be saved.")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose output")
    args = parser.parse_args()
    
    # Handle config-based execution
    if args.config:
        import yaml
        with open(args.config, 'r') as f:
            config = yaml.safe_load(f)
        
        # Extract output directory from config
        trial_dir = config.get('trial_dir', '')
        output_dir = trial_dir if trial_dir else None
        
        if not output_dir:
            parser.error("trial_dir must be specified in config file")
        
        runner = AnalysisRunner(output_dir)
        runner.run()
    
    # Handle CLI-based execution
    elif args.output_dir:
        runner = AnalysisRunner(args.output_dir)
        runner.run()
    
    else:
        parser.error("Either --config or --output_dir must be provided")

if __name__ == "__main__":
    main() 