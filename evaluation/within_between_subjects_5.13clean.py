#!/usr/bin/env python
# coding: utf-8

# # Within- and Between-Subjects Tests

# This notebook creates an excel file, by default named `experiment_analysis.xlsx` (so it needs to be renamed for clarity after generating). Each code block adds a tab to this excel file with another analysis from one of the sub-studies across the four waves.
# 
# The tabs are labeled with the wave number first, followed by an indicator of the sub-study (e.g. `W1-Base Rate Neglect`, or `W2-Framing`).
# 
# Version 05/01/25: add MAD. 

# # Wave 1

# ## Base rate problem (Kahneman & Tversky, 1973)
# 

# In[73]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_1samp
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import os

# output_filename = "experiment_analysis basic_all 051325 json test.xlsx"

# ---------------------------
# Command line argument parsing
# ---------------------------
parser = argparse.ArgumentParser(description="Run within- and between-subjects analysis.")
parser.add_argument("--output_dir", required=True, help="Directory containing the simulation trial outputs and where the analysis Excel file will be saved.")
args = parser.parse_args()

output_dir = args.output_dir

# ---------------------------
# Construct output Excel filename
# ---------------------------
excel_output_dir = os.path.join(output_dir, "accuracy_evaluation")
os.makedirs(excel_output_dir, exist_ok=True)
output_filename = os.path.join(excel_output_dir, "within_subject_analysis.xlsx")

# ---------------------------
# Helper: Filter valid rows
# ---------------------------
def valid_rows(df):
    condition1 = (df["PROLIFIC_PID"].astype(str).str.strip() != "")
    condition2 = (df["Finished"].astype(str).str.strip().str.upper() == "TRUE")
    return df[condition1 & condition2].copy()

# ---------------------------
# File names for the four waves (raw labels)
# ---------------------------
        # "wave1": "raw_data/Wave 1/wave 1 final labels 042525.csv",
        # "wave2": "raw_data/Wave 2/Wave 2 final labels 042525.csv",
        # "wave3": "raw_data/Wave 3/wave 3 final labels 042825.csv",
        # "wave4": "raw_data/Wave 4/wave 4 complete data labels.csv"
label_files = {
    "wave1": "raw_data/Wave 1/wave 1 final labels 042525.csv",
    "wave2": "raw_data/Wave 2/Wave 2 final labels 042525.csv",
    "wave3": "raw_data/Wave 3/wave 3 final labels 042825.csv",
    "wave4": "raw_data/Wave 4/wave 4 complete data labels.csv",
#    "LLM": "responses_llm_imputed_formatted_label.csv",
     "LLM": os.path.join(output_dir, "csv_comparison/csv_formatted_label/responses_llm_imputed_label_formatted.csv"),
}

# ---------------------------
# Determine respondents who completed all 4 waves
# ---------------------------
def get_common_ids(label_files):
    common_ids = None
    for wave, file in label_files.items():
        if wave == "LLM":
            continue  # Skip LLM data (uses TWIN_ID instead of PROLIFIC_PID)
        df = pd.read_csv(file, low_memory=False)
        df = valid_rows(df)
        ids = set(df["PROLIFIC_PID"].astype(str).str.strip())
        if common_ids is None:
            common_ids = ids
        else:
            common_ids = common_ids.intersection(ids)
    print(f"Number of respondents who finished all 4 waves: {len(common_ids)}")
    return common_ids


common_ids = get_common_ids(label_files)

# ---------------------------
# Load wave 1 data and filter to common respondents
# ---------------------------
wave1_df = pd.read_csv(label_files["wave1"], low_memory=False)
wave1_df = valid_rows(wave1_df)
wave1_df = wave1_df[ wave1_df["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]
df_wave1 = wave1_df
df_wave1.columns = [col.strip() for col in df_wave1.columns]


# ---------------------------
# Load Wave 2 data
# ---------------------------
df_wave2 = pd.read_csv(label_files["wave2"], low_memory=False)
df_wave2 = valid_rows(df_wave2)
df_wave2 = df_wave2[df_wave2["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]
df_wave2["PROLIFIC_PID"] = df_wave2["PROLIFIC_PID"].astype(str).str.strip()

# ---------------------------
# Load Wave 3 data and filter to common respondents
# ---------------------------
df_wave3 = pd.read_csv(label_files["wave3"], low_memory=False)
df_wave3 = valid_rows(df_wave3)
df_wave3 = df_wave3[df_wave3["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]
df_wave3["PROLIFIC_PID"] = df_wave3["PROLIFIC_PID"].astype(str).str.strip()
df_wave3.columns = [col.strip() for col in df_wave3.columns]

# Load wave 4 data and filter to common respondents
wave4_df = pd.read_csv(label_files["wave4"], low_memory=False)
wave4_df = valid_rows(wave4_df)
wave4_df = wave4_df[ wave4_df["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]

# ---------------------------
# Load Wave 4 (other file names)
# ---------------------------
df_wave4 = pd.read_csv(label_files["wave4"], low_memory=False)
df_wave4 = valid_rows(df_wave4)
df_wave4 = df_wave4[df_wave4["PROLIFIC_PID"].astype(str).str.strip().isin(common_ids)]
df_wave4["PROLIFIC_PID"] = df_wave4["PROLIFIC_PID"].astype(str).str.strip()
df_wave4.columns = [col.strip() for col in df_wave4.columns]


# Load and filter LLM data
df_llm = pd.read_csv(label_files["LLM"], low_memory=False)
df_llm = df_llm[df_llm["TWIN_ID"].astype(str).str.strip() != ""]
df_llm["TWIN_ID"] = df_llm["TWIN_ID"].astype(str).str.strip()
df_llm = df_llm[df_llm["TWIN_ID"].str.isnumeric()]  # Keep only rows with numeric TWIN_IDs
df_llm.columns = [col.strip() for col in df_llm.columns]


# ---------------------------
# Define the Base Rate Neglect columns (between-subjects)
# ---------------------------
# Each subject sees only one of these questions.
base_rate_cols = [
    "Form B - 70 eng _1",  # Normative value: 30%
    "Q156_1"              # Normative value: 70%
]

# Extract the two base rate columns along with PROLIFIC_PID
base_rate_df = wave1_df[["PROLIFIC_PID"] + base_rate_cols].copy()

# ---------------------------
# Convert responses to numeric (expected scale: 0-100)
# ---------------------------
for col in base_rate_cols:
    base_rate_df[col] = pd.to_numeric(base_rate_df[col], errors='coerce')

# ---------------------------
# Compute descriptive statistics for each question
# ---------------------------
desc_stats_list = []
for col in base_rate_cols:
    data = base_rate_df[col].dropna()
    desc_stats_list.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df = pd.DataFrame(desc_stats_list)


# ---------------------------
# Two-sample t-test between the two base rate conditions
# ---------------------------
from scipy.stats import ttest_ind

data1 = base_rate_df[base_rate_cols[0]].dropna()
data2 = base_rate_df[base_rate_cols[1]].dropna()

t_stat, p_val = ttest_ind(data1, data2, equal_var=False)  # Welch's t-test

ttest_results_df = pd.DataFrame([{
    "Comparison": f"{base_rate_cols[0]} vs. {base_rate_cols[1]}",
    "Mean 1": round(data1.mean(), 2),
    "Mean 2": round(data2.mean(), 2),
    "t-statistic": round(t_stat, 3),
    "p-value": round(p_val, 3),
    "N1": len(data1),
    "N2": len(data2)
}])

# ---------------------------
# header note explanation
# ---------------------------
header_note = (
    "Analysis: Two-sample t-test comparing participants' responses across the two base rate conditions. \n"
    "Each participant responded to one question only. This test evaluates whether the average responses differ significantly."
)

# ---------------------------
# unique new sheet name
# ---------------------------
base_sheet_name = "W1-Base Rate Neglect"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name


# ---------------------------
# Write everything to the new sheet
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    # Write descriptive statistics starting at row 4
    startrow = 3
    desc_stats_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    
    # Get the worksheet object
    ws = writer.sheets[sheet_name]
    # Write header note in cell A1
    ws.cell(row=1, column=1, value=header_note)
    
    # Determine the row at which to write the t-test results (leaving a gap)
    ttest_startrow = startrow + len(desc_stats_df) + 3
    
    # Write the t-test results into the worksheet using openpyxl's dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df, index=False, header=True), ttest_startrow):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")



# ---------------------------
# Now repeat the same operations for wave 4
# ---------------------------


#print(wave4_df.columns.tolist())

# Each subject sees only one of these questions.
base_rate_cols_w4 = [
    "Form A _1",  # Normative value: 30%
    "Q156_1"              # Normative value: 70%
]

# Extract the two base rate columns along with PROLIFIC_PID
base_rate_df_w4 = wave4_df[["PROLIFIC_PID"] + base_rate_cols_w4].copy()

# Convert responses to numeric
for col in base_rate_cols_w4:
    base_rate_df_w4[col] = pd.to_numeric(base_rate_df_w4[col], errors='coerce')

# Compute descriptive statistics for each question in wave 4
desc_stats_list_w4 = []
for col in base_rate_cols_w4:
    data = base_rate_df_w4[col].dropna()
    desc_stats_list_w4.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_wave4_df = pd.DataFrame(desc_stats_list_w4)

# Two-sample t-test between the two base rate conditions in wave 4
data1_w4 = base_rate_df_w4[base_rate_cols_w4[0]].dropna()
data2_w4 = base_rate_df_w4[base_rate_cols_w4[1]].dropna()

t_stat_w4, p_val_w4 = ttest_ind(data1_w4, data2_w4, equal_var=False)  # Welch's t-test

ttest_results_wave4_df = pd.DataFrame([{
    "Comparison": f"{base_rate_cols_w4[0]} vs. {base_rate_cols_w4[1]} (Wave 4)",
    "Mean 1": round(data1_w4.mean(), 2),
    "Mean 2": round(data2_w4.mean(), 2),
    "t-statistic": round(t_stat_w4, 3),
    "p-value": round(p_val_w4, 3),
    "N1": len(data1_w4),
    "N2": len(data2_w4)
}])


# ---------------------------
# Now repeat the same operations for LLM data (appended to same sheet)
# ---------------------------

# Define columns for LLM analysis
base_rate_cols_llm = ["Form A _1", "Q156_1"]
base_rate_df_llm = df_llm[["TWIN_ID"] + base_rate_cols_llm].copy()

# Convert to numeric
for col in base_rate_cols_llm:
    base_rate_df_llm[col] = pd.to_numeric(base_rate_df_llm[col], errors='coerce')

# Descriptive statistics
desc_stats_list_llm = []
for col in base_rate_cols_llm:
    data = base_rate_df_llm[col].dropna()
    desc_stats_list_llm.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_llm_df = pd.DataFrame(desc_stats_list_llm)

# Two-sample t-test
data1_llm = base_rate_df_llm[base_rate_cols_llm[0]].dropna()
data2_llm = base_rate_df_llm[base_rate_cols_llm[1]].dropna()

t_stat_llm, p_val_llm = ttest_ind(data1_llm, data2_llm, equal_var=False)

ttest_results_llm_df = pd.DataFrame([{
    "Comparison": f"{base_rate_cols_llm[0]} vs. {base_rate_cols_llm[1]} (LLM)",
    "Mean 1": round(data1_llm.mean(), 2),
    "Mean 2": round(data2_llm.mean(), 2),
    "t-statistic": round(t_stat_llm, 3),
    "p-value": round(p_val_llm, 3),
    "N1": len(data1_llm),
    "N2": len(data2_llm)
}])

# Append LLM results to same sheet as Wave 1 & 4
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]

    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value="LLM Results:")
    
    # Find next available row after wave 4 + correlation
    next_row = ws.max_row + 2
    
    # Write LLM Descriptive Stats
    desc_stats_llm_df.to_excel(writer, sheet_name=sheet_name, startrow=next_row, index=False)
    
    # Write t-test results below descriptive stats
    ttest_startrow_llm = next_row + len(desc_stats_llm_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_llm_df, index=False, header=True), ttest_startrow_llm):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"LLM results appended to '{sheet_name}' tab in '{output_filename}'.")


# ---------------------------
# Append everything into the same Excel tab
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    # Find the next empty row
    current_max_row = ws.max_row+2
    
    ws.cell(row=current_max_row, column=1, value="Wave 4 Results:")
    current_max_row = ws.max_row + 1

    
    # Write Wave 4 Descriptive Stats
    desc_stats_wave4_df.to_excel(writer, sheet_name=sheet_name, startrow=current_max_row + 2, index=False)
    
    # Write Wave 4 T-Test Results
    ttest_startrow_w4 = current_max_row + 2 + len(desc_stats_wave4_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_wave4_df, index=False, header=True), ttest_startrow_w4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
print(f"Wave 4 results appended to '{sheet_name}' tab in '{output_filename}'.")

print("done")


# ## Outcome bias (Baron & Hershey, 1988)

# In[74]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------
# Define Outcome Bias columns (between-subjects)
# ---------------------------
# Q161: Positive outcome scenario
# Q162: Negative outcome scenario
outcome_cols = ["Q161", "Q162"]

# Extract the two outcome columns along with PROLIFIC_PID
outcome_df = wave1_df[["PROLIFIC_PID"] + outcome_cols].copy()

# ---------------------------
# Map the text-based responses to numeric values (from +3 to -3)
# ---------------------------
outcome_mapping = {
    "Clearly correct, an excellent decision": 3,
    "Correct, all things considered": 2,
    "Correct, but the opposite would be reasonable too": 1,
    "The decision and its opposite are equally good": 0,
    "Incorrect, but not unreasonable": -1,
    "Incorrect, all things considered": -2,
    "Incorrect, a very bad decision": -3
}

# Create new columns with the numeric conversion (in-memory only)
outcome_df["Q161_conv"] = outcome_df["Q161"].str.strip().map(outcome_mapping)
outcome_df["Q162_conv"] = outcome_df["Q162"].str.strip().map(outcome_mapping)

# ---------------------------
# Compute descriptive statistics for each scenario (using converted values)
# ---------------------------
desc_stats_list = []
for col in ["Q161_conv", "Q162_conv"]:
    data = outcome_df[col].dropna()
    desc_stats_list.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df = pd.DataFrame(desc_stats_list)

# ---------------------------
# Compute Outcome Bias and perform independent samples t-test
# ---------------------------
# Group 1: Q161_conv (positive outcome)
# Group 2: Q162_conv (negative outcome)
group_pos = outcome_df["Q161_conv"].dropna()
group_neg = outcome_df["Q162_conv"].dropna()

# Outcome bias: positive group mean minus negative group mean
outcome_bias = group_pos.mean() - group_neg.mean()

# Independent samples t-test (using Welch's t-test)
t_stat, p_val = ttest_ind(group_pos, group_neg, equal_var=False)

ttest_results = {
    "Outcome Bias (Mean Difference)": round(outcome_bias, 3),
    "t-statistic": round(t_stat, 3),
    "p-value": round(p_val, 3)
}
ttest_results_df = pd.DataFrame([ttest_results])

# ---------------------------
# Prepare a header note explaining the analysis
# ---------------------------
header_note = (
    "Analysis: Outcome Bias (Baron & Hershey, 1988).\n"
    "Responses for Q161 and Q162 were originally text-based Likert ratings (ranging from -3 to +3).\n"
    "They were converted as follows:\n"
    "  'Clearly correct, an excellent decision' -> +3\n"
    "  'Correct, all things considered' -> +2\n"
    "  'Correct, but the opposite would be reasonable too' -> +1\n"
    "  'The decision and its opposite are equally good' -> 0\n"
    "  'Incorrect, but not unreasonable' -> -1\n"
    "  'Incorrect, all things considered' -> -2\n"
    "  'Incorrect, a very bad decision' -> -3\n\n"
    "Per the paper, we define outcome bias as the mean rating in the positive outcome condition (Q161_conv) minus\n"
    "the mean rating in the negative outcome condition (Q162_conv). A positive and statistically significant\n"
    "difference supports the presence of outcome bias."
)

# ---------------------------
# determine a unique new sheet name
# ---------------------------
##output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W1-Outcome Bias"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

# ---------------------------
# Write descriptive stats and t-test results to the new sheet
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:
    startrow = 3
    # Write descriptive statistics starting at row 4
    desc_stats_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    
    # Get the worksheet object
    ws = writer.sheets[sheet_name]
    # Write header note in cell A1
    ws.cell(row=1, column=1, value=header_note)
    
    # Determine the row at which to write the t-test results (leaving a gap)
    ttest_startrow = startrow + len(desc_stats_df) + 3
    # Write the t-test results into the worksheet using openpyxl's dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df, index=False, header=True), ttest_startrow):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


# ---------------------------
# Now process Wave 4
# ---------------------------

# Extract and map the outcome columns
outcome_df_w4 = wave4_df[["PROLIFIC_PID"] + outcome_cols].copy()
outcome_df_w4["Q161_conv"] = outcome_df_w4["Q161"].str.strip().map(outcome_mapping)
outcome_df_w4["Q162_conv"] = outcome_df_w4["Q162"].str.strip().map(outcome_mapping)

# ---------------------------
# Compute descriptive statistics for Wave 4
# ---------------------------
desc_stats_list_w4 = []
for col in ["Q161_conv", "Q162_conv"]:
    data = outcome_df_w4[col].dropna()
    desc_stats_list_w4.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df_w4 = pd.DataFrame(desc_stats_list_w4)

# ---------------------------
# T-test for Wave 4
# ---------------------------
group_pos_w4 = outcome_df_w4["Q161_conv"].dropna()
group_neg_w4 = outcome_df_w4["Q162_conv"].dropna()

outcome_bias_w4 = group_pos_w4.mean() - group_neg_w4.mean()
t_stat_w4, p_val_w4 = ttest_ind(group_pos_w4, group_neg_w4, equal_var=False)

ttest_results_w4 = {
    "Outcome Bias (Mean Difference)": round(outcome_bias_w4, 3),
    "t-statistic": round(t_stat_w4, 3),
    "p-value": round(p_val_w4, 3)
}
ttest_results_df_w4 = pd.DataFrame([ttest_results_w4])


# ---------------------------
# Now process LLM
# ---------------------------

# Extract and map the outcome columns
outcome_df_llm = df_llm[["TWIN_ID"] + outcome_cols].copy()
outcome_df_llm["Q161_conv"] = outcome_df_llm["Q161"].str.strip().map(outcome_mapping)
outcome_df_llm["Q162_conv"] = outcome_df_llm["Q162"].str.strip().map(outcome_mapping)

# ---------------------------
# Compute descriptive statistics for LLM
# ---------------------------
desc_stats_list_llm = []
for col in ["Q161_conv", "Q162_conv"]:
    data = outcome_df_llm[col].dropna()
    desc_stats_list_llm.append({
        "Question": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df_llm = pd.DataFrame(desc_stats_list_llm)

# ---------------------------
# T-test for LLM
# ---------------------------
group_pos_llm = outcome_df_llm["Q161_conv"].dropna()
group_neg_llm = outcome_df_llm["Q162_conv"].dropna()

outcome_bias_llm = group_pos_llm.mean() - group_neg_llm.mean()
t_stat_llm, p_val_llm = ttest_ind(group_pos_llm, group_neg_llm, equal_var=False)

ttest_results_llm = {
    "Outcome Bias (Mean Difference)": round(outcome_bias_llm, 3),
    "t-statistic": round(t_stat_llm, 3),
    "p-value": round(p_val_llm, 3)
}
ttest_results_df_llm = pd.DataFrame([ttest_results_llm])


# Append LLM results to same sheet as Wave 1 & 4
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]

    next_row = ws.max_row + 2
    ws.cell(row=next_row, column=1, value="LLM Results:")
    
    # Find next available row after wave 4 + correlation
    next_row = ws.max_row + 2
    
    # Write LLM Descriptive Stats
    desc_stats_df_llm.to_excel(writer, sheet_name=sheet_name, startrow=next_row, index=False)
    
    # Write t-test results below descriptive stats
    ttest_startrow_llm = next_row + len(desc_stats_df_llm) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df_llm, index=False, header=True), start=ttest_startrow_llm):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


print(f"LLM results appended to '{sheet_name}' tab in '{output_filename}'.")

# ---------------------------
# Append everything to the same Excel sheet, cleanly
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    # Find the next empty row
    current_max_row = ws.max_row
    
    # Write Section Title: Wave 4 Summary
    ws.cell(row=current_max_row + 2, column=1, value="Wave 4 - Outcome Bias Summary")
    
    # Write Wave 4 Descriptive Statistics
    desc_stats_startrow_w4 = current_max_row + 4
    desc_stats_df_w4.to_excel(writer, sheet_name=sheet_name, startrow=desc_stats_startrow_w4, index=False)
    
    # Write Wave 4 T-Test Results
    ttest_startrow_w4 = desc_stats_startrow_w4 + len(desc_stats_df_w4) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df_w4, index=False, header=True), ttest_startrow_w4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
  
print(f"Wave 4 descriptive stats and t-test results appended to '{sheet_name}' tab in '{output_filename}'.")

# ---------------------------
# Write distribution of CONVERTED responses (Q161_conv and Q162_conv) 
# ---------------------------

# Build distribution tables from converted columns
def build_converted_distribution(df, wave_label):
    dist_list = []
    for q in ["Q161_conv", "Q162_conv"]:
        dist = df[q].value_counts(dropna=False).sort_index()
        dist_df = pd.DataFrame({
            "Wave": wave_label,
            "Question": q,
            "Converted Response": dist.index,
            "Count": dist.values
        })
        dist_list.append(dist_df)
    return pd.concat(dist_list, ignore_index=True)

dist_conv_w1 = build_converted_distribution(outcome_df, "Wave 1")
dist_conv_llm = build_converted_distribution(outcome_df_llm, "LLM")
dist_conv_w4 = build_converted_distribution(outcome_df_w4, "Wave 4")

converted_dist_df = pd.concat([dist_conv_w1,dist_conv_llm, dist_conv_w4], ignore_index=True)

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    # Determine where to write (leave space after correlation block)
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="Converted Response Distributions for Q161 and Q162")

    raw_startrow = next_row + 2
    converted_dist_df.to_excel(writer, sheet_name=sheet_name, startrow=raw_startrow, index=False)

print("Converted response distributions for Q161 and Q162 written to Excel.")


print("done")


# ## Sunk cost fallacy (Stanovich & West, 2008)

# In[75]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import math


# ---------------------------
# Define the Sunk Cost Fallacy columns (between-subjects)
# ---------------------------
# Q181: No-sunk-cost condition
# Q182: Sunk-cost condition
sunk_cost_cols = ["Q181", "Q182"]

# Extract the two columns with PROLIFIC_PID
sunk_cost_df = wave1_df[["PROLIFIC_PID"] + sunk_cost_cols].copy()

    
def map_to_ten_point(x):
    try:
        v = int(float(x))
    except:
        return np.nan
    if 0 <= v <= 20:
        return v
    else:
        return np.nan


# ---------------------------
# Create new columns with the mapped responses (in-memory conversion only)
# ---------------------------
sunk_cost_df["Q181_conv"] = sunk_cost_df["Q181"].apply(map_to_ten_point)
sunk_cost_df["Q182_conv"] = sunk_cost_df["Q182"].apply(map_to_ten_point)


# ---------------------------
# Compute descriptive statistics for each condition (using converted responses)
# ---------------------------
desc_stats_list = []
for col in ["Q181_conv", "Q182_conv"]:
    data = sunk_cost_df[col].dropna()
    desc_stats_list.append({
        "Condition": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df = pd.DataFrame(desc_stats_list)

# ---------------------------
# Perform independent samples t-test between the two conditions
# ---------------------------
group_no_sunk = sunk_cost_df["Q181_conv"].dropna()  # no-sunk-cost condition
group_sunk = sunk_cost_df["Q182_conv"].dropna()       # sunk-cost condition

t_stat, p_val = ttest_ind(group_no_sunk, group_sunk, equal_var=False)

# Compute pooled standard deviation for Cohen's d
n1, n2 = group_no_sunk.shape[0], group_sunk.shape[0]
s1, s2 = group_no_sunk.std(), group_sunk.std()
pooled_sd = math.sqrt(((n1 - 1) * s1**2 + (n2 - 1) * s2**2) / (n1 + n2 - 2))
cohens_d = (group_no_sunk.mean() - group_sunk.mean()) / pooled_sd

ttest_results = {
    "No-sunk-cost Mean": round(group_no_sunk.mean(), 2),
    "Sunk-cost Mean": round(group_sunk.mean(), 2),
    "t-statistic": round(t_stat, 3),
    "p-value": round(p_val, 3),
    "Cohen's d": round(cohens_d, 3)
}
ttest_results_df = pd.DataFrame([ttest_results])

# ---------------------------
# Prepare a header note explaining the analysis
# ---------------------------
header_note = (
    "Analysis: Sunk Cost Fallacy.\n"
    "Participants in the no-sunk-cost condition (Q181) and sunk-cost condition (Q182) provided\n"
    "responses (raw values out of 20) that were, per Stanovich & West (2008), mapped to a 10-point scale as follows:\n"
    "  0–1 -> 1, 2–3 -> 2, 4–5 -> 3, 6–7 -> 4, 8–9 -> 5, 10–11 -> 6, 11–12 -> 7, 13–14 -> 8,\n"
    "  15–16 -> 9, 17–20 -> 10.\n\n"
    "We expect that participants in the no-sunk-cost condition will show a higher score\n"
    "(indicating a stronger preference for convenience) than those in the sunk-cost condition.\n"
    "A significant independent samples t-test would support the presence\n"
    "of a sunk cost effect."
)

# ---------------------------
# Determine a unique new sheet name
# ---------------------------
##output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W1-Sunk Cost"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

# ---------------------------
# Write all results to the new sheet at once
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
    startrow = 3
    # Write descriptive statistics table
    desc_stats_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    
    # Get worksheet object and add header note
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)
    
    # Write the t-test results below the descriptive statistics
    ttest_startrow = startrow + len(desc_stats_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df, index=False, header=True), ttest_startrow):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


################modifications;
# ---------------------------
# REMOVE the mapping and just use raw numeric values directly
# ---------------------------
for col in sunk_cost_cols:
    sunk_cost_df[col] = pd.to_numeric(sunk_cost_df[col], errors='coerce')

# ---------------------------
# Compute descriptive statistics for Wave 1 (raw numbers now)
# ---------------------------
desc_stats_list = []
for col in sunk_cost_cols:
    data = sunk_cost_df[col].dropna()
    desc_stats_list.append({
        "Condition": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df = pd.DataFrame(desc_stats_list)

# ---------------------------
# Perform independent samples t-test between the two conditions (Wave 1)
# ---------------------------
group_no_sunk = sunk_cost_df["Q181"].dropna()
group_sunk = sunk_cost_df["Q182"].dropna()

t_stat, p_val = ttest_ind(group_no_sunk, group_sunk, equal_var=False)

# Compute pooled standard deviation for Cohen's d
n1, n2 = group_no_sunk.shape[0], group_sunk.shape[0]
s1, s2 = group_no_sunk.std(), group_sunk.std()
pooled_sd = math.sqrt(((n1 - 1) * s1**2 + (n2 - 1) * s2**2) / (n1 + n2 - 2))
cohens_d = (group_no_sunk.mean() - group_sunk.mean()) / pooled_sd

ttest_results = {
    "No-sunk-cost Mean": round(group_no_sunk.mean(), 2),
    "Sunk-cost Mean": round(group_sunk.mean(), 2),
    "t-statistic": round(t_stat, 3),
    "p-value": round(p_val, 3),
    "Cohen's d": round(cohens_d, 3)
}
ttest_results_df = pd.DataFrame([ttest_results])

# ---------------------------
# Write Wave 1 results first
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
    startrow = 3
    desc_stats_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)
    
    ttest_startrow = startrow + len(desc_stats_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df, index=False, header=True), ttest_startrow):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab for Wave 1 results.")

# ---------------------------
# Now process Wave 4 similarly
# ---------------------------


# Extract and convert the sunk cost columns
sunk_cost_df_w4 = wave4_df[["PROLIFIC_PID"] + sunk_cost_cols].copy()
for col in sunk_cost_cols:
    sunk_cost_df_w4[col] = pd.to_numeric(sunk_cost_df_w4[col], errors='coerce')

# Descriptive stats for Wave 4
desc_stats_list_w4 = []
for col in sunk_cost_cols:
    data = sunk_cost_df_w4[col].dropna()
    desc_stats_list_w4.append({
        "Condition": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df_w4 = pd.DataFrame(desc_stats_list_w4)

# T-test for Wave 4
group_no_sunk_w4 = sunk_cost_df_w4["Q181"].dropna()
group_sunk_w4 = sunk_cost_df_w4["Q182"].dropna()

t_stat_w4, p_val_w4 = ttest_ind(group_no_sunk_w4, group_sunk_w4, equal_var=False)

n1_w4, n2_w4 = group_no_sunk_w4.shape[0], group_sunk_w4.shape[0]
s1_w4, s2_w4 = group_no_sunk_w4.std(), group_sunk_w4.std()
pooled_sd_w4 = math.sqrt(((n1_w4 - 1) * s1_w4**2 + (n2_w4 - 1) * s2_w4**2) / (n1_w4 + n2_w4 - 2))
cohens_d_w4 = (group_no_sunk_w4.mean() - group_sunk_w4.mean()) / pooled_sd_w4

ttest_results_w4 = {
    "No-sunk-cost Mean (W4)": round(group_no_sunk_w4.mean(), 2),
    "Sunk-cost Mean (W4)": round(group_sunk_w4.mean(), 2),
    "t-statistic": round(t_stat_w4, 3),
    "p-value": round(p_val_w4, 3),
    "Cohen's d": round(cohens_d_w4, 3)
}
ttest_results_df_w4 = pd.DataFrame([ttest_results_w4])


# ---------------------------
# Now process LLM data similarly
# ---------------------------

# Extract and convert the sunk cost columns using TWIN_ID
sunk_cost_df_llm = df_llm[["TWIN_ID"] + sunk_cost_cols].copy()
for col in sunk_cost_cols:
    sunk_cost_df_llm[col] = pd.to_numeric(sunk_cost_df_llm[col], errors='coerce')

# Descriptive stats for LLM
desc_stats_list_llm = []
for col in sunk_cost_cols:
    data = sunk_cost_df_llm[col].dropna()
    desc_stats_list_llm.append({
        "Condition": col,
        "N": data.shape[0],
        "Mean": round(data.mean(), 2),
        "Median": round(data.median(), 2),
        "Std": round(data.std(), 2),
        "Min": data.min(),
        "Max": data.max()
    })
desc_stats_df_llm = pd.DataFrame(desc_stats_list_llm)

# T-test for LLM
group_no_sunk_llm = sunk_cost_df_llm["Q181"].dropna()
group_sunk_llm = sunk_cost_df_llm["Q182"].dropna()

t_stat_llm, p_val_llm = ttest_ind(group_no_sunk_llm, group_sunk_llm, equal_var=False)

n1_llm, n2_llm = group_no_sunk_llm.shape[0], group_sunk_llm.shape[0]
s1_llm, s2_llm = group_no_sunk_llm.std(), group_sunk_llm.std()
pooled_sd_llm = math.sqrt(((n1_llm - 1) * s1_llm**2 + (n2_llm - 1) * s2_llm**2) / (n1_llm + n2_llm - 2))
cohens_d_llm = (group_no_sunk_llm.mean() - group_sunk_llm.mean()) / pooled_sd_llm

ttest_results_llm = {
    "No-sunk-cost Mean (LLM)": round(group_no_sunk_llm.mean(), 2),
    "Sunk-cost Mean (LLM)": round(group_sunk_llm.mean(), 2),
    "t-statistic": round(t_stat_llm, 3),
    "p-value": round(p_val_llm, 3),
    "Cohen's d": round(cohens_d_llm, 3)
}
ttest_results_df_llm = pd.DataFrame([ttest_results_llm])

# ---------------------------
# Append LLM results to same sheet
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    current_max_row = ws.max_row

    # LLM Section Title
    ws.cell(row=current_max_row + 2, column=1, value="LLM - Sunk Cost Fallacy Summary")

    # LLM Descriptive Statistics
    desc_stats_startrow_llm = current_max_row + 4
    desc_stats_df_llm.to_excel(writer, sheet_name=sheet_name, startrow=desc_stats_startrow_llm, index=False)

    # LLM t-test
    ttest_startrow_llm = desc_stats_startrow_llm + len(desc_stats_df_llm) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df_llm, index=False, header=True), start=ttest_startrow_llm):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print("LLM Sunk Cost Fallacy results appended to Excel.")



# ---------------------------
# Append Wave 4 results into the same sheet
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    current_max_row = ws.max_row

    # Wave 4 Section Title
    ws.cell(row=current_max_row + 2, column=1, value="Wave 4 - Sunk Cost Fallacy Summary")

    # Wave 4 Descriptive Statistics
    desc_stats_startrow_w4 = current_max_row + 4
    desc_stats_df_w4.to_excel(writer, sheet_name=sheet_name, startrow=desc_stats_startrow_w4, index=False)

    # Wave 4 t-test
    ttest_startrow_w4 = desc_stats_startrow_w4 + len(desc_stats_df_w4) + 3
    for r_idx, row in enumerate(dataframe_to_rows(ttest_results_df_w4, index=False, header=True), ttest_startrow_w4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"Excel file '{output_filename}' updated with Wave 4 descriptive stats, t-test, and correlation results.")
print("done")


# ## Allais problem (Stanovich & West, 2008)

# In[76]:


import pandas as pd
import numpy as np
from scipy.stats import binomtest  # Using scipy's binomtest for the binomial test
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------
# Helper: Filter valid rows
# ---------------------------
def valid_rows(df):
    cond1 = (df["PROLIFIC_PID"].astype(str).str.strip() != "")
    cond2 = (df["Finished"].astype(str).str.strip().str.upper() == "TRUE")
    return df[cond1 & cond2].copy()

# ---------------------------
# Determine respondents who completed all 4 waves
# ---------------------------
def get_common_ids(label_files):
    common_ids = None
    for wave, file in label_files.items():
        df = pd.read_csv(file, low_memory=False)
        df = valid_rows(df)
        ids = set(df["PROLIFIC_PID"].astype(str).str.strip())
        if common_ids is None:
            common_ids = ids
        else:
            common_ids = common_ids.intersection(ids)
    print(f"Number of respondents who finished all 4 waves: {len(common_ids)}")
    return common_ids


# ---------------------------
# Define the Allais Problem questions
# ---------------------------
# Q192: Form 1 – Choice between:
#      Option A: One million dollars for sure
#      Option B: 89% chance of one million dollars, 10% chance of five million dollars, 1% chance of nothing
# Q193: Form 2 – Choice between:
#      Option C: 11% chance of one million dollars, 89% chance of nothing
#      Option D: 10% chance of five million dollars, 90% chance of nothing
allais_cols = ["Q192", "Q193"]

# Extract the columns along with PROLIFIC_PID
allais_df = wave1_df[["PROLIFIC_PID"] + allais_cols].copy()

# ---------------------------
# Define mapping functions for the two forms
# ---------------------------
def map_allais_form1(resp):
    if pd.isna(resp):
        return np.nan
    text = resp.lower()
    # If the response indicates a sure thing, map to Option A
    if "for sure" in text or "100" in text:
        return "A"
    # Otherwise, if it includes the three key probabilities, map to Option B
    elif "89%" in text and "10%" in text and "1%" in text:
        return "B"
    else:
        return "Unknown"

def map_allais_form2(resp):
    if pd.isna(resp):
        return np.nan
    text = resp.lower()
    # If it mentions 11% and 89%, map to Option C
    if "11%" in text and "89%" in text:
        return "C"
    # If it mentions 10% and 90%, map to Option D
    elif "10%" in text and "90%" in text:
        return "D"
    else:
        return "Unknown"

# Create new columns with the mapped responses (in-memory conversion only)
allais_df["Q192_mapped"] = allais_df["Q192"].apply(map_allais_form1)
allais_df["Q193_mapped"] = allais_df["Q193"].apply(map_allais_form2)

# ---------------------------
# Compute frequency tables for each question
# ---------------------------
def freq_table(series):
    # Get counts sorted by option alphabetically
    counts = series.value_counts(dropna=True).sort_index()
    percents = series.value_counts(normalize=True, dropna=True).sort_index() * 100
    return pd.DataFrame({
        "Option": counts.index,
        "Count": counts.values,
        "Percent": percents.round(1).values
    })

freq_Q192 = freq_table(allais_df["Q192_mapped"])
freq_Q193 = freq_table(allais_df["Q193_mapped"])

freq_Q192["Form"] = "Form 1 (Q192)"
freq_Q193["Form"] = "Form 2 (Q193)"

# Combine the frequency tables for reporting
freq_allais_df = pd.concat([freq_Q192, freq_Q193], ignore_index=True)

# ---------------------------
# Perform binomial tests:
# For Form 1: Test if the proportion choosing Option A is > 0.5.
# For Form 2: Test if the proportion choosing Option D is > 0.5.
# (Exclude responses labeled "Unknown" if any.)
form1_data = allais_df["Q192_mapped"].dropna()
n_form1 = form1_data.shape[0]
n_A = (form1_data == "A").sum()
result_form1 = binomtest(n_A, n_form1, p=0.5, alternative='greater')
p_val_form1 = result_form1.pvalue

form2_data = allais_df["Q193_mapped"].dropna()
n_form2 = form2_data.shape[0]
n_D = (form2_data == "D").sum()
result_form2 = binomtest(n_D, n_form2, p=0.5, alternative='greater')
p_val_form2 = result_form2.pvalue

binom_results_df = pd.DataFrame({
    "Form": ["Form 1 (Q192)", "Form 2 (Q193)"],
    "n_total": [n_form1, n_form2],
    "n_target": [n_A, n_D],
    "Proportion_target": [round(n_A/n_form1, 3) if n_form1 > 0 else np.nan, 
                           round(n_D/n_form2, 3) if n_form2 > 0 else np.nan],
    "p-value": [round(p_val_form1, 3), round(p_val_form2, 3)]
})

# ---------------------------
# Prepare a header note explaining the analysis
# ---------------------------
header_note = (
    "Analysis: Allais Problem.\n\n"
    "Participants were randomly assigned to one of two forms of the Allais problem.\n"
    "Form 1 (Q192) presents a choice between:\n"
    "  Option A: One million dollars for sure\n"
    "  Option B: 89% chance of one million dollars, 10% chance of five million dollars, 1% chance of nothing\n\n"
    "Form 2 (Q193) presents a choice between:\n"
    "  Option C: 11% chance of one million dollars, 89% chance of nothing\n"
    "  Option D: 10% chance of five million dollars, 90% chance of nothing\n\n"
    "Responses were mapped to options A, B for Form 1 and C, D for Form 2.\n"
    "For completeness, we perform binomial tests for each form against a null of 0.5. \n"
    "Stanovich & West expect most participants to choose Option A in Form 1 and Option D in Form 2."
)

# ---------------------------
# Determine a unique new sheet name
# ---------------------------
base_sheet_name = "W1-Allais Problem"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

# ---------------------------
# Write frequency tables and binomial test results to the new sheet in one go
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    # Write the combined frequency table starting at row 4
    freq_allais_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)
    
    # Write the binomial test results below the frequency table (leaving a gap)
    binom_startrow = startrow + len(freq_allais_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(binom_results_df, index=False, header=True), binom_startrow):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


# ---------------------------
# Now process Wave 4 similarly
# ---------------------------

# Extract Allais columns
allais_df_w4 = wave4_df[["PROLIFIC_PID"] + allais_cols].copy()

# Apply mappings
allais_df_w4["Q192_mapped"] = allais_df_w4["Q192"].apply(map_allais_form1)
allais_df_w4["Q193_mapped"] = allais_df_w4["Q193"].apply(map_allais_form2)

# Frequency tables for Wave 4
freq_Q192_w4 = freq_table(allais_df_w4["Q192_mapped"])
freq_Q193_w4 = freq_table(allais_df_w4["Q193_mapped"])

freq_Q192_w4["Form"] = "Form 1 (Q192) - Wave 4"
freq_Q193_w4["Form"] = "Form 2 (Q193) - Wave 4"

freq_allais_df_w4 = pd.concat([freq_Q192_w4, freq_Q193_w4], ignore_index=True)

# Binomial tests for Wave 4
form1_data_w4 = allais_df_w4["Q192_mapped"].dropna()
n_form1_w4 = form1_data_w4.shape[0]
n_A_w4 = (form1_data_w4 == "A").sum()
result_form1_w4 = binomtest(n_A_w4, n_form1_w4, p=0.5, alternative='greater')
p_val_form1_w4 = result_form1_w4.pvalue

form2_data_w4 = allais_df_w4["Q193_mapped"].dropna()
n_form2_w4 = form2_data_w4.shape[0]
n_D_w4 = (form2_data_w4 == "D").sum()
result_form2_w4 = binomtest(n_D_w4, n_form2_w4, p=0.5, alternative='greater')
p_val_form2_w4 = result_form2_w4.pvalue

binom_results_df_w4 = pd.DataFrame({
    "Form": ["Form 1 (Q192) - Wave 4", "Form 2 (Q193) - Wave 4"],
    "n_total": [n_form1_w4, n_form2_w4],
    "n_target": [n_A_w4, n_D_w4],
    "Proportion_target": [round(n_A_w4/n_form1_w4, 3) if n_form1_w4 > 0 else np.nan, 
                           round(n_D_w4/n_form2_w4, 3) if n_form2_w4 > 0 else np.nan],
    "p-value": [round(p_val_form1_w4, 3), round(p_val_form2_w4, 3)]
})



# ---------------------------
# Now process LLM data similarly
# ---------------------------

# Extract Allais columns
allais_df_llm = df_llm[["TWIN_ID"] + allais_cols].copy()

# Apply mappings
allais_df_llm["Q192_mapped"] = allais_df_llm["Q192"].apply(map_allais_form1)
allais_df_llm["Q193_mapped"] = allais_df_llm["Q193"].apply(map_allais_form2)

# Frequency tables for LLM
freq_Q192_llm = freq_table(allais_df_llm["Q192_mapped"])
freq_Q193_llm = freq_table(allais_df_llm["Q193_mapped"])

freq_Q192_llm["Form"] = "Form 1 (Q192) - LLM"
freq_Q193_llm["Form"] = "Form 2 (Q193) - LLM"

freq_allais_df_llm = pd.concat([freq_Q192_llm, freq_Q193_llm], ignore_index=True)

# Binomial tests for LLM
form1_data_llm = allais_df_llm["Q192_mapped"].dropna()
n_form1_llm = form1_data_llm.shape[0]
n_A_llm = (form1_data_llm == "A").sum()
result_form1_llm = binomtest(n_A_llm, n_form1_llm, p=0.5, alternative='greater')
p_val_form1_llm = result_form1_llm.pvalue

form2_data_llm = allais_df_llm["Q193_mapped"].dropna()
n_form2_llm = form2_data_llm.shape[0]
n_D_llm = (form2_data_llm == "D").sum()
result_form2_llm = binomtest(n_D_llm, n_form2_llm, p=0.5, alternative='greater')
p_val_form2_llm = result_form2_llm.pvalue

binom_results_df_llm = pd.DataFrame({
    "Form": ["Form 1 (Q192) - LLM", "Form 2 (Q193) - LLM"],
    "n_total": [n_form1_llm, n_form2_llm],
    "n_target": [n_A_llm, n_D_llm],
    "Proportion_target": [round(n_A_llm/n_form1_llm, 3) if n_form1_llm > 0 else np.nan, 
                          round(n_D_llm/n_form2_llm, 3) if n_form2_llm > 0 else np.nan],
    "p-value": [round(p_val_form1_llm, 3), round(p_val_form2_llm, 3)]
})

# Append LLM results to the same sheet
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    current_max_row = ws.max_row

    # Section title for LLM
    ws.cell(row=current_max_row + 2, column=1, value="LLM - Allais Problem Summary")
    
    # Write LLM frequency tables
    freq_startrow_llm = current_max_row + 4
    freq_allais_df_llm.to_excel(writer, sheet_name=sheet_name, startrow=freq_startrow_llm, index=False)

    # Write LLM binomial results
    binom_startrow_llm = freq_startrow_llm + len(freq_allais_df_llm) + 3
    for r_idx, row in enumerate(dataframe_to_rows(binom_results_df_llm, index=False, header=True), start=binom_startrow_llm):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print("LLM Allais Problem results appended to Excel.")


# ---------------------------
# Append Wave 4 results to the same sheet
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    
    current_max_row = ws.max_row

    # Section title for Wave 4
    ws.cell(row=current_max_row + 2, column=1, value="Wave 4 - Allais Problem Summary")
    
    # Write Wave 4 frequency tables
    freq_startrow_w4 = current_max_row + 4
    freq_allais_df_w4.to_excel(writer, sheet_name=sheet_name, startrow=freq_startrow_w4, index=False)

    # Write Wave 4 binomial results
    binom_startrow_w4 = freq_startrow_w4 + len(freq_allais_df_w4) + 3
    for r_idx, row in enumerate(dataframe_to_rows(binom_results_df_w4, index=False, header=True), binom_startrow_w4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

print(f"Wave 4 Allais Problem results appended to '{sheet_name}' tab in '{output_filename}'.")
print("done")


# ## False consensus (Furnas & LaPira, 2024)

# In[77]:


import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
import statsmodels.formula.api as smf
from scipy.stats import pearsonr
import math
from tqdm import tqdm

# ---------------------------
# Helper: Filter valid rows
# ---------------------------
def valid_rows(df):
    cond1 = (df["PROLIFIC_PID"].astype(str).str.strip() != "")
    cond2 = (df["Finished"].astype(str).str.strip().str.upper() == "TRUE")
    return df[cond1 & cond2].copy()


# ---------------------------
# Determine respondents who completed all 4 waves
# ---------------------------
def get_common_ids(label_files):
    common_ids = None
    for wave, file in label_files.items():
        df = pd.read_csv(file, low_memory=False)
        df = valid_rows(df)
        ids = set(df["PROLIFIC_PID"].astype(str).str.strip())
        if common_ids is None:
            common_ids = ids
        else:
            common_ids = common_ids.intersection(ids)
    print(f"Number of respondents who finished all 4 waves: {len(common_ids)}")
    return common_ids

#common_ids = get_common_ids(label_files)


# ---------------------------
# Define column names
# ---------------------------
self_cols = [f"False Cons. self _{i}" for i in range(1, 11)]
others_cols = [f"False cons. others _{i}" for i in [1,2,3,4,5,6,7,10,11,12]]

# ---------------------------
# Map self responses to numeric
# ---------------------------
self_mapping = {
    "strongly support": 2,
    "somewhat support": 1,
    "neither oppose nor support": 0,
    "somewhat oppose": -1,
    "strongly oppose": -2
}

for col in self_cols:
    wave1_df[col + "_num"] = wave1_df[col].astype(str).str.lower().str.strip().map(self_mapping)

for col in others_cols:
    wave1_df[col] = pd.to_numeric(wave1_df[col], errors='coerce')

# ---------------------------
# Compute public support dynamically (Wave 1)
# ---------------------------
public_support_w1 = {}
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    valid_responses = wave1_df[self_col].dropna()
    n_total = valid_responses.shape[0]
    n_support = valid_responses[(valid_responses == 1) | (valid_responses == 2)].shape[0]
    support_pct = (n_support / n_total) * 100 if n_total > 0 else np.nan
    public_support_w1[i] = {"Policy": f"Policy {i}", "Actual": support_pct}

public_support_df_w1 = pd.DataFrame([
    {"Item": i, "Policy": v["Policy"], "Support_W1": v["Actual"]}
    for i, v in public_support_w1.items()
])

# ---------------------------
# Build long-format data for Wave 1
# ---------------------------
long_data = []
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    others_col = f"False cons. others _{i}" if i <= 7 else f"False cons. others _{i+2}"
    df_item = wave1_df[["PROLIFIC_PID", self_col, others_col]].dropna()
    df_item = df_item.rename(columns={self_col: "Self", others_col: "Predicted"})
    df_item["Policy"] = f"Policy {i}"
    df_item["Actual"] = public_support_w1[i]["Actual"]
    df_item["Misperception"] = df_item["Predicted"] - df_item["Actual"]
    df_item["Item"] = i
    long_data.append(df_item)

long_df = pd.concat(long_data, ignore_index=True)
long_df["Self_cat"] = pd.Categorical(long_df["Self"], categories=[-2, -1, 0, 1, 2])
long_df["Self_cat_str"] = long_df["Self_cat"].astype(str)

# ---------------------------
# TWFE Regression for Wave 1
# ---------------------------
model = smf.ols("Misperception ~ C(Self_cat_str, Treatment(reference='0')) + C(PROLIFIC_PID) + C(Policy)", data=long_df).fit()
summary_table = model.summary2().tables[1]
self_effects = summary_table.loc[summary_table.index.str.contains("Self_cat_str")].copy()
self_effects["Self_Rating"] = self_effects.index.str.extract(r"T\.(.*)\]", expand=False).astype(float)
self_effects = self_effects.reset_index(drop=True)
self_effects = self_effects[["Self_Rating", "Coef.", "Std.Err.", "t", "P>|t|", "[0.025", "0.975]"]]
self_effects = self_effects.rename(columns={
    "Coef.": "Coefficient", "Std.Err.": "Std_Error",
    "t": "t_value", "P>|t|": "p_value", "[0.025": "CI_lower", "0.975]": "CI_upper"
})

# ---------------------------
# Process Wave 4
# ---------------------------

for col in self_cols:
    wave4_df[col + "_num"] = wave4_df[col].astype(str).str.lower().str.strip().map(self_mapping)

for col in others_cols:
    wave4_df[col] = pd.to_numeric(wave4_df[col], errors='coerce')

# Compute public support (Wave 4)
public_support_w4 = {}
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    valid_responses = wave4_df[self_col].dropna()
    n_total = valid_responses.shape[0]
    n_support = valid_responses[(valid_responses == 1) | (valid_responses == 2)].shape[0]
    support_pct = (n_support / n_total) * 100 if n_total > 0 else np.nan
    public_support_w4[i] = {"Policy": f"Policy {i}", "Actual": support_pct}

public_support_df_w4 = pd.DataFrame([
    {"Item": i, "Policy": v["Policy"], "Support_W4": v["Actual"]}
    for i, v in public_support_w4.items()
])

# Build long-format data for Wave 4
long_data_w4 = []
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    others_col = f"False cons. others _{i}" if i <= 7 else f"False cons. others _{i+2}"
    df_item = wave4_df[["PROLIFIC_PID", self_col, others_col]].dropna()
    df_item = df_item.rename(columns={self_col: "Self", others_col: "Predicted"})
    df_item["Policy"] = f"Policy {i}"
    df_item["Actual"] = public_support_w4[i]["Actual"]
    df_item["Misperception"] = df_item["Predicted"] - df_item["Actual"]
    df_item["Item"] = i
    long_data_w4.append(df_item)

long_df_w4 = pd.concat(long_data_w4, ignore_index=True)
long_df_w4["Self_cat"] = pd.Categorical(long_df_w4["Self"], categories=[-2, -1, 0, 1, 2])
long_df_w4["Self_cat_str"] = long_df_w4["Self_cat"].astype(str)

# TWFE Regression for Wave 4
model_w4 = smf.ols("Misperception ~ C(Self_cat_str, Treatment(reference='0')) + C(PROLIFIC_PID) + C(Policy)", data=long_df_w4).fit()
summary_table_w4 = model_w4.summary2().tables[1]
self_effects_w4 = summary_table_w4.loc[summary_table_w4.index.str.contains("Self_cat_str")].copy()
self_effects_w4["Self_Rating"] = self_effects_w4.index.str.extract(r"T\.(.*)\]", expand=False).astype(float)
self_effects_w4 = self_effects_w4.reset_index(drop=True)
self_effects_w4 = self_effects_w4[["Self_Rating", "Coef.", "Std.Err.", "t", "P>|t|", "[0.025", "0.975]"]]
self_effects_w4 = self_effects_w4.rename(columns={
    "Coef.": "Coefficient", "Std.Err.": "Std_Error",
    "t": "t_value", "P>|t|": "p_value", "[0.025": "CI_lower", "0.975]": "CI_upper"
})


# ---------------------------
# Process LLM similarly
# ---------------------------
# Assumes df_llm is already loaded and cleaned with TWIN_ID and relevant columns
for col in self_cols:
    df_llm[col + "_num"] = df_llm[col].astype(str).str.lower().str.strip().map(self_mapping)

for col in others_cols:
    df_llm[col] = pd.to_numeric(df_llm[col], errors='coerce')

# Compute public support (LLM)
public_support_llm = {}
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    valid_responses = df_llm[self_col].dropna()
    n_total = valid_responses.shape[0]
    n_support = valid_responses[(valid_responses == 1) | (valid_responses == 2)].shape[0]
    support_pct = (n_support / n_total) * 100 if n_total > 0 else np.nan
    public_support_llm[i] = {"Policy": f"Policy {i}", "Actual": support_pct}

public_support_df_llm = pd.DataFrame([
    {"Item": i, "Policy": v["Policy"], "Support_LLM": v["Actual"]}
    for i, v in public_support_llm.items()
])

# Long format LLM
long_data_llm = []
for i in range(1, 11):
    self_col = f"False Cons. self _{i}_num"
    others_col = f"False cons. others _{i}" if i <= 7 else f"False cons. others _{i+2}"
    df_item = df_llm[["TWIN_ID", self_col, others_col]].dropna()
    df_item = df_item.rename(columns={self_col: "Self", others_col: "Predicted"})
    df_item["Policy"] = f"Policy {i}"
    df_item["Actual"] = public_support_llm[i]["Actual"]
    df_item["Misperception"] = df_item["Predicted"] - df_item["Actual"]
    df_item["Item"] = i
    long_data_llm.append(df_item)

long_df_llm = pd.concat(long_data_llm, ignore_index=True)
long_df_llm["Self_cat"] = pd.Categorical(long_df_llm["Self"], categories=[-2, -1, 0, 1, 2])
long_df_llm["Self_cat_str"] = long_df_llm["Self_cat"].astype(str)

# Regression for LLM
model_llm = smf.ols("Misperception ~ C(Self_cat_str, Treatment(reference='0')) + C(TWIN_ID) + C(Policy)", data=long_df_llm).fit()
summary_table_llm = model_llm.summary2().tables[1]
self_effects_llm = summary_table_llm.loc[summary_table_llm.index.str.contains("Self_cat_str")].copy()
self_effects_llm["Self_Rating"] = self_effects_llm.index.str.extract(r"T\.(.*)\]", expand=False).astype(float)
self_effects_llm = self_effects_llm.reset_index(drop=True)
self_effects_llm = self_effects_llm[["Self_Rating", "Coef.", "Std.Err.", "t", "P>|t|", "[0.025", "0.975]"]]
self_effects_llm = self_effects_llm.rename(columns={
    "Coef.": "Coefficient", "Std_Error": "Std_Error",
    "t": "t_value", "P>|t|": "p_value", "[0.025": "CI_lower", "0.975]": "CI_upper"
})


# ---------------------------
# Write everything to Excel
# ---------------------------
##output_filename = "experiment_analysis.xlsx"
wb = openpyxl.load_workbook(output_filename)
sheet_name = "W1-False Consensus"
if sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(sheet_name)
else:
    ws = wb[sheet_name]

# Write each section nicely separated
row_cursor = 1

# TWFE Wave 1
ws.cell(row=row_cursor, column=1, value="Two-Way Fixed Effects Regression - Wave 1")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(self_effects, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(self_effects) + 3

#LLM
ws.cell(row=row_cursor, column=1, value="Two-Way Fixed Effects Regression - LLM")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(self_effects_llm, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

row_cursor += len(self_effects_llm) + 3


# TWFE Wave 4
ws.cell(row=row_cursor, column=1, value="Two-Way Fixed Effects Regression - Wave 4")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(self_effects_w4, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(self_effects_w4) + 3

# Public Support Wave 1
ws.cell(row=row_cursor, column=1, value="Public Support - Wave 1")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(public_support_df_w1, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(public_support_df_w1) + 3

#LLM
ws.cell(row=row_cursor, column=1, value="Public Support - LLM")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(public_support_df_llm, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(public_support_df_llm) + 3

# Public Support Wave 4
ws.cell(row=row_cursor, column=1, value="Public Support - Wave 4")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(public_support_df_w4, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(public_support_df_w4) + 3

wb.save(output_filename)

print(f"All results written to '{output_filename}' in sheet '{sheet_name}'.")

# ---------------------------
# Write Self and Actual distributions to Excel
# ---------------------------
# Compute value counts
self_dist_w1 = long_df["Self"].value_counts(dropna=False).sort_index().reset_index()
self_dist_w1.columns = ["Self Rating", "Count"]
self_dist_w1.insert(0, "Wave", "Wave 1")

self_dist_w4 = long_df_w4["Self"].value_counts(dropna=False).sort_index().reset_index()
self_dist_w4.columns = ["Self Rating", "Count"]
self_dist_w4.insert(0, "Wave", "Wave 4")

self_dist_llm = long_df_llm["Self"].value_counts(dropna=False).sort_index().reset_index()
self_dist_llm.columns = ["Self Rating", "Count"]
self_dist_llm.insert(0, "Wave", "LLM")

actual_dist_w1 = long_df["Actual"].round().value_counts(dropna=False).sort_index().reset_index()
actual_dist_w1.columns = ["Actual (%)", "Count"]
actual_dist_w1.insert(0, "Wave", "Wave 1")

actual_dist_w4 = long_df_w4["Actual"].round().value_counts(dropna=False).sort_index().reset_index()
actual_dist_w4.columns = ["Actual (%)", "Count"]
actual_dist_w4.insert(0, "Wave", "Wave 4")

actual_dist_llm = long_df_llm["Actual"].round().value_counts(dropna=False).sort_index().reset_index()
actual_dist_llm.columns = ["Actual (%)", "Count"]
actual_dist_llm.insert(0, "Wave", "LLM")


# Combine into single tables
combined_self_dist = pd.concat([self_dist_w1,self_dist_llm, self_dist_w4], ignore_index=True)
combined_actual_dist = pd.concat([actual_dist_w1,actual_dist_llm, actual_dist_w4], ignore_index=True)

# Append to Excel
wb = openpyxl.load_workbook(output_filename)
ws = wb[sheet_name]
row_cursor = ws.max_row + 3

# Self rating distribution
ws.cell(row=row_cursor, column=1, value="Distribution of Self Ratings")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(combined_self_dist, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

row_cursor += len(combined_self_dist) + 3

# Actual public support distribution
ws.cell(row=row_cursor, column=1, value="Distribution of Actual Public Support (%)")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(combined_actual_dist, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Save file
wb.save(output_filename)

print("✅ Self and Actual response distributions added to Excel.")

# ---------------------------
# Append model fit stats and distributions to Excel
# ---------------------------
from collections import OrderedDict

# Summary stats for model fit
model_fit_summary = pd.DataFrame([
    {
        "Wave": "Wave 1",
        "R-squared": round(model.rsquared, 3),
        "Adj. R-squared": round(model.rsquared_adj, 3),
        "F-statistic": round(model.fvalue, 3) if model.fvalue is not None else np.nan,
        "p-value (F)": round(model.f_pvalue, 3) if model.f_pvalue is not None else np.nan,
        "n_obs": int(model.nobs)
    },
    {
        "Wave": "Wave 4",
        "R-squared": round(model_w4.rsquared, 3),
        "Adj. R-squared": round(model_w4.rsquared_adj, 3),
        "F-statistic": round(model_w4.fvalue, 3) if model_w4.fvalue is not None else np.nan,
        "p-value (F)": round(model_w4.f_pvalue, 3) if model_w4.f_pvalue is not None else np.nan,
        "n_obs": int(model_w4.nobs)
    },
    {
        "Wave": "LLM",
        "R-squared": round(model_llm.rsquared, 3),
        "Adj. R-squared": round(model_llm.rsquared_adj, 3),
        "F-statistic": round(model_llm.fvalue, 3) if model_llm.fvalue is not None else np.nan,
        "p-value (F)": round(model_llm.f_pvalue, 3) if model_llm.f_pvalue is not None else np.nan,
        "n_obs": int(model_llm.nobs)
    }
])

# Distributions of "False Cons. others" items
def get_others_distribution(df, wave_label):
    rows = []
    for col in others_cols:
        counts = df[col].value_counts(dropna=False).sort_index()
        for val, count in counts.items():
            rows.append(OrderedDict({
                "Wave": wave_label,
                "Item": col,
                "Value": val,
                "Count": count
            }))
    return pd.DataFrame(rows)

dist_wave1 = get_others_distribution(wave1_df, "Wave 1")
dist_wave4 = get_others_distribution(wave4_df, "Wave 4")
dist_llm = get_others_distribution(df_llm, "LLM")
dist_combined = pd.concat([dist_wave1, dist_llm,dist_wave4], ignore_index=True)

# Reopen and append to Excel
wb = openpyxl.load_workbook(output_filename)
ws = wb[sheet_name]
row_cursor = ws.max_row + 3

# Model Fit Summary
ws.cell(row=row_cursor, column=1, value="Model Fit Summary (TWFE)")
row_cursor += 2
for r_idx, row in enumerate(dataframe_to_rows(model_fit_summary, index=False, header=True), start=row_cursor):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
row_cursor += len(model_fit_summary) + 3


wb.save(output_filename)

print("✅ Regression summary stats and response distributions added to Excel.")



print("done")


# ## Nonseparability of risks and benefits (Stanovich & West, 2008)

# In[78]:


import pandas as pd
import numpy as np
from scipy.stats import pearsonr
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm


df_wave1_original = df_wave1.copy()
df_wave4_original = df_wave4.copy()

# Detect benefit and risk columns
bene_cols = sorted([col for col in df_wave1.columns if col.lower().startswith("nonseparabilty bene")])
risk_cols = sorted([col for col in df_wave1.columns if col.lower().startswith("nonseparability ris")])

#print("df_wave1 columns:", list(df_wave1.columns))

print(bene_cols)
print(risk_cols)

# Mapping
def robust_map_response(resp, mapping):
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    for key, val in mapping.items():
        if key in text:
            return float(val)
    return np.nan

bene_mapping = {
    "not at all beneficial": 1,
    "low benefit": 2,
    "slightly beneficial": 3,
    "neutral": 4,
    "moderately beneficial": 5,
    "very beneficial": 6,
    "extremely beneficial": 7
}
risk_mapping = {
    "not at all risky": 1,
    "low risk": 2,
    "slightly risky": 3,
    "neutral": 4,
    "moderately risky": 5,
    "very risky": 6,
    "extremely risky": 7
}

for col in bene_cols:
    df_wave1[col + "_num"] = df_wave1[col].apply(lambda x: robust_map_response(x, bene_mapping))
for col in risk_cols:
    df_wave1[col + "_num"] = df_wave1[col].apply(lambda x: robust_map_response(x, risk_mapping))

bene_num_cols = [col + "_num" for col in bene_cols]
risk_num_cols = [col + "_num" for col in risk_cols]

# ---------------------------
# Analyze Wave 1
# ---------------------------
tech_names = {1: "Bicycles", 2: "Alcoholic beverages", 3: "Chemical plants", 4: "Pesticides"}
results_w1 = []

for i in range(4):
    bene_col = bene_num_cols[i]
    risk_col = risk_num_cols[i]
    df_item = df_wave1[[bene_col, risk_col]].dropna()
    n = df_item.shape[0]
    if n > 1:
        r, p = pearsonr(df_item[bene_col], df_item[risk_col])
        t_stat = r * np.sqrt((n-2) / (1-r**2)) if n > 2 else np.nan
    else:
        r, p, t_stat = np.nan, np.nan, np.nan
    mean_bene = df_item[bene_col].mean()
    mean_risk = df_item[risk_col].mean()
    results_w1.append({
        "Item": i+1,
        "Technology": tech_names.get(i+1, f"Item {i+1}"),
        "n": n,
        "Mean Benefit": round(mean_bene, 2),
        "Mean Risk": round(mean_risk, 2),
        "Correlation": round(r, 3),
        "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
        "p-value": round(p, 3) if not np.isnan(p) else np.nan
    })

results_df_w1 = pd.DataFrame(results_w1)

# ---------------------------
# Wave 4
# ---------------------------

for col in bene_cols:
    df_wave4[col + "_num"] = df_wave4[col].apply(lambda x: robust_map_response(x, bene_mapping))
for col in risk_cols:
    df_wave4[col + "_num"] = df_wave4[col].apply(lambda x: robust_map_response(x, risk_mapping))

results_w4 = []

for i in range(4):
    bene_col = bene_cols[i] + "_num"
    risk_col = risk_cols[i] + "_num"
    df_item = df_wave4[[bene_col, risk_col]].dropna()
    n = df_item.shape[0]
    if n > 1:
        r, p = pearsonr(df_item[bene_col], df_item[risk_col])
        t_stat = r * np.sqrt((n-2) / (1-r**2)) if n > 2 else np.nan
    else:
        r, p, t_stat = np.nan, np.nan, np.nan
    mean_bene = df_item[bene_col].mean()
    mean_risk = df_item[risk_col].mean()
    results_w4.append({
        "Item": i+1,
        "Technology": tech_names.get(i+1, f"Item {i+1}"),
        "n": n,
        "Mean Benefit": round(mean_bene, 2),
        "Mean Risk": round(mean_risk, 2),
        "Correlation": round(r, 3),
        "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
        "p-value": round(p, 3) if not np.isnan(p) else np.nan
    })

results_df_w4 = pd.DataFrame(results_w4)


# ---------------------------
# LLM Data
# ---------------------------

# Apply numeric mappings to LLM data
for col in bene_cols:
    df_llm[col + "_num"] = df_llm[col].apply(lambda x: robust_map_response(x, bene_mapping))
for col in risk_cols:
    df_llm[col + "_num"] = df_llm[col].apply(lambda x: robust_map_response(x, risk_mapping))

results_llm = []

for i in range(4):
    bene_col = bene_cols[i] + "_num"
    risk_col = risk_cols[i] + "_num"
    df_item = df_llm[[bene_col, risk_col]].dropna()
    n = df_item.shape[0]
    if n > 1:
        r, p = pearsonr(df_item[bene_col], df_item[risk_col])
        t_stat = r * np.sqrt((n-2) / (1-r**2)) if n > 2 else np.nan
    else:
        r, p, t_stat = np.nan, np.nan, np.nan
    mean_bene = df_item[bene_col].mean()
    mean_risk = df_item[risk_col].mean()
    results_llm.append({
        "Item": i+1,
        "Technology": tech_names.get(i+1, f"Item {i+1}"),
        "n": n,
        "Mean Benefit": round(mean_bene, 2),
        "Mean Risk": round(mean_risk, 2),
        "Correlation": round(r, 3),
        "t-statistic": round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
        "p-value": round(p, 3) if not np.isnan(p) else np.nan
    })

results_df_llm = pd.DataFrame(results_llm)

# ---------------------------
# Write all results to Excel
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W1 Nonseparability"

try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
except FileNotFoundError:
    sheet_name_new = base_sheet_name

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    results_df_w1.to_excel(writer, sheet_name=sheet_name_new, startrow=3, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value="Wave 1 Results")

    startrow = ws.max_row + 3
    results_df_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws.cell(row=startrow - 2, column=1, value="LLM Results")
    
    startrow = ws.max_row + 3
    results_df_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws.cell(row=startrow-2, column=1, value="Wave 4 Results")

print(f"Excel file '{output_filename}' updated with '{sheet_name_new}' tab.")

# ---------------------------
# Compute and write value distributions of _num variables
# ---------------------------
def get_distribution(df, cols, wave_label):
    rows = []
    for col in cols:
        counts = df[col].value_counts(dropna=False).sort_index()
        for val, count in counts.items():
            rows.append({
                "Wave": wave_label,
                "Variable": col,
                "Response Value": val,
                "Count": count
            })
    return pd.DataFrame(rows)

# Get distributions for benefit and risk _num variables
dist_w1 = get_distribution(df_wave1, bene_num_cols + risk_num_cols, "Wave 1")
dist_llm = get_distribution(df_llm, bene_num_cols + risk_num_cols, "LLM")
dist_w4 = get_distribution(df_wave4, bene_num_cols + risk_num_cols, "Wave 4")
dist_all = pd.concat([dist_w1, dist_llm, dist_w4], ignore_index=True)

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    dist_startrow = ws.max_row + 4
    ws.cell(row=dist_startrow, column=1, value="Distribution of _num Variables")
    dist_startrow += 2
    dist_all.to_excel(writer, sheet_name=sheet_name_new, startrow=dist_startrow, index=False)

print("✅ Response distributions for _num variables written to Excel.")

df_wave1=df_wave1_original 
df_wave4=df_wave4_original 

print("done")


# # Wave 2

# ## Framing problem (Tversky & Kahneman, 1981)

# In[79]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl


# ---------------------------
# Define mapping for framing responses (6-point Likert)
# ---------------------------
def map_framing_response(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    # Check if response refers to Program A or B.
    if "program a" in text:
        if "strongly" in text:
            return 1
        elif "slightly" in text:
            return 3
        else:
            return 2
    elif "program b" in text:
        if "strongly" in text:
            return 6
        elif "slightly" in text:
            return 4
        else:
            return 5
    else:
        return np.nan

# Map responses for Q157 and Q158
df_wave2["Q157_mapped"] = df_wave2["Q157"].apply(map_framing_response)
df_wave2["Q158_mapped"] = df_wave2["Q158"].apply(map_framing_response)

# ---------------------------
# Separate the two groups (each subject sees one question)
# ---------------------------
group_gain = df_wave2.loc[df_wave2["Q157_mapped"].notna(), "Q157_mapped"]
group_loss = df_wave2.loc[df_wave2["Q158_mapped"].notna(), "Q158_mapped"]

# Compute descriptive statistics for each group.
desc_gain = {
    "n": group_gain.shape[0],
    "Mean": group_gain.mean(),
    "SD": group_gain.std(ddof=1)
}
desc_loss = {
    "n": group_loss.shape[0],
    "Mean": group_loss.mean(),
    "SD": group_loss.std(ddof=1)
}

# Perform independent samples t-test (assuming unequal variances).
if group_gain.shape[0] > 1 and group_loss.shape[0] > 1:
    t_stat, p_val = ttest_ind(group_gain, group_loss, equal_var=False)
else:
    t_stat, p_val = np.nan, np.nan

# ---------------------------
# Prepare summary tables
# ---------------------------
summary_table = pd.DataFrame({
    "Group": ["Gain Frame (Q157)", "Loss Frame (Q158)"],
    "n": [desc_gain["n"], desc_loss["n"]],
    "Mean": [round(desc_gain["Mean"], 2), round(desc_loss["Mean"], 2)],
    "SD": [round(desc_gain["SD"], 2), round(desc_loss["SD"], 2)]
})
ttest_table = pd.DataFrame({
    "t-statistic": [round(t_stat, 3) if not np.isnan(t_stat) else np.nan],
    "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan]
})

# Combine the two tables into one DataFrame (insert a blank row between them).
blank_row = pd.DataFrame([[""] * summary_table.shape[1]], columns=summary_table.columns)
combined_df = pd.concat([summary_table, blank_row, ttest_table], ignore_index=True)

# ---------------------------
# Prepare header note
# ---------------------------
header_note = (
    "Analysis: Framing Problem \n\n"
    "Participants were randomly assigned to one of two versions of the framing problem:\n\n"
    "Positive Frame (Q157):\n"
    "  - If Program A is adopted, 200 people will be saved.\n"
    "  - If Program B is adopted, there is a 1/3 probability that 600 people will be saved, and a 2/3 probability that no people will be saved.\n\n"
    "Negative Frame (Q158):\n"
    "  - If Program A is adopted, 400 people will die.\n"
    "  - If Program B is adopted, there is a 1/3 probability that nobody will die, and a 2/3 probability that 600 people will die.\n\n"
    "Responses were provided on a 6-point Likert scale:\n"
    "  1 = strongly favor Program A, 2 = slightly favor Program A, 3 = favor Program A,\n"
    "  4 = favor Program B, 5 = slightly favor Program B, 6 = strongly favor Program B.\n\n"
    "Expectation from Tversky & Kahneman is that in the loss frame, participants will be more inclined to choose Program B -> higher scores.\n\n"
)

# ---------------------------
# Determine a unique new sheet name
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Framing"
try:
    wb = openpyxl.load_workbook(output_filename)
    # Find a unique sheet name by appending a suffix if necessary.
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

print("Final chosen sheet name:", sheet_name)

# ---------------------------
# Write everything to the new sheet in one go
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    # Write the combined DataFrame once.
    combined_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    # Get the worksheet and write the header note in cell A1.
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


# ---------------------------
# Now add Wave 4 analysis 
# ---------------------------

# Map framing responses in Wave 4
df_wave4["Q157_mapped"] = df_wave4["Q157"].apply(map_framing_response)
df_wave4["Q158_mapped"] = df_wave4["Q158"].apply(map_framing_response)

group_gain_w4 = df_wave4.loc[df_wave4["Q157_mapped"].notna(), "Q157_mapped"]
group_loss_w4 = df_wave4.loc[df_wave4["Q158_mapped"].notna(), "Q158_mapped"]

# Compute descriptive statistics for Wave 4
desc_gain_w4 = {
    "n": group_gain_w4.shape[0],
    "Mean": group_gain_w4.mean(),
    "SD": group_gain_w4.std(ddof=1)
}
desc_loss_w4 = {
    "n": group_loss_w4.shape[0],
    "Mean": group_loss_w4.mean(),
    "SD": group_loss_w4.std(ddof=1)
}

# Independent samples t-test for Wave 4
if group_gain_w4.shape[0] > 1 and group_loss_w4.shape[0] > 1:
    t_stat_w4, p_val_w4 = ttest_ind(group_gain_w4, group_loss_w4, equal_var=False)
else:
    t_stat_w4, p_val_w4 = np.nan, np.nan

    
# ---------------------------
# LLM Data
# ---------------------------

# Map framing responses in LLM data
df_llm["Q157_mapped"] = df_llm["Q157"].apply(map_framing_response)
df_llm["Q158_mapped"] = df_llm["Q158"].apply(map_framing_response)

group_gain_llm = df_llm.loc[df_llm["Q157_mapped"].notna(), "Q157_mapped"]
group_loss_llm = df_llm.loc[df_llm["Q158_mapped"].notna(), "Q158_mapped"]

# Compute descriptive statistics for LLM
desc_gain_llm = {
    "n": group_gain_llm.shape[0],
    "Mean": group_gain_llm.mean(),
    "SD": group_gain_llm.std(ddof=1)
}
desc_loss_llm = {
    "n": group_loss_llm.shape[0],
    "Mean": group_loss_llm.mean(),
    "SD": group_loss_llm.std(ddof=1)
}

# Independent samples t-test for LLM
if group_gain_llm.shape[0] > 1 and group_loss_llm.shape[0] > 1:
    t_stat_llm, p_val_llm = ttest_ind(group_gain_llm, group_loss_llm, equal_var=False)
else:
    t_stat_llm, p_val_llm = np.nan, np.nan

# Build LLM summary tables
summary_llm = pd.DataFrame({
    "Group": ["Gain Frame (Q157)", "Loss Frame (Q158)"],
    "n": [desc_gain_llm["n"], desc_loss_llm["n"]],
    "Mean": [round(desc_gain_llm["Mean"], 2), round(desc_loss_llm["Mean"], 2)],
    "SD": [round(desc_gain_llm["SD"], 2), round(desc_loss_llm["SD"], 2)]
})
ttest_llm = pd.DataFrame({
    "t-statistic": [round(t_stat_llm, 3)],
    "p-value": [round(p_val_llm, 3)]
})    
     
# ---------------------------
# Build new tables
# ---------------------------

# Wave 4 summary
summary_w4 = pd.DataFrame({
    "Group": ["Gain Frame (Q157)", "Loss Frame (Q158)"],
    "n": [desc_gain_w4["n"], desc_loss_w4["n"]],
    "Mean": [round(desc_gain_w4["Mean"], 2), round(desc_loss_w4["Mean"], 2)],
    "SD": [round(desc_gain_w4["SD"], 2), round(desc_loss_w4["SD"], 2)]
})
ttest_w4 = pd.DataFrame({
    "t-statistic": [round(t_stat_w4, 3)],
    "p-value": [round(p_val_w4, 3)]
})

# ---------------------------
# Append new outputs to the same Excel tab
# ---------------------------
from openpyxl.styles import Font

wb = openpyxl.load_workbook(output_filename)
ws = wb[sheet_name]


startrow = ws.max_row + 3

# Write LLM Summary
ws.cell(row=startrow-1, column=1, value="LLM Results").font = Font(bold=True, size=12)
for idx, col in enumerate(summary_llm.columns, 1):
    ws.cell(row=startrow, column=idx, value=col)
for i, row in summary_llm.iterrows():
    for j, value in enumerate(row, 1):
        ws.cell(row=startrow + i + 1, column=j, value=value)

# Write LLM T-test
startrow = startrow + len(summary_llm) + 3
ws.cell(row=startrow-1, column=1, value="LLM T-test Results").font = Font(bold=True, size=12)
for idx, col in enumerate(ttest_llm.columns, 1):
    ws.cell(row=startrow, column=idx, value=col)
for i, row in ttest_llm.iterrows():
    for j, value in enumerate(row, 1):
        ws.cell(row=startrow + i + 1, column=j, value=value)

startrow = ws.max_row + 3

# Write Wave 4 Summary   
ws.cell(row=startrow-1, column=1, value="Wave 4 Results").font = Font(bold=True, size=12)
for idx, col in enumerate(summary_w4.columns, 1):
    ws.cell(row=startrow, column=idx, value=col)
for i, row in summary_w4.iterrows():
    for j, value in enumerate(row, 1):
        ws.cell(row=startrow + i + 1, column=j, value=value)
        
# Write Wave 4 T-test
startrow = startrow + len(summary_w4) + 3
ws.cell(row=startrow-1, column=1, value="Wave 4 T-test Results").font = Font(bold=True, size=12)
for idx, col in enumerate(ttest_w4.columns, 1):
    ws.cell(row=startrow, column=idx, value=col)
for i, row in ttest_w4.iterrows():
    for j, value in enumerate(row, 1):
        ws.cell(row=startrow+i+1, column=j, value=value)

wb.save(output_filename)
print(f"Excel file '{output_filename}' updated with Wave 4 analysis.")


# ---------------------------
# Write distribution of Q157_mapped and Q158_mapped for Wave 2 and Wave 4
# ---------------------------

def get_mapped_distribution(df, wave_label):
    dist_rows = []
    for q in ["Q157_mapped", "Q158_mapped"]:
        counts = df[q].value_counts(dropna=False).sort_index()
        for val, count in counts.items():
            dist_rows.append({
                "Wave": wave_label,
                "Question": q,
                "Mapped Value": val,
                "Count": count
            })
    return pd.DataFrame(dist_rows)

dist_wave2 = get_mapped_distribution(df_wave2, "Wave 2")
dist_llm = get_mapped_distribution(df_llm, "LLM")
dist_wave4 = get_mapped_distribution(df_wave4, "Wave 4")
dist_combined = pd.concat([dist_wave2, dist_llm, dist_wave4], ignore_index=True)

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow, column=1, value="Distribution of Mapped Responses for Q157 and Q158").font = Font(bold=True, size=12)
    startrow += 2
    dist_combined.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

print("Distributions of Q157_mapped and Q158_mapped written to Excel.")

print("done")


# ## Conjunction fallacy / "Linda problem" (Tversky & Kahneman, 1983)

# In[80]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import re
import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm
from scipy.stats import pearsonr


# ---------------------------
# Linda problem: Use six columns Q159_1, Q159_2, Q159_3 and Q160_1, Q160_2, Q160_3.
# We will use the third question from each set as the target:
#   - Condition 1 (Single event): Q159_3 ("Linda is a bank teller")
#   - Condition 2 (Conjunction): Q160_3 ("Linda is a bank teller and is active in the feminist movement")
# ---------------------------
# Define a robust mapping function for 6-point Likert responses.
def robust_map_linda(resp, mapping):
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    for key, val in mapping.items():
        if key in text:
            return float(val)
    return np.nan

# Define the mapping dictionary.
linda_mapping = {
    "extremely improbable": 1,
    "very improbable": 2,
    "somewhat probable": 3,
    "moderately probable": 4,
    "very probable": 5,
    "extremely probable": 6
}

# Map responses for the target items.
df_wave2["Linda_Single"] = df_wave2["Q159_3"].apply(lambda x: robust_map_linda(x, linda_mapping))
df_wave2["Linda_Conjunction"] = df_wave2["Q160_3"].apply(lambda x: robust_map_linda(x, linda_mapping))

# ---------------------------
# Separate the two conditions.
# ---------------------------
group_single = df_wave2.loc[df_wave2["Linda_Single"].notna(), "Linda_Single"]
group_conj = df_wave2.loc[df_wave2["Linda_Conjunction"].notna(), "Linda_Conjunction"]

# ---------------------------
# Compute descriptive statistics for each condition.
# ---------------------------
desc_single = {
    "n": group_single.shape[0],
    "Mean": group_single.mean(),
    "SD": group_single.std(ddof=1)
}
desc_conj = {
    "n": group_conj.shape[0],
    "Mean": group_conj.mean(),
    "SD": group_conj.std(ddof=1)
}

# Perform independent samples t-test (assuming unequal variances).
if group_single.shape[0] > 1 and group_conj.shape[0] > 1:
    t_stat, p_val = ttest_ind(group_single, group_conj, equal_var=False)
else:
    t_stat, p_val = np.nan, np.nan

# ---------------------------
# Prepare summary tables.
# ---------------------------
summary_table = pd.DataFrame({
    "Group": ["Single Event (Q159_3)", "Conjunction (Q160_3)"],
    "n": [desc_single["n"], desc_conj["n"]],
    "Mean": [round(desc_single["Mean"], 2), round(desc_conj["Mean"], 2)],
    "SD": [round(desc_single["SD"], 2), round(desc_conj["SD"], 2)]
})
ttest_table = pd.DataFrame({
    "t-statistic": [round(t_stat, 3) if not np.isnan(t_stat) else np.nan],
    "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan]
})

# Combine the tables into one DataFrame (with a blank row between).
blank_row = pd.DataFrame([[""] * summary_table.shape[1]], columns=summary_table.columns)
combined_df = pd.concat([summary_table, blank_row, ttest_table], ignore_index=True)

# ---------------------------
# Prepare header note.
# ---------------------------
header_note = (
    "Analysis: Linda Problem \n\n"
    "Participants read a description of Linda and then completed one of two sets of questions.\n\n"
    "Condition 1 (Q159): Subjects answered three questions, with the third asking:\n"
    "  'It is ___ that Linda is a bank teller.'\n\n"
    "Condition 2 (Q160): Subjects answered three questions, with the third asking:\n"
    "  'It is ___ that Linda is a bank teller and is active in the feminist movement.'\n\n"
    "Responses were provided on a 6-point Likert scale with text options such as:\n"
    "  'Extremely improbable', 'Very improbable', 'Somewhat probable', 'Moderately probable', 'Very probable', 'Extremely probable'.\n"
    "These were mapped numerically (1 = Extremely improbable, 2 = Very improbable, 3 = Somewhat probable,\n"
    "4 = Moderately probable, 5 = Very probable, 6 = Extremely probable).\n\n"
    "If subjects exhibit the conjunction fallacy, the mean probability rating for the conjunction\n"
    "condition (Q160_3) will be higher than for the single event condition (Q159_3).\n"
)

# ---------------------------
# Determine a unique new sheet name and write everything at once.
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Linda"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

print("Final chosen sheet name:", sheet_name)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    combined_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


# --------------------------------------------------
# NEW ANALYSIS: Repeat for Wave 4
# --------------------------------------------------
# Map responses for Wave 4
df_wave4["Linda_Single"] = df_wave4["Q159_3"].apply(lambda x: robust_map_linda(x, linda_mapping))
df_wave4["Linda_Conjunction"] = df_wave4["Q160_3"].apply(lambda x: robust_map_linda(x, linda_mapping))

# Separate the two conditions for Wave 4
group_single_w4 = df_wave4.loc[df_wave4["Linda_Single"].notna(), "Linda_Single"]
group_conj_w4 = df_wave4.loc[df_wave4["Linda_Conjunction"].notna(), "Linda_Conjunction"]

# Descriptive stats for Wave 4
desc_single_w4 = {
    "n": group_single_w4.shape[0],
    "Mean": group_single_w4.mean(),
    "SD": group_single_w4.std(ddof=1)
}
desc_conj_w4 = {
    "n": group_conj_w4.shape[0],
    "Mean": group_conj_w4.mean(),
    "SD": group_conj_w4.std(ddof=1)
}

# T-test for Wave 4
if group_single_w4.shape[0] > 1 and group_conj_w4.shape[0] > 1:
    t_stat_w4, p_val_w4 = ttest_ind(group_single_w4, group_conj_w4, equal_var=False)
else:
    t_stat_w4, p_val_w4 = np.nan, np.nan

# Prepare Wave 4 tables
summary_table_w4 = pd.DataFrame({
    "Group": ["Single Event (Q159_3)", "Conjunction (Q160_3)"],
    "n": [desc_single_w4["n"], desc_conj_w4["n"]],
    "Mean": [round(desc_single_w4["Mean"], 2), round(desc_conj_w4["Mean"], 2)],
    "SD": [round(desc_single_w4["SD"], 2), round(desc_conj_w4["SD"], 2)]
})
ttest_table_w4 = pd.DataFrame({
    "t-statistic": [round(t_stat_w4, 3) if not np.isnan(t_stat_w4) else np.nan],
    "p-value": [round(p_val_w4, 3) if not np.isnan(p_val_w4) else np.nan]
})


# --------------------------------------------------
# NEW ANALYSIS: Repeat for LLM
# --------------------------------------------------

# Map responses for LLM
df_llm["Linda_Single"] = df_llm["Q159_3"].apply(lambda x: robust_map_linda(x, linda_mapping))
df_llm["Linda_Conjunction"] = df_llm["Q160_3"].apply(lambda x: robust_map_linda(x, linda_mapping))

# Separate the two conditions
group_single_llm = df_llm.loc[df_llm["Linda_Single"].notna(), "Linda_Single"]
group_conj_llm = df_llm.loc[df_llm["Linda_Conjunction"].notna(), "Linda_Conjunction"]

# Descriptive stats
desc_single_llm = {
    "n": group_single_llm.shape[0],
    "Mean": group_single_llm.mean(),
    "SD": group_single_llm.std(ddof=1)
}
desc_conj_llm = {
    "n": group_conj_llm.shape[0],
    "Mean": group_conj_llm.mean(),
    "SD": group_conj_llm.std(ddof=1)
}

# T-test
if group_single_llm.shape[0] > 1 and group_conj_llm.shape[0] > 1:
    t_stat_llm, p_val_llm = ttest_ind(group_single_llm, group_conj_llm, equal_var=False)
else:
    t_stat_llm, p_val_llm = np.nan, np.nan

# Prepare summary tables
summary_llm = pd.DataFrame({
    "Group": ["Single Event (Q159_3)", "Conjunction (Q160_3)"],
    "n": [desc_single_llm["n"], desc_conj_llm["n"]],
    "Mean": [round(desc_single_llm["Mean"], 2), round(desc_conj_llm["Mean"], 2)],
    "SD": [round(desc_single_llm["SD"], 2), round(desc_conj_llm["SD"], 2)]
})
ttest_llm = pd.DataFrame({
    "t-statistic": [round(t_stat_llm, 3)],
    "p-value": [round(p_val_llm, 3)]
})

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow - 1, column=1, value="LLM Summary").font = Font(bold=True, size=12)

    # Write summary
    summary_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

    # Write t-test
    startrow += len(summary_llm) + 3
    ttest_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

# --------------------------------------------------
# Write new results into the same Excel sheet
# --------------------------------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    # Write the summary table with a header above it
    ws = writer.book[sheet_name]
    startrow_new = startrow + combined_df.shape[0] + 3
    ws.cell(row=startrow_new - 1, column=1, value="Wave 4 Summary").font = Font(bold=True, size=12)
    
    summary_table_w4.to_excel(writer, sheet_name=sheet_name, startrow=startrow_new, index=False)
    ttest_table_w4.to_excel(writer, sheet_name=sheet_name, startrow=startrow_new + len(summary_table_w4) + 3, index=False)
  #  corr_table.to_excel(writer, sheet_name=sheet_name, startrow=startrow_new + len(summary_table_w4) + len(ttest_table_w4) + 8, index=False)

print(f"Wave 4 analysis added to '{sheet_name}' in '{output_filename}'.")

# ---------------------------
# Write distribution of Linda_Single and Linda_Conjunction for Wave 2 and Wave 4
# ---------------------------

def get_mapped_distribution(df, wave_label):
    dist_rows = []
    for q in ["Linda_Single", "Linda_Conjunction"]:
        counts = df[q].value_counts(dropna=False).sort_index()
        for val, count in counts.items():
            dist_rows.append({
                "Wave": wave_label,
                "Question": q,
                "Mapped Value": val,
                "Count": count
            })
    return pd.DataFrame(dist_rows)

dist_wave2 = get_mapped_distribution(df_wave2, "Wave 2")
dist_llm = get_mapped_distribution(df_llm, "LLM")
dist_wave4 = get_mapped_distribution(df_wave4, "Wave 4")
dist_combined = pd.concat([dist_wave2,dist_llm, dist_wave4], ignore_index=True)

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow, column=1, value="Distribution of Mapped Responses for Linda_Single and Linda_Conjunction").font = Font(bold=True, size=12)
    startrow += 2
    dist_combined.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

print("Distributions of Linda_Single and Linda_Conjunction (Wave 2 and 4) written to Excel.")


print("done")


# ## Anchoring and adjustment

# In[81]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl

# -----------
# ---------------------------
# African Countries:
#   - Low Anchor: Q163 is the anchor question; Q164 is the numeric estimate (anchor = 12)
#   - High Anchor: Q165 is the anchor question; Q166 is the numeric estimate (anchor = 65)
# Trees:
#   - Low Anchor: Q167 is the anchor question; Q168 is the numeric estimate (anchor = 85 feet)
#   - High Anchor: Q169 is the anchor question; Q170 is the numeric estimate (anchor = 1000 feet)
# ---------------------------
# Convert relevant columns to numeric.
for col in ["Q164", "Q166", "Q168", "Q170"]:
    df_wave2[col] = pd.to_numeric(df_wave2[col], errors='coerce')

# ---------------------------
# African countries analysis
# ---------------------------
africa_low = df_wave2.loc[df_wave2["Q164"].notna(), "Q164"]
africa_high = df_wave2.loc[df_wave2["Q166"].notna(), "Q166"]

desc_africa_low = {
    "n": africa_low.shape[0],
    "Mean": africa_low.mean(),
    "SD": africa_low.std(ddof=1),
    "Median": africa_low.median(),
    "Min": africa_low.min(),
    "Max": africa_low.max()
}
desc_africa_high = {
    "n": africa_high.shape[0],
    "Mean": africa_high.mean(),
    "SD": africa_high.std(ddof=1),
    "Median": africa_high.median(),
    "Min": africa_high.min(),
    "Max": africa_high.max()
}

if africa_low.shape[0] > 1 and africa_high.shape[0] > 1:
    t_stat_africa, p_val_africa = ttest_ind(africa_high, africa_low, equal_var=False)
else:
    t_stat_africa, p_val_africa = np.nan, np.nan

africa_summary = pd.DataFrame({
    "Domain": ["African Countries", "African Countries"],
    "Anchor": ["Low (12)", "High (65)"],
    "n": [desc_africa_low["n"], desc_africa_high["n"]],
    "Mean Estimate": [round(desc_africa_low["Mean"], 2), round(desc_africa_high["Mean"], 2)],
    "SD": [round(desc_africa_low["SD"], 2), round(desc_africa_high["SD"], 2)],
    "Median": [round(desc_africa_low["Median"], 2), round(desc_africa_high["Median"], 2)],
    "Min": [desc_africa_low["Min"], desc_africa_high["Min"]],
    "Max": [desc_africa_low["Max"], desc_africa_high["Max"]]
})
africa_ttest = pd.DataFrame({
    "Domain": ["African Countries"],
    "t-statistic": [round(t_stat_africa, 3) if not np.isnan(t_stat_africa) else np.nan],
    "p-value": [round(p_val_africa, 3) if not np.isnan(p_val_africa) else np.nan]
})

# ---------------------------
# Tree analysis
# ---------------------------
tree_low = df_wave2.loc[df_wave2["Q168"].notna(), "Q168"]
tree_high = df_wave2.loc[df_wave2["Q170"].notna(), "Q170"]

desc_tree_low = {
    "n": tree_low.shape[0],
    "Mean": tree_low.mean(),
    "SD": tree_low.std(ddof=1),
    "Median": tree_low.median(),
    "Min": tree_low.min(),
    "Max": tree_low.max()
}
desc_tree_high = {
    "n": tree_high.shape[0],
    "Mean": tree_high.mean(),
    "SD": tree_high.std(ddof=1),
    "Median": tree_high.median(),
    "Min": tree_high.min(),
    "Max": tree_high.max()
}

if tree_low.shape[0] > 1 and tree_high.shape[0] > 1:
    t_stat_tree, p_val_tree = ttest_ind(tree_high, tree_low, equal_var=False)
else:
    t_stat_tree, p_val_tree = np.nan, np.nan

tree_summary = pd.DataFrame({
    "Domain": ["Trees", "Trees"],
    "Anchor": ["Low (85 feet)", "High (1000 feet)"],
    "n": [desc_tree_low["n"], desc_tree_high["n"]],
    "Mean": [round(desc_tree_low["Mean"], 2), round(desc_tree_high["Mean"], 2)],
    "SD": [round(desc_tree_low["SD"], 2), round(desc_tree_high["SD"], 2)],
    "Median": [round(desc_tree_low["Median"], 2), round(desc_tree_high["Median"], 2)],
    "Min": [desc_tree_low["Min"], desc_tree_high["Min"]],
    "Max": [desc_tree_low["Max"], desc_tree_high["Max"]]
})
tree_ttest = pd.DataFrame({
    "Domain": ["Trees"],
    "t-statistic": [round(t_stat_tree, 3) if not np.isnan(t_stat_tree) else np.nan],
    "p-value": [round(p_val_tree, 3) if not np.isnan(p_val_tree) else np.nan]
})

# ---------------------------
# Combine African and tree results into one final table.
# ---------------------------
combined_summary = pd.concat([africa_summary, pd.DataFrame([[""]*africa_summary.shape[1]], columns=africa_summary.columns),
                                tree_summary], ignore_index=True)
combined_ttest = pd.concat([africa_ttest, tree_ttest], ignore_index=True)

final_combined = pd.concat([combined_summary, pd.DataFrame([[""]*combined_summary.shape[1]], columns=combined_summary.columns), combined_ttest], ignore_index=True)

# ---------------------------
# Prepare header note.
# ---------------------------
header_note = (
    "Analysis: Anchoring Effects \n\n"
    "Participants were assigned to either a low or high anchor condition in two domains:\n\n"
    "African Countries:\n"
    "  - Low Anchor: Q163 asks about whether there are more or fewer than 12 African countries in the UN, then Q164 asks for a numeric estimate.\n"
    "  - High Anchor: Q165 asks about whether there are more or fewer than 65 African countries in the UN, then Q166 asks for a numeric estimate.\n\n"
    "Trees:\n"
    "  - Low Anchor: Q167 asks if the tallest redwood is more or less than 85 feet tall, then Q168 asks for a numeric estimate.\n"
    "  - High Anchor: Q169 asks if the tallest redwood is more or less than 1000 feet tall, then Q170 asks for a numeric estimate.\n\n"
    "For each domain, descriptive statistics (n, mean, SD, median, min, max) of the numeric estimates are reported by anchor condition,\n"
    "and an independent samples t-test compares the high vs. low anchor groups.\n"
    "In both domains, we expect subjects anchored high will provide higher estimates than the subjects anchored low."
)

# ---------------------------
# Determine a unique new sheet name and write everything in one go.
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Anchoring"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name = base_sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name = base_sheet_name

print("Final chosen sheet name:", sheet_name)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
   # writer.book = openpyxl.load_workbook(output_filename) 
    startrow = 3
    final_combined.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name}' tab.")


# --------------------------------------------------
# NEW ANALYSIS: Repeat for Wave 4
# --------------------------------------------------

# Convert relevant columns to numeric
for col in ["Q164", "Q166", "Q168", "Q170"]:
    df_wave4[col] = pd.to_numeric(df_wave4[col], errors='coerce')

# --------------------------------------------------
# Merge low and high anchors for each wave (not average, just pick available value)
# --------------------------------------------------
# Wave 2
df_wave2["Africa_Estimate"] = df_wave2["Q164"].combine_first(df_wave2["Q166"])
df_wave2["Tree_Estimate"] = df_wave2["Q168"].combine_first(df_wave2["Q170"])

# Wave 4
df_wave4["Africa_Estimate"] = df_wave4["Q164"].combine_first(df_wave4["Q166"])
df_wave4["Tree_Estimate"] = df_wave4["Q168"].combine_first(df_wave4["Q170"])

# --------------------------------------------------
# Merge Wave 2 and Wave 4 for matched respondents
# --------------------------------------------------
merged_africa = pd.merge(
    df_wave2[["PROLIFIC_PID", "Africa_Estimate"]],
    df_wave4[["PROLIFIC_PID", "Africa_Estimate"]],
    on="PROLIFIC_PID",
    suffixes=("_W2", "_W4")
).dropna()

merged_tree = pd.merge(
    df_wave2[["PROLIFIC_PID", "Tree_Estimate"]],
    df_wave4[["PROLIFIC_PID", "Tree_Estimate"]],
    on="PROLIFIC_PID",
    suffixes=("_W2", "_W4")
).dropna()


# --------------------------------------------------
# Descriptive statistics and t-tests for Wave 4
# --------------------------------------------------

# African countries Wave 4
africa_low_w4 = df_wave4.loc[df_wave4["Q164"].notna(), "Q164"]
africa_high_w4 = df_wave4.loc[df_wave4["Q166"].notna(), "Q166"]

desc_africa_low_w4 = {
    "n": africa_low_w4.shape[0],
    "Mean": africa_low_w4.mean(),
    "SD": africa_low_w4.std(ddof=1),
    "Median": africa_low_w4.median(),
    "Min": africa_low_w4.min(),
    "Max": africa_low_w4.max()
}
desc_africa_high_w4 = {
    "n": africa_high_w4.shape[0],
    "Mean": africa_high_w4.mean(),
    "SD": africa_high_w4.std(ddof=1),
    "Median": africa_high_w4.median(),
    "Min": africa_high_w4.min(),
    "Max": africa_high_w4.max()
}

if africa_low_w4.shape[0] > 1 and africa_high_w4.shape[0] > 1:
    t_stat_africa_w4, p_val_africa_w4 = ttest_ind(africa_high_w4, africa_low_w4, equal_var=False)
else:
    t_stat_africa_w4, p_val_africa_w4 = np.nan, np.nan

africa_summary_w4 = pd.DataFrame({
    "Domain": ["African Countries", "African Countries"],
    "Anchor": ["Low (12)", "High (65)"],
    "n": [desc_africa_low_w4["n"], desc_africa_high_w4["n"]],
    "Mean Estimate": [round(desc_africa_low_w4["Mean"], 2), round(desc_africa_high_w4["Mean"], 2)],
    "SD": [round(desc_africa_low_w4["SD"], 2), round(desc_africa_high_w4["SD"], 2)],
    "Median": [round(desc_africa_low_w4["Median"], 2), round(desc_africa_high_w4["Median"], 2)],
    "Min": [desc_africa_low_w4["Min"], desc_africa_high_w4["Min"]],
    "Max": [desc_africa_low_w4["Max"], desc_africa_high_w4["Max"]]
})
africa_ttest_w4 = pd.DataFrame({
    "Domain": ["African Countries"],
    "t-statistic": [round(t_stat_africa_w4, 3) if not np.isnan(t_stat_africa_w4) else np.nan],
    "p-value": [round(p_val_africa_w4, 3) if not np.isnan(p_val_africa_w4) else np.nan]
})

# Tree Wave 4
tree_low_w4 = df_wave4.loc[df_wave4["Q168"].notna(), "Q168"]
tree_high_w4 = df_wave4.loc[df_wave4["Q170"].notna(), "Q170"]

desc_tree_low_w4 = {
    "n": tree_low_w4.shape[0],
    "Mean": tree_low_w4.mean(),
    "SD": tree_low_w4.std(ddof=1),
    "Median": tree_low_w4.median(),
    "Min": tree_low_w4.min(),
    "Max": tree_low_w4.max()
}
desc_tree_high_w4 = {
    "n": tree_high_w4.shape[0],
    "Mean": tree_high_w4.mean(),
    "SD": tree_high_w4.std(ddof=1),
    "Median": tree_high_w4.median(),
    "Min": tree_high_w4.min(),
    "Max": tree_high_w4.max()
}

if tree_low_w4.shape[0] > 1 and tree_high_w4.shape[0] > 1:
    t_stat_tree_w4, p_val_tree_w4 = ttest_ind(tree_high_w4, tree_low_w4, equal_var=False)
else:
    t_stat_tree_w4, p_val_tree_w4 = np.nan, np.nan

tree_summary_w4 = pd.DataFrame({
    "Domain": ["Trees", "Trees"],
    "Anchor": ["Low (85 feet)", "High (1000 feet)"],
    "n": [desc_tree_low_w4["n"], desc_tree_high_w4["n"]],
    "Mean Estimate": [round(desc_tree_low_w4["Mean"], 2), round(desc_tree_high_w4["Mean"], 2)],
    "SD": [round(desc_tree_low_w4["SD"], 2), round(desc_tree_high_w4["SD"], 2)],
    "Median": [round(desc_tree_low_w4["Median"], 2), round(desc_tree_high_w4["Median"], 2)],
    "Min": [desc_tree_low_w4["Min"], desc_tree_high_w4["Min"]],
    "Max": [desc_tree_low_w4["Max"], desc_tree_high_w4["Max"]]
})
tree_ttest_w4 = pd.DataFrame({
    "Domain": ["Trees"],
    "t-statistic": [round(t_stat_tree_w4, 3) if not np.isnan(t_stat_tree_w4) else np.nan],
    "p-value": [round(p_val_tree_w4, 3) if not np.isnan(p_val_tree_w4) else np.nan]
})

# Final Wave 4 combined summary
final_combined_w4 = pd.concat([
    africa_summary_w4,
    pd.DataFrame([[""]*africa_summary_w4.shape[1]], columns=africa_summary_w4.columns),
    tree_summary_w4,
    pd.DataFrame([[""]*tree_summary_w4.shape[1]], columns=tree_summary_w4.columns),
    africa_ttest_w4,
    tree_ttest_w4
], ignore_index=True)



# --------------------------------------------------
# NEW ANALYSIS: Repeat for LLM (identified by TWIN_ID)
# --------------------------------------------------

# Convert columns to numeric
for col in ["Q164", "Q166", "Q168", "Q170"]:
    df_llm[col] = pd.to_numeric(df_llm[col], errors='coerce')

# African countries
africa_low_llm = df_llm.loc[df_llm["Q164"].notna(), "Q164"]
africa_high_llm = df_llm.loc[df_llm["Q166"].notna(), "Q166"]

desc_africa_low_llm = {
    "n": africa_low_llm.shape[0],
    "Mean": africa_low_llm.mean(),
    "SD": africa_low_llm.std(ddof=1),
    "Median": africa_low_llm.median(),
    "Min": africa_low_llm.min(),
    "Max": africa_low_llm.max()
}
desc_africa_high_llm = {
    "n": africa_high_llm.shape[0],
    "Mean": africa_high_llm.mean(),
    "SD": africa_high_llm.std(ddof=1),
    "Median": africa_high_llm.median(),
    "Min": africa_high_llm.min(),
    "Max": africa_high_llm.max()
}

if africa_low_llm.shape[0] > 1 and africa_high_llm.shape[0] > 1:
    t_stat_africa_llm, p_val_africa_llm = ttest_ind(africa_high_llm, africa_low_llm, equal_var=False)
else:
    t_stat_africa_llm, p_val_africa_llm = np.nan, np.nan

africa_summary_llm = pd.DataFrame({
    "Domain": ["African Countries", "African Countries"],
    "Anchor": ["Low (12)", "High (65)"],
    "n": [desc_africa_low_llm["n"], desc_africa_high_llm["n"]],
    "Mean Estimate": [round(desc_africa_low_llm["Mean"], 2), round(desc_africa_high_llm["Mean"], 2)],
    "SD": [round(desc_africa_low_llm["SD"], 2), round(desc_africa_high_llm["SD"], 2)],
    "Median": [round(desc_africa_low_llm["Median"], 2), round(desc_africa_high_llm["Median"], 2)],
    "Min": [desc_africa_low_llm["Min"], desc_africa_high_llm["Min"]],
    "Max": [desc_africa_low_llm["Max"], desc_africa_high_llm["Max"]]
})
africa_ttest_llm = pd.DataFrame({
    "Domain": ["African Countries"],
    "t-statistic": [round(t_stat_africa_llm, 3) if not np.isnan(t_stat_africa_llm) else np.nan],
    "p-value": [round(p_val_africa_llm, 3) if not np.isnan(p_val_africa_llm) else np.nan]
})

# Trees
tree_low_llm = df_llm.loc[df_llm["Q168"].notna(), "Q168"]
tree_high_llm = df_llm.loc[df_llm["Q170"].notna(), "Q170"]

desc_tree_low_llm = {
    "n": tree_low_llm.shape[0],
    "Mean": tree_low_llm.mean(),
    "SD": tree_low_llm.std(ddof=1),
    "Median": tree_low_llm.median(),
    "Min": tree_low_llm.min(),
    "Max": tree_low_llm.max()
}
desc_tree_high_llm = {
    "n": tree_high_llm.shape[0],
    "Mean": tree_high_llm.mean(),
    "SD": tree_high_llm.std(ddof=1),
    "Median": tree_high_llm.median(),
    "Min": tree_high_llm.min(),
    "Max": tree_high_llm.max()
}

if tree_low_llm.shape[0] > 1 and tree_high_llm.shape[0] > 1:
    t_stat_tree_llm, p_val_tree_llm = ttest_ind(tree_high_llm, tree_low_llm, equal_var=False)
else:
    t_stat_tree_llm, p_val_tree_llm = np.nan, np.nan

tree_summary_llm = pd.DataFrame({
    "Domain": ["Trees", "Trees"],
    "Anchor": ["Low (85 feet)", "High (1000 feet)"],
    "n": [desc_tree_low_llm["n"], desc_tree_high_llm["n"]],
    "Mean Estimate": [round(desc_tree_low_llm["Mean"], 2), round(desc_tree_high_llm["Mean"], 2)],
    "SD": [round(desc_tree_low_llm["SD"], 2), round(desc_tree_high_llm["SD"], 2)],
    "Median": [round(desc_tree_low_llm["Median"], 2), round(desc_tree_high_llm["Median"], 2)],
    "Min": [desc_tree_low_llm["Min"], desc_tree_high_llm["Min"]],
    "Max": [desc_tree_low_llm["Max"], desc_tree_high_llm["Max"]]
})
tree_ttest_llm = pd.DataFrame({
    "Domain": ["Trees"],
    "t-statistic": [round(t_stat_tree_llm, 3) if not np.isnan(t_stat_tree_llm) else np.nan],
    "p-value": [round(p_val_tree_llm, 3) if not np.isnan(p_val_tree_llm) else np.nan]
})

# Combine all LLM results
final_combined_llm = pd.concat([
    africa_summary_llm,
    pd.DataFrame([[""]*africa_summary_llm.shape[1]], columns=africa_summary_llm.columns),
    tree_summary_llm,
    pd.DataFrame([[""]*tree_summary_llm.shape[1]], columns=tree_summary_llm.columns),
    africa_ttest_llm,
    tree_ttest_llm
], ignore_index=True)

# Write to Excel (append to same sheet)
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = writer.sheets[sheet_name]
    startrow_llm = ws.max_row + 3
    ws.cell(row=startrow_llm - 1, column=1, value="LLM Summary Statistics and t-tests")
    final_combined_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow_llm, index=False)

print("✅ LLM descriptive stats and t-tests written to Excel.")



# --------------------------------------------------
# Write Wave 4 Descriptives to Excel
# --------------------------------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    # Re-access the sheet properly for writing cell notes
    wb = writer.book
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    
    # Determine where to start writing Wave 4
    startrow_new = ws.max_row + 3

    # Write label above the table
    ws.cell(row=startrow_new - 1, column=1, value="Wave 4 descriptive stats and t-tests")

    # Write the actual table
    final_combined_w4.to_excel(writer, sheet_name=sheet_name, startrow=startrow_new, index=False)


print("Writing Wave 4 summary starting at row:", startrow_new)
    
    # --------------------------------------------------
# Verify that PROLIFIC_PIDs kept the same anchor condition between Wave 2 and Wave 4 (Africa question)
# --------------------------------------------------

# Step 1: Define function to assign condition (Low or High)
def get_africa_condition(row, low_col, high_col):
    if pd.notna(row[low_col]) and pd.isna(row[high_col]):
        return "Low"
    elif pd.isna(row[low_col]) and pd.notna(row[high_col]):
        return "High"
    else:
        return np.nan  # Shouldn't happen ideally

# Step 2: Assign condition in both waves
df_wave2["Africa_Condition"] = df_wave2.apply(lambda row: get_africa_condition(row, "Q164", "Q166"), axis=1)
df_wave4["Africa_Condition"] = df_wave4.apply(lambda row: get_africa_condition(row, "Q164", "Q166"), axis=1)

# Step 3: Merge to compare
condition_check = pd.merge(
    df_wave2[["PROLIFIC_PID", "Africa_Condition"]],
    df_wave4[["PROLIFIC_PID", "Africa_Condition"]],
    on="PROLIFIC_PID",
    suffixes=("_W2", "_W4")
)

# Step 4: Identify mismatches
condition_mismatch = condition_check[
    (condition_check["Africa_Condition_W2"].notna()) &
    (condition_check["Africa_Condition_W4"].notna()) &
    (condition_check["Africa_Condition_W2"] != condition_check["Africa_Condition_W4"])
]

print(f"Number of participants with mismatched Africa condition between Wave 2 and Wave 4: {condition_mismatch.shape[0]}")
if not condition_mismatch.empty:
    print(condition_mismatch)
else:
    print("✅ All participants kept the same anchor condition (low vs high) across waves for Africa question.")

    
print("done")


# ## Absolute vs Relative Savings (Stanovich & West, 2008)

# In[82]:


import pandas as pd
import numpy as np
from scipy.stats import chi2_contingency
import openpyxl

# ---------------------------
# Map responses for Q183 and Q184.
# We assume responses should be either "Yes" or "No" (case-insensitive).
# ---------------------------
def map_yes_no(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).strip().lower()
    if "yes" in text:
        return "yes"
    elif "no" in text:
        return "no"
    else:
        return np.nan

df_wave2["Q183_mapped"] = df_wave2["Q183"].apply(map_yes_no)
df_wave2["Q184_mapped"] = df_wave2["Q184"].apply(map_yes_no)

# ---------------------------
# Ensure each subject appears in only one condition.
# Since the design is between-subjects, drop subjects that answered both or neither.
# ---------------------------
mask = ((df_wave2["Q183_mapped"].notna()) ^ (df_wave2["Q184_mapped"].notna()))
df_wave2 = df_wave2[mask]

# ---------------------------
# Now, define the two groups.
# ---------------------------
group_calc = df_wave2.loc[df_wave2["Q183_mapped"].notna(), "Q183_mapped"]
group_jacket = df_wave2.loc[df_wave2["Q184_mapped"].notna(), "Q184_mapped"]

# ---------------------------
# Compute frequency counts for each condition.
# ---------------------------
def freq_table(series):
    counts = series.value_counts(dropna=True).sort_index()
    percents = series.value_counts(normalize=True, dropna=True).sort_index() * 100
    return pd.DataFrame({
        "Response": counts.index,
        "Count": counts.values,
        "Percent": percents.round(1).values
    })

freq_calc = freq_table(group_calc)
freq_calc["Condition"] = "Calculator (Q183)"
freq_jacket = freq_table(group_jacket)
freq_jacket["Condition"] = "Jacket (Q184)"

freq_combined = pd.concat([freq_calc, freq_jacket], ignore_index=True)
freq_combined = freq_combined[["Condition", "Response", "Count", "Percent"]]

# ---------------------------
# Build a 2x2 contingency table for yes/no responses.
# ---------------------------
def get_yes_no_counts(series):
    series = series.str.lower().str.strip()
    yes = (series == "yes").sum()
    no = (series == "no").sum()
    return yes, no

yes_calc, no_calc = get_yes_no_counts(group_calc)
yes_jacket, no_jacket = get_yes_no_counts(group_jacket)

contingency = np.array([[yes_calc, no_calc],
                         [yes_jacket, no_jacket]])

# ---------------------------
# Perform chi-square test.
# ---------------------------
chi2, p_chi2, dof, expected = chi2_contingency(contingency)
chi2_results = pd.DataFrame({
    "Chi2": [round(chi2, 3)],
    "Degrees of Freedom": [dof],
    "p-value": [round(p_chi2, 3)]
})

# ---------------------------
# Prepare descriptive summary table for both conditions.
# ---------------------------
desc_table = pd.DataFrame({
    "Condition": ["Calculator (Q183)", "Jacket (Q184)"],
    "n": [group_calc.shape[0], group_jacket.shape[0]],
    "Yes (%)": [round(yes_calc/(yes_calc+no_calc)*100,1) if (yes_calc+no_calc) > 0 else np.nan,
                round(yes_jacket/(yes_jacket+no_jacket)*100,1) if (yes_jacket+no_jacket) > 0 else np.nan],
    "No (%)": [round(no_calc/(yes_calc+no_calc)*100,1) if (yes_calc+no_calc) > 0 else np.nan,
               round(no_jacket/(yes_jacket+no_jacket)*100,1) if (yes_jacket+no_jacket) > 0 else np.nan]
})

# ---------------------------
# Combine outputs into one final DataFrame.
# We'll stack the descriptive table, then a blank row, then the frequency table, then a blank row, then the chi-square table.
# ---------------------------
blank_row_desc = pd.DataFrame([[""] * desc_table.shape[1]], columns=desc_table.columns)
blank_row_freq = pd.DataFrame([[""] * freq_combined.shape[1]], columns=freq_combined.columns)
final_combined = pd.concat([desc_table, blank_row_desc, freq_combined, blank_row_freq, chi2_results], ignore_index=True)

# ---------------------------
# Prepare header note.
# ---------------------------
header_note = (
    "Analysis: Relative vs. Absolute Savings \n\n"
    "Two scenarios were presented to participants:\n\n"
    "1. Calculator scenario (Q183):\n"
    "   - You go to purchase a calculator for $30. The salesperson informs you that the calculator is on sale for $20 at the other branch (a $10 saving).\n\n"
    "2. Jacket scenario (Q184):\n"
    "   - You go to purchase a jacket for $250. The salesperson informs you that the jacket is on sale for $240 at the other branch (a $10 saving).\n\n"
    "The absolute saving is the same (at $10), but we expect people will more likely drive for the calculator than for the jacket.\n\n"
    "Chi-square test was used to compare binary categorical choices."
)

# ---------------------------
# Determine a unique new sheet name and write all output in one go.
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Relative Savings"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

print("Final chosen sheet name:", sheet_name_new)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    final_combined.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")


# --------------------------------------------------
# NEW ANALYSIS: Repeat for Wave 4
# --------------------------------------------------

# Map responses for Q183 and Q184
df_wave4["Q183_mapped"] = df_wave4["Q183"].apply(map_yes_no)
df_wave4["Q184_mapped"] = df_wave4["Q184"].apply(map_yes_no)

# Ensure each subject appears in only one condition
mask_w4 = ((df_wave4["Q183_mapped"].notna()) ^ (df_wave4["Q184_mapped"].notna()))
df_wave4 = df_wave4[mask_w4]

# Define groups for Wave 4
group_calc_w4 = df_wave4.loc[df_wave4["Q183_mapped"].notna(), "Q183_mapped"]
group_jacket_w4 = df_wave4.loc[df_wave4["Q184_mapped"].notna(), "Q184_mapped"]

# Compute frequency counts for Wave 4
freq_calc_w4 = freq_table(group_calc_w4)
freq_calc_w4["Condition"] = "Calculator (Q183) - W4"
freq_jacket_w4 = freq_table(group_jacket_w4)
freq_jacket_w4["Condition"] = "Jacket (Q184) - W4"

freq_combined_w4 = pd.concat([freq_calc_w4, freq_jacket_w4], ignore_index=True)
freq_combined_w4 = freq_combined_w4[["Condition", "Response", "Count", "Percent"]]

# 2x2 contingency table for Wave 4
yes_calc_w4, no_calc_w4 = get_yes_no_counts(group_calc_w4)
yes_jacket_w4, no_jacket_w4 = get_yes_no_counts(group_jacket_w4)

contingency_w4 = np.array([[yes_calc_w4, no_calc_w4],
                           [yes_jacket_w4, no_jacket_w4]])

# Chi-square test for Wave 4
chi2_w4, p_chi2_w4, dof_w4, expected_w4 = chi2_contingency(contingency_w4)
chi2_results_w4 = pd.DataFrame({
    "Chi2": [round(chi2_w4, 3)],
    "Degrees of Freedom": [dof_w4],
    "p-value": [round(p_chi2_w4, 3)]
})

# Descriptive table for Wave 4
desc_table_w4 = pd.DataFrame({
    "Condition": ["Calculator (Q183) - W4", "Jacket (Q184) - W4"],
    "n": [group_calc_w4.shape[0], group_jacket_w4.shape[0]],
    "Yes (%)": [round(yes_calc_w4/(yes_calc_w4+no_calc_w4)*100,1) if (yes_calc_w4+no_calc_w4) > 0 else np.nan,
                round(yes_jacket_w4/(yes_jacket_w4+no_jacket_w4)*100,1) if (yes_jacket_w4+no_jacket_w4) > 0 else np.nan],
    "No (%)": [round(no_calc_w4/(yes_calc_w4+no_calc_w4)*100,1) if (yes_calc_w4+no_calc_w4) > 0 else np.nan,
               round(no_jacket_w4/(yes_jacket_w4+no_jacket_w4)*100,1) if (yes_jacket_w4+no_jacket_w4) > 0 else np.nan]
})

# Final combined Wave 4 table
blank_row_desc_w4 = pd.DataFrame([[""] * desc_table_w4.shape[1]], columns=desc_table_w4.columns)
blank_row_freq_w4 = pd.DataFrame([[""] * freq_combined_w4.shape[1]], columns=freq_combined_w4.columns)
final_combined_w4 = pd.concat([desc_table_w4, blank_row_desc_w4, freq_combined_w4, blank_row_freq_w4, chi2_results_w4], ignore_index=True)


# ---------------------------
# NEW ANALYSIS: Repeat for LLM data (identified by TWIN_ID)
# ---------------------------

df_llm["Q183_mapped"] = df_llm["Q183"].apply(map_yes_no)
df_llm["Q184_mapped"] = df_llm["Q184"].apply(map_yes_no)

mask_llm = ((df_llm["Q183_mapped"].notna()) ^ (df_llm["Q184_mapped"].notna()))
df_llm = df_llm[mask_llm]

group_calc_llm = df_llm.loc[df_llm["Q183_mapped"].notna(), "Q183_mapped"]
group_jacket_llm = df_llm.loc[df_llm["Q184_mapped"].notna(), "Q184_mapped"]

freq_calc_llm = freq_table(group_calc_llm)
freq_calc_llm["Condition"] = "Calculator (Q183) - LLM"
freq_jacket_llm = freq_table(group_jacket_llm)
freq_jacket_llm["Condition"] = "Jacket (Q184) - LLM"

freq_combined_llm = pd.concat([freq_calc_llm, freq_jacket_llm], ignore_index=True)
freq_combined_llm = freq_combined_llm[["Condition", "Response", "Count", "Percent"]]

yes_calc_llm, no_calc_llm = get_yes_no_counts(group_calc_llm)
yes_jacket_llm, no_jacket_llm = get_yes_no_counts(group_jacket_llm)

contingency_llm = np.array([[yes_calc_llm, no_calc_llm],
                            [yes_jacket_llm, no_jacket_llm]])

chi2_llm, p_chi2_llm, dof_llm, expected_llm = chi2_contingency(contingency_llm)
chi2_results_llm = pd.DataFrame({
    "Chi2": [round(chi2_llm, 3)],
    "Degrees of Freedom": [dof_llm],
    "p-value": [round(p_chi2_llm, 3)]
})

desc_table_llm = pd.DataFrame({
    "Condition": ["Calculator (Q183) - LLM", "Jacket (Q184) - LLM"],
    "n": [group_calc_llm.shape[0], group_jacket_llm.shape[0]],
    "Yes (%)": [round(yes_calc_llm / (yes_calc_llm + no_calc_llm) * 100, 1) if (yes_calc_llm + no_calc_llm) > 0 else np.nan,
                round(yes_jacket_llm / (yes_jacket_llm + no_jacket_llm) * 100, 1) if (yes_jacket_llm + no_jacket_llm) > 0 else np.nan],
    "No (%)": [round(no_calc_llm / (yes_calc_llm + no_calc_llm) * 100, 1) if (yes_calc_llm + no_calc_llm) > 0 else np.nan,
               round(no_jacket_llm / (yes_jacket_llm + no_jacket_llm) * 100, 1) if (yes_jacket_llm + no_jacket_llm) > 0 else np.nan]
})

# Combine and write to Excel
blank_row_llm = pd.DataFrame([[""] * desc_table_llm.shape[1]], columns=desc_table_llm.columns)
final_combined_llm = pd.concat([desc_table_llm, blank_row_llm, freq_combined_llm, blank_row_llm, chi2_results_llm], ignore_index=True)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 3
    ws.cell(row=startrow - 1, column=1, value="LLM Descriptive Statistics and Chi-Square Test")
    final_combined_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)



# --------------------------------------------------
# Write everything into the same Excel sheet
# --------------------------------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = wb[sheet_name_new]

    startrow_new = startrow + final_combined.shape[0] + 6  # after Wave 2 results

    # Write label above Wave 4 table
    ws.cell(row=startrow_new - 1, column=1, value="Wave 4 Descriptive Statistics and Chi-Square Test")

    # Write Wave 4 table
    final_combined_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow_new, index=False)

#     # Correlation results
#     startrow_corr = startrow_new + final_combined_w4.shape[0] + 6
#     correlation_results.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow_corr, index=False)

print(f"Wave 4 analysis added to '{sheet_name_new}' in '{output_filename}'.")


print("done")


# ## Myside bias (Stanovich & West, 2008)

# In[83]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl


# ---------------------------
# Map responses numerically
# ---------------------------
# ---------------------------
# Correct mapping from text responses to numerical Likert scores
# ---------------------------
def map_likert_scale(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).strip().lower()
    if "definitely no" in text:
        return 1
    elif "probably no" in text:
        return 3
    elif text == "no":
        return 2
    elif "probably yes" in text:
        return 4
    elif "definitely yes" in text:
        return 6
    elif text == "yes":
        return 5
    else:
        return np.nan
    
df_wave2["Q194_mapped"] = df_wave2["Q194"].apply(map_likert_scale)
df_wave2["Q195_mapped"] = df_wave2["Q195"].apply(map_likert_scale)

# Sort so rows with non-missing Q194 come first
df_wave2 = df_wave2.sort_values(by="Q194_mapped", na_position="last")
# Then drop duplicates, keeping the non-missing response if available
df_wave2 = df_wave2.drop_duplicates(subset="PROLIFIC_PID", keep="first")

# ---------------------------
# Ensure each subject appears in only one condition
# ---------------------------
mask = (df_wave2["Q194_mapped"].notna()) ^ (df_wave2["Q195_mapped"].notna())
df_wave2 = df_wave2[mask]

# ---------------------------
# Define groups
# ---------------------------
group_germany = df_wave2.loc[df_wave2["Q194_mapped"].notna(), "Q194_mapped"]
group_us = df_wave2.loc[df_wave2["Q195_mapped"].notna(), "Q195_mapped"]

# ---------------------------
# Descriptive statistics
# ---------------------------
desc_table = pd.DataFrame({
    "Condition": ["Ford (Q194)", "German (Q195)"],
    "n": [group_germany.shape[0], group_us.shape[0]],
    "Mean": [round(group_germany.mean(), 2), round(group_us.mean(), 2)],
    "SD": [round(group_germany.std(ddof=1), 2), round(group_us.std(ddof=1), 2)],
    "Median": [round(group_germany.median(), 2), round(group_us.median(), 2)],
    "Min": [group_germany.min(), group_us.min()],
    "Max": [group_germany.max(), group_us.max()]
})

# ---------------------------
# Independent samples t-test
# ---------------------------
if group_germany.shape[0] > 1 and group_us.shape[0] > 1:
    t_stat, p_val = ttest_ind(group_germany, group_us, equal_var=False)
else:
    t_stat, p_val = np.nan, np.nan

t_test_table = pd.DataFrame({
    "Statistic": ["t-statistic", "p-value"],
    "Value": [round(t_stat, 3) if not np.isnan(t_stat) else np.nan,
              round(p_val, 3) if not np.isnan(p_val) else np.nan]
})

# ---------------------------
# Prepare header note
# ---------------------------
header_note = (
    "Analysis: Myside Bias \n\n"
    "Two versions of the question were administered between subjects:\n\n"
    "Version 1 (Q194): Ford Explorer ban question in Germany.\n"
    "Version 2 (Q195): German car ban question in U.S.\n\n"
    "Responses were on a 6-point Likert scale:\n"
    "1 = Definitely No, 2 = No, 3 = Probably No, 4 = Probably Yes, 5 = Yes, 6 = Definitely Yes.\n\n"
    "Independent samples t-test compares the mean responses between conditions."
)

# ---------------------------
# Write everything cleanly to Excel
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Myside"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

print("Final chosen sheet name:", sheet_name_new)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    # Write header
    ws = writer.book.create_sheet(sheet_name_new)
    ws.cell(row=1, column=1, value=header_note)
    
    # Write descriptive statistics
    desc_table.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)

    # Leave a gap and write t-test results
    startrow_ttest = startrow + len(desc_table) + 4
    t_test_table.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow_ttest, index=False)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")

# --------------------------------------------------
# NEW ANALYSIS: Wave 4 
# --------------------------------------------------

df_wave4 = df_wave4.drop_duplicates(subset="PROLIFIC_PID")

# Map responses numerically for Wave 4
df_wave4["Q194_mapped"] = df_wave4["Q194"].apply(map_likert_scale)
df_wave4["Q195_mapped"] = df_wave4["Q195"].apply(map_likert_scale)

# Ensure each subject appears in only one condition
mask_wave4 = (df_wave4["Q194_mapped"].notna()) ^ (df_wave4["Q195_mapped"].notna())
df_wave4 = df_wave4[mask_wave4]

mapped_count = df_wave4["Q194_mapped"].notna().sum()
print("Mapped Q194 participants (Wave 4):", mapped_count)

# Define groups for Wave 4
group_germany_w4 = df_wave4.loc[df_wave4["Q194_mapped"].notna(), "Q194_mapped"]
group_us_w4 = df_wave4.loc[df_wave4["Q195_mapped"].notna(), "Q195_mapped"]

# Descriptive statistics for Wave 4
desc_table_w4 = pd.DataFrame({
    "Condition": ["Ford (Q194) - W4", "German (Q195) - W4"],
    "n": [group_germany_w4.shape[0], group_us_w4.shape[0]],
    "Mean": [round(group_germany_w4.mean(), 2), round(group_us_w4.mean(), 2)],
    "SD": [round(group_germany_w4.std(ddof=1), 2), round(group_us_w4.std(ddof=1), 2)],
    "Median": [round(group_germany_w4.median(), 2), round(group_us_w4.median(), 2)],
    "Min": [group_germany_w4.min(), group_us_w4.min()],
    "Max": [group_germany_w4.max(), group_us_w4.max()]
})

# T-test for Wave 4
if group_germany_w4.shape[0] > 1 and group_us_w4.shape[0] > 1:
    t_stat_w4, p_val_w4 = ttest_ind(group_germany_w4, group_us_w4, equal_var=False)
else:
    t_stat_w4, p_val_w4 = np.nan, np.nan

t_test_table_w4 = pd.DataFrame({
    "Statistic": ["t-statistic", "p-value"],
    "Value": [round(t_stat_w4, 3) if not np.isnan(t_stat_w4) else np.nan,
              round(p_val_w4, 3) if not np.isnan(p_val_w4) else np.nan]
})


# --------------------------------------------------
# NEW ANALYSIS: LLM Data
# --------------------------------------------------

# Map responses numerically for LLM
df_llm["Q194_mapped"] = df_llm["Q194"].apply(map_likert_scale)
df_llm["Q195_mapped"] = df_llm["Q195"].apply(map_likert_scale)

# Ensure each subject appears in only one condition
mask_llm = (df_llm["Q194_mapped"].notna()) ^ (df_llm["Q195_mapped"].notna())
df_llm = df_llm[mask_llm].drop_duplicates(subset="TWIN_ID")

# Define groups for LLM
group_germany_llm = df_llm.loc[df_llm["Q194_mapped"].notna(), "Q194_mapped"]
group_us_llm = df_llm.loc[df_llm["Q195_mapped"].notna(), "Q195_mapped"]

# Descriptive statistics for LLM
llm_desc_table = pd.DataFrame({
    "Condition": ["Ford (Q194) - LLM", "German (Q195) - LLM"],
    "n": [group_germany_llm.shape[0], group_us_llm.shape[0]],
    "Mean": [round(group_germany_llm.mean(), 2), round(group_us_llm.mean(), 2)],
    "SD": [round(group_germany_llm.std(ddof=1), 2), round(group_us_llm.std(ddof=1), 2)],
    "Median": [round(group_germany_llm.median(), 2), round(group_us_llm.median(), 2)],
    "Min": [group_germany_llm.min(), group_us_llm.min()],
    "Max": [group_germany_llm.max(), group_us_llm.max()]
})

# T-test for LLM
if group_germany_llm.shape[0] > 1 and group_us_llm.shape[0] > 1:
    t_stat_llm, p_val_llm = ttest_ind(group_germany_llm, group_us_llm, equal_var=False)
else:
    t_stat_llm, p_val_llm = np.nan, np.nan

t_test_table_llm = pd.DataFrame({
    "Statistic": ["t-statistic", "p-value"],
    "Value": [round(t_stat_llm, 3) if not np.isnan(t_stat_llm) else np.nan,
              round(p_val_llm, 3) if not np.isnan(p_val_llm) else np.nan]
})

# Write LLM results to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = wb[sheet_name_new]

    llm_row = ws.max_row + 3
    ws.cell(row=llm_row - 1, column=1, value="LLM Descriptive Statistics and t-test")

    # Write descriptive table
    llm_desc_table.to_excel(writer, sheet_name=sheet_name_new, startrow=llm_row, index=False)
    
    # Write t-test
    llm_row += len(llm_desc_table) + 3
    t_test_table_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=llm_row, index=False)


# --------------------------------------------------
# Write Wave 4 results into Excel
# --------------------------------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = wb[sheet_name_new]

    # Find the next empty row
    nextrow = ws.max_row + 4

    # Write the title
    ws.cell(row=nextrow, column=1, value="Wave 4 descriptive statistics and t-test").font = Font(bold=True)
    
    # Write Wave 4 descriptive statistics
    nextrow += 2
    desc_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)
    
    nextrow += len(desc_table_w4) + 3
    
    # Write Wave 4 t-test
    t_test_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)

    nextrow += len(t_test_table_w4) + 3

print(f"✅ Wave 4 analysis successfully added to '{sheet_name_new}' in '{output_filename}'.")


# ---------------------------
# Write distribution of CONVERTED responses (Q194_mapped and Q195_mapped) for Wave 2 and Wave 4
# ---------------------------

# Build distribution tables from converted columns
def build_converted_distribution(df, wave_label):
    dist_list = []
    for q in ["Q194_mapped", "Q195_mapped"]:
        dist = df[q].value_counts(dropna=False).sort_index()
        dist_df = pd.DataFrame({
            "Wave": wave_label,
            "Question": q,
            "Converted Response": dist.index,
            "Count": dist.values
        })
        dist_list.append(dist_df)
    return pd.concat(dist_list, ignore_index=True)

dist_conv_w2 = build_converted_distribution(df_wave2, "Wave 2")
dist_conv_llm = build_converted_distribution(df_llm, "LLM")
dist_conv_w4 = build_converted_distribution(df_wave4, "Wave 4")

converted_dist_df = pd.concat([dist_conv_w2,dist_conv_llm, dist_conv_w4], ignore_index=True)

# Append to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = wb[sheet_name_new]
    
    # Determine where to write (leave space after correlation block)
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="Converted Response Distributions for Q194 and Q195")

    startrow=nextrow+10
    converted_dist_df.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)

print("Converted response distributions for Q194 and Q195 (Wave 2 and 4) written to Excel.")

print("done")


# ## Omission bias (Stanovich & West, 2008)

# In[84]:


import pandas as pd
import numpy as np
from scipy.stats import binomtest
import openpyxl



df_wave2_raw=df_wave2

# ---------------------------
# Omission Bias question is in column "Omission bias " (with trailing whitespace)
# ---------------------------
omission_col = "Omission bias "

# ---------------------------
# Define mapping for the 4-point Likert responses.
# Mapping: 1 = "I would definitely not...", 2 = "I would probably not...",
#          3 = "I would probably...", 4 = "I would definitely..."
# ---------------------------
def map_omission(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    if "definitely not" in text:
        return 1
    elif "probably not" in text:
        return 2
    elif "definitely" in text:
        return 4
    elif "probably" in text:
        return 3
    else:
        return np.nan

df_wave2["Omission_mapped"] = df_wave2[omission_col].apply(map_omission)

# ---------------------------
# Create binary grouping: group 1 = responses 1 or 2; group 2 = responses 3 or 4.
# ---------------------------
def assign_group(x):
    if pd.isna(x):
        return np.nan
    if x in [1, 2]:
        return "Low Vaccination"
    elif x in [3, 4]:
        return "High Vaccination"
    else:
        return np.nan

df_wave2["Omission_Group"] = df_wave2["Omission_mapped"].apply(assign_group)

# Add temporary column to prioritize rows with valid responses
df_wave2["has_response"] = df_wave2["Omission_mapped"].notna().astype(int)

# Sort so valid (non-NaN) responses are first
df_wave2 = df_wave2.sort_values(by="Omission_mapped", na_position="last")

# Drop duplicates — this will keep the non-NaN version if it exists
df_wave2 = df_wave2.drop_duplicates(subset="PROLIFIC_PID", keep="first")

# Now filter to keep only those with valid responses
df_wave2 = df_wave2[df_wave2["Omission_mapped"].notna()]

# ---------------------------
# Compute frequency counts for the two groups.
# ---------------------------
freq_table = df_wave2["Omission_Group"].value_counts(dropna=True).sort_index().reset_index()
freq_table.columns = ["Group", "Count"]
freq_table["Percent"] = (freq_table["Count"] / freq_table["Count"].sum() * 100).round(1)

# ---------------------------
# Compute 95% confidence intervals for proportions
# ---------------------------
from statsmodels.stats.proportion import proportion_confint

n_total = df_wave2.shape[0]
n_low = df_wave2[df_wave2["Omission_Group"] == "Low Vaccination"].shape[0]
n_high = df_wave2[df_wave2["Omission_Group"] == "High Vaccination"].shape[0]

low_prop = n_low / n_total
high_prop = n_high / n_total

low_ci_low, low_ci_high = proportion_confint(count=n_low, nobs=n_total, alpha=0.05, method="wilson")
high_ci_low, high_ci_high = proportion_confint(count=n_high, nobs=n_total, alpha=0.05, method="wilson")

test_results = pd.DataFrame({
    "Group": ["Low Vaccination", "High Vaccination"],
    "Proportion": [round(low_prop, 3), round(high_prop, 3)],
    "95% CI Lower": [round(low_ci_low, 3), round(high_ci_low, 3)],
    "95% CI Upper": [round(low_ci_high, 3), round(high_ci_high, 3)],
    "n": [n_low, n_high]
})

# ---------------------------
# Prepare a descriptive summary table.
# ---------------------------
desc_table = pd.DataFrame({
    "Group": ["Low Vaccination (1-2)", "High Vaccination (3-4)"],
    "n": [df_wave2[df_wave2["Omission_Group"]=="Low Vaccination"].shape[0],
          df_wave2[df_wave2["Omission_Group"]=="High Vaccination"].shape[0]]
})

# ---------------------------
# Combine outputs into one final DataFrame.
# We'll stack the descriptive summary table, a blank row, the frequency table, a blank row, then the test results.
# ---------------------------
blank_row_desc = pd.DataFrame([[""] * desc_table.shape[1]], columns=desc_table.columns)
blank_row_freq = pd.DataFrame([[""] * freq_table.shape[1]], columns=freq_table.columns)
final_output = pd.concat([desc_table, blank_row_desc, freq_table, blank_row_freq, test_results], ignore_index=True)

# ---------------------------
# Prepare header note.
# ---------------------------
header_note = (
    "Analysis: Omission Bias \n\n"
    "Scenario:\n"
    "  There will be a deadly flu in your area next winter. Your doctor says you have a 10% chance of dying from the flu.\n"
    "  A new flu vaccine is available that prevents you from catching the flu but carries a 5% risk of death from a weaker flu virus.\n\n"
    "Participants answered the question on a 4-point Likert scale:\n"
    "  'I would definitely not...', 'I would probably not...', 'I would probably...', 'I would definitely...'\n\n"
    "Per Stanovich & West, we then bin subjects into two categories:\n"
    "  - Low Vaccination: responses 1 or 2\n"
    "  - High Vaccination: responses 3 or 4\n\n"
    "We compute 95% confidence intervals for the proportion of responses in each group."
)

# ---------------------------
# Determine a unique new sheet name and write all output in one go.
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W2-Omission"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

print("Final chosen sheet name:", sheet_name_new)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    final_output.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")

# --------------------------------------------------
# Wave 4:
# --------------------------------------------------
#print("df_wave4 columns:", list(df_wave4.columns))
omission_col_w4 = "Omission bias"
# Apply mapping to omission column in wave 4
df_wave4["Omission_mapped"] = df_wave4[omission_col_w4].apply(map_omission)
df_wave4["Omission_Group"] = df_wave4["Omission_mapped"].apply(assign_group)
df_wave4 = df_wave4[df_wave4["Omission_mapped"].notna()]

# Frequency table for wave 4
freq_table_w4 = df_wave4["Omission_Group"].value_counts(dropna=True).sort_index().reset_index()
freq_table_w4.columns = ["Group", "Count"]
freq_table_w4["Percent"] = (freq_table_w4["Count"] / freq_table_w4["Count"].sum() * 100).round(1)

# Proportions and 95% CI
n_total_w4 = df_wave4.shape[0]
n_low_w4 = df_wave4[df_wave4["Omission_Group"] == "Low Vaccination"].shape[0]
n_high_w4 = df_wave4[df_wave4["Omission_Group"] == "High Vaccination"].shape[0]

low_prop_w4 = n_low_w4 / n_total_w4
high_prop_w4 = n_high_w4 / n_total_w4

low_ci_low_w4, low_ci_high_w4 = proportion_confint(n_low_w4, n_total_w4, alpha=0.05, method="wilson")
high_ci_low_w4, high_ci_high_w4 = proportion_confint(n_high_w4, n_total_w4, alpha=0.05, method="wilson")

test_results_w4 = pd.DataFrame({
    "Group": ["Low Vaccination", "High Vaccination"],
    "Proportion": [round(low_prop_w4, 3), round(high_prop_w4, 3)],
    "95% CI Lower": [round(low_ci_low_w4, 3), round(high_ci_low_w4, 3)],
    "95% CI Upper": [round(low_ci_high_w4, 3), round(high_ci_high_w4, 3)],
    "n": [n_low_w4, n_high_w4]
})


# --------------------------------------------------
# LLM Analysis (identified by TWIN_ID)
# --------------------------------------------------
omission_col_llm = "Omission bias"
df_llm["Omission_mapped"] = df_llm[omission_col_llm].apply(map_omission)
df_llm["Omission_Group"] = df_llm["Omission_mapped"].apply(assign_group)
df_llm = df_llm[df_llm["Omission_mapped"].notna()]

# Frequency table for LLM
freq_table_llm = df_llm["Omission_Group"].value_counts(dropna=True).sort_index().reset_index()
freq_table_llm.columns = ["Group", "Count"]
freq_table_llm["Percent"] = (freq_table_llm["Count"] / freq_table_llm["Count"].sum() * 100).round(1)

# Proportions and 95% CI
n_total_llm = df_llm.shape[0]
n_low_llm = df_llm[df_llm["Omission_Group"] == "Low Vaccination"].shape[0]
n_high_llm = df_llm[df_llm["Omission_Group"] == "High Vaccination"].shape[0]

low_prop_llm = n_low_llm / n_total_llm
high_prop_llm = n_high_llm / n_total_llm

low_ci_low_llm, low_ci_high_llm = proportion_confint(n_low_llm, n_total_llm, alpha=0.05, method="wilson")
high_ci_low_llm, high_ci_high_llm = proportion_confint(n_high_llm, n_total_llm, alpha=0.05, method="wilson")

test_results_llm = pd.DataFrame({
    "Group": ["Low Vaccination", "High Vaccination"],
    "Proportion": [round(low_prop_llm, 3), round(high_prop_llm, 3)],
    "95% CI Lower": [round(low_ci_low_llm, 3), round(high_ci_low_llm, 3)],
    "95% CI Upper": [round(low_ci_high_llm, 3), round(high_ci_high_llm, 3)],
    "n": [n_low_llm, n_high_llm]
})

# Append LLM results to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = writer.sheets[sheet_name_new]
    
    nextrow = ws.max_row + 4
    ws.cell(row=nextrow, column=1, value="LLM Descriptive Statistics and Confidence Intervals")

    nextrow += 2
    freq_table_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)

    nextrow += len(freq_table_llm) + 3
    test_results_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)

print("✅ LLM omission bias results successfully written to Excel.")

# --------------------------------------------------
# Append Wave 4 results to Excel
# --------------------------------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    wb = writer.book
    ws = writer.sheets[sheet_name_new]

    # Find the next empty row
    nextrow = ws.max_row + 4

    # Write label above the table
    ws.cell(row=nextrow, column=1, value="Wave 4 Descriptive Statistics and Confidence Intervals")
    nextrow += 2
    
    # Write Wave 4 frequency table
    freq_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)
    nextrow += len(freq_table_w4) + 3

    # Write Wave 4 proportions and CI
    test_results_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)
    nextrow += len(test_results_w4) + 3

#     # Write correlation results
#     correlation_table.to_excel(writer, sheet_name=sheet_name_new, startrow=nextrow, index=False)

print("✅ Wave 4 omission bias analysis successfully written to Excel.")


df_wave2=df_wave2_raw

print("done")


# # Wave 3

# ## Less is More (Stanovich & West, 2008)

# In[85]:


# --- Modified full analysis code to write all results into one tab ---
import pandas as pd
import numpy as np
from scipy.stats import f_oneway, pearsonr
import openpyxl
import random
from openpyxl.styles import Font
from tqdm import tqdm 


# ---------------------------
# Define column mappings
# ---------------------------
formA_cols = ["Q171", "Q174", "Q177"]
formB_cols = ["Q172", "Q175", "Q178"]
formC_cols = ["Q173", "Q176", "Q179"]

# ---------------------------
# Map responses to numeric values
# ---------------------------
def map_less_is_more(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).strip().lower()

    mapping = {
        "strongly disagree": 1,
        "disagree strongly": 1,
        "very unlikely": 1,
        "disagree a little": 2,
        "moderately disagree": 2,
        "somewhat unlikely": 2,
        "neither agree nor disagree": 3,
        "neutral": 3,
        "slightly disagree": 3,
        "agree a little": 4,
        "slightly agree": 4,
        "somewhat likely": 4,
        "agree strongly": 5,
        "moderately agree": 5,    
        "strongly agree": 6,
        "very likely": 6,
        "extremely likely": 6,
        "likely": 5,
    }

    for key, val in mapping.items():
        if key in text:
            return val
    return np.nan  # fallback

for col in formA_cols + formB_cols + formC_cols:
    new_col = col + "_num"
    df_wave3[new_col] = df_wave3[col].apply(map_less_is_more)

# ---------------------------
# Form assignment
# ---------------------------
def get_form_response(df, cols):
    sub_df = df[[col + "_num" for col in cols]]
    return sub_df.mean(axis=1)  # DO NOT drop rows with partial missingness


df_wave3["FormA_avg"] = get_form_response(df_wave3, formA_cols)
df_wave3["FormB_avg"] = get_form_response(df_wave3, formB_cols)
df_wave3["FormC_avg"] = get_form_response(df_wave3, formC_cols)

def assign_form(row):
    if not pd.isna(row["Q171_num"]):
        return "Form A"
    elif not pd.isna(row["Q172_num"]):
        return "Form B"
    elif not pd.isna(row["Q173_num"]):
        return "Form C"
    else:
        return np.nan

df_wave3["Form_Assignment"] = df_wave3.apply(assign_form, axis=1)
df_wave3_valid = df_wave3[df_wave3["Form_Assignment"].notna()]


# ---------------------------
# Analyze each question separately
# ---------------------------
def get_question_scores(df, form_cols):
    return {
        "Form A": df.loc[df["Form_Assignment"] == "Form A", form_cols["Form A"]],
        "Form B": df.loc[df["Form_Assignment"] == "Form B", form_cols["Form B"]],
        "Form C": df.loc[df["Form_Assignment"] == "Form C", form_cols["Form C"]]
    }

question_mappings = {
    "Gamble": {"Form A": "Q171_num", "Form B": "Q172_num", "Form C": "Q173_num"},
    "Proportion Dominance 1": {"Form A": "Q174_num", "Form B": "Q175_num", "Form C": "Q176_num"},
    "Proportion Dominance 2": {"Form A": "Q177_num", "Form B": "Q178_num", "Form C": "Q179_num"}
}

# Collect all results
all_output = []

for question, mapping in question_mappings.items():
    scores = get_question_scores(df_wave3_valid, mapping)
    desc_stats = pd.DataFrame({
        "Form": list(scores.keys()),
        "n": [s.size for s in scores.values()],
        "Mean Score": [round(s.mean(), 2) for s in scores.values()],
        "SD": [round(s.std(ddof=1), 2) for s in scores.values()],
        "Median": [round(s.median(), 2) for s in scores.values()],
        "Min": [s.min() for s in scores.values()],
        "Max": [s.max() for s in scores.values()]
    })
    if all(len(s) > 0 for s in scores.values()):
        F_stat, p_val = f_oneway(*scores.values())
    else:
        F_stat, p_val = np.nan, np.nan
    anova = pd.DataFrame({
        "F-statistic": [round(F_stat, 3) if not np.isnan(F_stat) else np.nan],
        "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan]
    })
    all_output.append((question, desc_stats, anova))

# ---------------------------
# Write all results to one Excel tab
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
try:
    wb = openpyxl.load_workbook(output_filename)
    wb.close()
    mode = 'a'
except FileNotFoundError:
    mode = 'w'

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet_name = "W3-Less is More"
    startrow = 0
    for question, desc_df, anova_df in all_output:
        # Write header for each question
        pd.DataFrame({f"Analysis: {question} (Wave 3)": [""]}).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += 2
        desc_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(desc_df) + 2
        anova_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(anova_df) + 4

print(f"Excel file '{output_filename}' updated with all results in the 'W3-Less is More' tab.")


# ---------------------------
# LLM
# ---------------------------

for col in formA_cols + formB_cols + formC_cols:
    new_col = col + "_num"
    df_llm[new_col] = df_llm[col].apply(map_less_is_more)

df_llm["FormA_avg"] = get_form_response(df_llm, formA_cols)
df_llm["FormB_avg"] = get_form_response(df_llm, formB_cols)
df_llm["FormC_avg"] = get_form_response(df_llm, formC_cols)
df_llm["Form_Assignment"] = df_llm.apply(assign_form, axis=1)
df_llm_valid = df_llm[df_llm["Form_Assignment"].notna()]

# Analyze each question separately for LLM
all_output_llm = []
for question, mapping in question_mappings.items():
    scores = get_question_scores(df_llm_valid, mapping)
    desc_stats = pd.DataFrame({
        "Form": list(scores.keys()),
        "n": [s.size for s in scores.values()],
        "Mean Score": [round(s.mean(), 2) for s in scores.values()],
        "SD": [round(s.std(ddof=1), 2) for s in scores.values()],
        "Median": [round(s.median(), 2) for s in scores.values()],
        "Min": [s.min() for s in scores.values()],
        "Max": [s.max() for s in scores.values()]
    })
    if sum(len(s) > 0 for s in scores.values()) >= 2:
        available_scores = [s for s in scores.values() if len(s) > 0]
        F_stat, p_val = f_oneway(*available_scores)
        k = len(available_scores)
        n = sum([len(s) for s in available_scores])
        df_between = k - 1
        df_within = n - k
    else:
        F_stat, p_val, df_between, df_within = np.nan, np.nan, np.nan, np.nan    
    anova = pd.DataFrame({
        "F-statistic": [round(F_stat, 3) if not np.isnan(F_stat) else np.nan],
        "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan],
        "df_between": [df_between],
        "df_within": [df_within]
    })
    all_output_llm.append((question, desc_stats, anova))

# Write LLM results into Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet_name = "W3-Less is More"
    startrow = writer.sheets[sheet_name].max_row + 5
    for question, desc_df, anova_df in all_output_llm:
        pd.DataFrame({f"Analysis: {question} (LLM)": [""]}).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += 2
        desc_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(desc_df) + 2
        anova_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(anova_df) + 4



# ---------------------------
# Wave 4
# ---------------------------

for col in formA_cols + formB_cols + formC_cols:
    new_col = col + "_num"
    df_wave4[new_col] = df_wave4[col].apply(map_less_is_more)

df_wave4["FormA_avg"] = get_form_response(df_wave4, formA_cols)
df_wave4["FormB_avg"] = get_form_response(df_wave4, formB_cols)
df_wave4["FormC_avg"] = get_form_response(df_wave4, formC_cols)
df_wave4["Form_Assignment"] = df_wave4.apply(assign_form, axis=1)
df_wave4_valid = df_wave4[df_wave4["Form_Assignment"].notna()]

# Analyze each question separately for Wave 4
all_output_wave4 = []
for question, mapping in question_mappings.items():
    scores = get_question_scores(df_wave4_valid, mapping)
    desc_stats = pd.DataFrame({
        "Form": list(scores.keys()),
        "n": [s.size for s in scores.values()],
        "Mean Score": [round(s.mean(), 2) for s in scores.values()],
        "SD": [round(s.std(ddof=1), 2) for s in scores.values()],
        "Median": [round(s.median(), 2) for s in scores.values()],
        "Min": [s.min() for s in scores.values()],
        "Max": [s.max() for s in scores.values()]
    })
    if sum(len(s) > 0 for s in scores.values()) >= 2:
        available_scores = [s for s in scores.values() if len(s) > 0]
        F_stat, p_val = f_oneway(*available_scores)
        k = len(available_scores)
        n = sum([len(s) for s in available_scores])
        df_between = k - 1
        df_within = n - k
    else:
        F_stat, p_val, df_between, df_within = np.nan, np.nan, np.nan, np.nan    
    anova = pd.DataFrame({
        "F-statistic": [round(F_stat, 3) if not np.isnan(F_stat) else np.nan],
        "p-value": [round(p_val, 3) if not np.isnan(p_val) else np.nan],
        "df_between": [df_between],
        "df_within": [df_within]
    })
    all_output_wave4.append((question, desc_stats, anova))

# Write Wave 4 results into Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet_name = "W3-Less is More"
    startrow = writer.sheets[sheet_name].max_row + 5
    for question, desc_df, anova_df in all_output_wave4:
        pd.DataFrame({f"Analysis: {question} (Wave 4)": [""]}).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += 2
        desc_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(desc_df) + 2
        anova_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(anova_df) + 4

# ---------------------------
# Prepare response distributions for Q171_num to Q179_num for wave 3
# ---------------------------
distribution_data_w3 = []

def collect_distributions(df, question_col, form_name):
    value_counts = df[question_col].value_counts(dropna=False).sort_index()
    total = df[question_col].notna().sum()
    for val, count in value_counts.items():
        distribution_data_w3.append({
            "Form": form_name,
            "Question": question_col,
            "Response": val,
            "Count": count
        })
    distribution_data_w3.append({
        "Form": form_name,
        "Question": question_col,
        "Response": "Non-missing total",
        "Count": total
    })

# Collect for each form and question
for form, cols in {
    "Form A": ["Q171_num", "Q174_num", "Q177_num"],
    "Form B": ["Q172_num", "Q175_num", "Q178_num"],
    "Form C": ["Q173_num", "Q176_num", "Q179_num"]
}.items():
    for col in cols:
        subset = df_wave3_valid[df_wave3_valid["Form_Assignment"] == form]
        collect_distributions(subset, col, form)

# Convert to DataFrame
distribution_df_w3 = pd.DataFrame(distribution_data_w3)

# ---------------------------
# Append to Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow, column=1, value="Response Distributions for Wave 3 (Q171–Q179)").font = Font(bold=True, size=12)
    startrow += 2
    distribution_df_w3.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)

# ---------------------------
# Prepare response distributions for Q171_num to Q179_num for LLM
# ---------------------------
distribution_data_llm = []

def collect_distributions(df, question_col, form_name):
    value_counts = df[question_col].value_counts(dropna=False).sort_index()
    total = df[question_col].notna().sum()
    for val, count in value_counts.items():
        distribution_data_llm.append({
            "Form": form_name,
            "Question": question_col,
            "Response": val,
            "Count": count
        })
    distribution_data_llm.append({
        "Form": form_name,
        "Question": question_col,
        "Response": "Non-missing total",
        "Count": total
    })

# Collect for each form and question
for form, cols in {
    "Form A": ["Q171_num", "Q174_num", "Q177_num"],
    "Form B": ["Q172_num", "Q175_num", "Q178_num"],
    "Form C": ["Q173_num", "Q176_num", "Q179_num"]
}.items():
    for col in cols:
        subset = df_llm_valid[df_llm_valid["Form_Assignment"] == form]
        collect_distributions(subset, col, form)

# Convert to DataFrame
distribution_df_llm = pd.DataFrame(distribution_data_llm)

# ---------------------------
# Append to Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow, column=1, value="Response Distributions for LLM (Q171–Q179)").font = Font(bold=True, size=12)
    startrow += 2
    distribution_df_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)


# ---------------------------
# Prepare response distributions for Q171_num to Q179_num for wave 4
# ---------------------------
distribution_data_w4 = []

def collect_distributions(df, question_col, form_name):
    value_counts = df[question_col].value_counts(dropna=False).sort_index()
    total = df[question_col].notna().sum()
    for val, count in value_counts.items():
        distribution_data_w4.append({
            "Form": form_name,
            "Question": question_col,
            "Response": val,
            "Count": count
        })
    distribution_data_w4.append({
        "Form": form_name,
        "Question": question_col,
        "Response": "Non-missing total",
        "Count": total
    })

# Collect for each form and question
for form, cols in {
    "Form A": ["Q171_num", "Q174_num", "Q177_num"],
    "Form B": ["Q172_num", "Q175_num", "Q178_num"],
    "Form C": ["Q173_num", "Q176_num", "Q179_num"]
}.items():
    for col in cols:
        subset = df_wave4_valid[df_wave4_valid["Form_Assignment"] == form]
        collect_distributions(subset, col, form)

# Convert to DataFrame
distribution_df_w4 = pd.DataFrame(distribution_data_w4)

# ---------------------------
# Append to Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name]
    startrow = ws.max_row + 3
    ws.cell(row=startrow, column=1, value="Response Distributions for Wave 4 (Q171–Q179)").font = Font(bold=True, size=12)
    startrow += 2
    distribution_df_w4.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)


print("done")


# ## WTA/WTP Thaler problem

# In[86]:


import pandas as pd
import numpy as np
from scipy.stats import ttest_ind
import openpyxl
import random
from scipy.stats import pearsonr, norm


# ---------------------------
# Thaler problem columns:
#   - Q189 => WTP-certainty
#   - Q190 => WTA-certainty
#   - Q191 => WTP-noncertainty
#
# Each subject sees only one of these three columns.
# ---------------------------
wtp_cert_col = "Q189"
wta_cert_col = "Q190"
wtp_noncert_col = "Q191"

# ---------------------------
# Mapping from text responses to integer 1–10
# (Matches the scale in the original paper)
# ---------------------------
thaler_mapping = {
    "$5,000,000 or more": 10,
    "$1,000,000": 9,
    "$500,000": 8,
    "$250,000": 7,
    "$100,000": 6,
    "$50,000": 5,
    "$10,000": 4,
    "$1,000": 3,
    "$100": 2,
    "$10": 1
}

def map_thaler_response(resp):
    """Return an integer 1–10 based on the text response (e.g. '$100', '$5,000,000 or more')."""
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    for key, val in thaler_mapping.items():
        # We'll do a substring match ignoring commas/spaces
        # E.g. '$5,000,000 or more' -> '5,000,000'
        # We'll remove commas/spaces from both sides
        short_key = key.lower().replace(",", "").replace(" ", "")
        short_text = text.replace(",", "").replace(" ", "")
        if short_key in short_text:
            return val
    return np.nan

# Map each of the three columns
df_wave3["WTP_cert_mapped"] = df_wave3[wtp_cert_col].apply(map_thaler_response)
df_wave3["WTA_cert_mapped"] = df_wave3[wta_cert_col].apply(map_thaler_response)
df_wave3["WTP_noncert_mapped"] = df_wave3[wtp_noncert_col].apply(map_thaler_response)

# ---------------------------
# Separate the data into three groups
# (XOR across the three columns so each subject is in only one group)
# ---------------------------
mask = (
    df_wave3["WTP_cert_mapped"].notna() & df_wave3["WTA_cert_mapped"].isna() & df_wave3["WTP_noncert_mapped"].isna()
) | (
    df_wave3["WTP_cert_mapped"].isna() & df_wave3["WTA_cert_mapped"].notna() & df_wave3["WTP_noncert_mapped"].isna()
) | (
    df_wave3["WTP_cert_mapped"].isna() & df_wave3["WTA_cert_mapped"].isna() & df_wave3["WTP_noncert_mapped"].notna()
)

df_wave3 = df_wave3[mask]

group_wtp_cert = df_wave3.loc[df_wave3["WTP_cert_mapped"].notna(), "WTP_cert_mapped"]
group_wta_cert = df_wave3.loc[df_wave3["WTA_cert_mapped"].notna(), "WTA_cert_mapped"]
group_wtp_noncert = df_wave3.loc[df_wave3["WTP_noncert_mapped"].notna(), "WTP_noncert_mapped"]

# ---------------------------
# Compute descriptive stats
# ---------------------------
def desc_stats(series):
    return {
        "n": series.shape[0],
        "Mean": series.mean(),
        "SD": series.std(ddof=1)
    }

desc_wtp_cert = desc_stats(group_wtp_cert)
desc_wta_cert = desc_stats(group_wta_cert)
desc_wtp_noncert = desc_stats(group_wtp_noncert)

desc_table = pd.DataFrame({
    "Condition": ["WTP-certainty (Q189)", "WTA-certainty (Q190)", "WTP-noncertainty (Q191)"],
    "n": [desc_wtp_cert["n"], desc_wta_cert["n"], desc_wtp_noncert["n"]],
    "Mean": [round(desc_wtp_cert["Mean"], 2), round(desc_wta_cert["Mean"], 2), round(desc_wtp_noncert["Mean"], 2)],
    "SD": [round(desc_wtp_cert["SD"], 2), round(desc_wta_cert["SD"], 2), round(desc_wtp_noncert["SD"], 2)]
})

# ---------------------------
# Two t-tests as in the paper:
# 1) WTA-certainty vs. WTP-certainty
# 2) WTP-certainty vs. WTP-noncertainty
# ---------------------------
def run_ttest(series1, series2):
    if series1.shape[0] > 1 and series2.shape[0] > 1:
        t_stat, p_val = ttest_ind(series1, series2, equal_var=False)
    else:
        t_stat, p_val = np.nan, np.nan
    return (round(t_stat,3), round(p_val,3))

t_wta_wtp, p_wta_wtp = run_ttest(group_wta_cert, group_wtp_cert)
t_wtp_cert_noncert, p_wtp_cert_noncert = run_ttest(group_wtp_cert, group_wtp_noncert)

ttest_table = pd.DataFrame({
    "Comparison": ["WTA-certainty vs. WTP-certainty", "WTP-certainty vs. WTP-noncertainty"],
    "t-statistic": [t_wta_wtp, t_wtp_cert_noncert],
    "p-value": [p_wta_wtp, p_wtp_cert_noncert]
})

# ---------------------------
# Combine descriptive stats and t-tests into one final DataFrame
# ---------------------------
blank_row = pd.DataFrame([["","","",""]], columns=desc_table.columns)
final_combined = pd.concat([desc_table, blank_row, ttest_table], ignore_index=True)

# ---------------------------
# Header note
# ---------------------------
header_note = (
    "Analysis: Thaler Problem (Wave 3)\n\n"
    "Three conditions (each subject sees one):\n"
    "  - WTP-certainty (Q189)\n"
    "  - WTA-certainty (Q190)\n"
    "  - WTP-noncertainty (Q191)\n\n"
    "Text responses (e.g. '$10,000', '$5,000,000 or more') are mapped to 1–10 as in the paper:\n"
    "  1 = $10\n"
    "  2 = $100\n"
    "  3 = $1,000\n"
    "  4 = $10,000\n"
    "  5 = $50,000\n"
    "  6 = $100,000\n"
    "  7 = $250,000\n"
    "  8 = $500,000\n"
    "  9 = $1,000,000\n"
    "  10 = $5,000,000 or more\n\n"
    "We replicate the original two t-tests:\n"
    "  1) WTA-certainty vs. WTP-certainty\n"
    "  2) WTP-certainty vs. WTP-noncertainty\n\n"
)

# ---------------------------
# Determine a unique sheet name and write all results
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W3-Thalerproblem"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    final_combined.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")


# ---------------------------
# Wave 4
# ---------------------------

# Map responses
df_wave4["WTP_cert_mapped"] = df_wave4[wtp_cert_col].apply(map_thaler_response)
df_wave4["WTA_cert_mapped"] = df_wave4[wta_cert_col].apply(map_thaler_response)
df_wave4["WTP_noncert_mapped"] = df_wave4[wtp_noncert_col].apply(map_thaler_response)

# Filter valid respondents (only one condition per person)
mask_w4 = (
    (df_wave4["WTP_cert_mapped"].notna() & df_wave4["WTA_cert_mapped"].isna() & df_wave4["WTP_noncert_mapped"].isna())
    | (df_wave4["WTP_cert_mapped"].isna() & df_wave4["WTA_cert_mapped"].notna() & df_wave4["WTP_noncert_mapped"].isna())
    | (df_wave4["WTP_cert_mapped"].isna() & df_wave4["WTA_cert_mapped"].isna() & df_wave4["WTP_noncert_mapped"].notna())
)
df_wave4 = df_wave4[mask_w4]

# Descriptive statistics for Wave 4

def desc_stats(series):
    return {
        "n": series.shape[0],
        "Mean": series.mean(),
        "SD": series.std(ddof=1)
    }

group_wtp_cert_w4 = df_wave4.loc[df_wave4["WTP_cert_mapped"].notna(), "WTP_cert_mapped"]
group_wta_cert_w4 = df_wave4.loc[df_wave4["WTA_cert_mapped"].notna(), "WTA_cert_mapped"]
group_wtp_noncert_w4 = df_wave4.loc[df_wave4["WTP_noncert_mapped"].notna(), "WTP_noncert_mapped"]

desc_table_w4 = pd.DataFrame({
    "Condition": ["WTP-certainty (Q189)", "WTA-certainty (Q190)", "WTP-noncertainty (Q191)"],
    "n": [group_wtp_cert_w4.shape[0], group_wta_cert_w4.shape[0], group_wtp_noncert_w4.shape[0]],
    "Mean": [round(group_wtp_cert_w4.mean(), 2), round(group_wta_cert_w4.mean(), 2), round(group_wtp_noncert_w4.mean(), 2)],
    "SD": [round(group_wtp_cert_w4.std(ddof=1), 2), round(group_wta_cert_w4.std(ddof=1), 2), round(group_wtp_noncert_w4.std(ddof=1), 2)]
})

# T-tests for Wave 4
def run_ttest(series1, series2):
    if series1.shape[0] > 1 and series2.shape[0] > 1:
        t_stat, p_val = ttest_ind(series1, series2, equal_var=False)
    else:
        t_stat, p_val = np.nan, np.nan
    return (round(t_stat, 3), round(p_val, 3))

t_wta_wtp_w4, p_wta_wtp_w4 = run_ttest(group_wta_cert_w4, group_wtp_cert_w4)
t_wtp_cert_noncert_w4, p_wtp_cert_noncert_w4 = run_ttest(group_wtp_cert_w4, group_wtp_noncert_w4)

ttest_table_w4 = pd.DataFrame({
    "Comparison": ["WTA-certainty vs. WTP-certainty", "WTP-certainty vs. WTP-noncertainty"],
    "t-statistic": [t_wta_wtp_w4, t_wtp_cert_noncert_w4],
    "p-value": [p_wta_wtp_w4, p_wtp_cert_noncert_w4]
})


# ---------------------------
# LLM Data
# ---------------------------

# Map responses
df_llm["WTP_cert_mapped"] = df_llm[wtp_cert_col].apply(map_thaler_response)
df_llm["WTA_cert_mapped"] = df_llm[wta_cert_col].apply(map_thaler_response)
df_llm["WTP_noncert_mapped"] = df_llm[wtp_noncert_col].apply(map_thaler_response)

# Filter valid respondents (only one condition per person)
mask_llm = (
    (df_llm["WTP_cert_mapped"].notna() & df_llm["WTA_cert_mapped"].isna() & df_llm["WTP_noncert_mapped"].isna()) |
    (df_llm["WTP_cert_mapped"].isna() & df_llm["WTA_cert_mapped"].notna() & df_llm["WTP_noncert_mapped"].isna()) |
    (df_llm["WTP_cert_mapped"].isna() & df_llm["WTA_cert_mapped"].isna() & df_llm["WTP_noncert_mapped"].notna())
)
df_llm = df_llm[mask_llm]

# Descriptive statistics for LLM
group_wtp_cert_llm = df_llm.loc[df_llm["WTP_cert_mapped"].notna(), "WTP_cert_mapped"]
group_wta_cert_llm = df_llm.loc[df_llm["WTA_cert_mapped"].notna(), "WTA_cert_mapped"]
group_wtp_noncert_llm = df_llm.loc[df_llm["WTP_noncert_mapped"].notna(), "WTP_noncert_mapped"]

desc_table_llm = pd.DataFrame({
    "Condition": ["WTP-certainty (Q189)", "WTA-certainty (Q190)", "WTP-noncertainty (Q191)"],
    "n": [group_wtp_cert_llm.shape[0], group_wta_cert_llm.shape[0], group_wtp_noncert_llm.shape[0]],
    "Mean": [round(group_wtp_cert_llm.mean(), 2), round(group_wta_cert_llm.mean(), 2), round(group_wtp_noncert_llm.mean(), 2)],
    "SD": [round(group_wtp_cert_llm.std(ddof=1), 2), round(group_wta_cert_llm.std(ddof=1), 2), round(group_wtp_noncert_llm.std(ddof=1), 2)]
})

# T-tests for LLM
t_wta_wtp_llm, p_wta_wtp_llm = run_ttest(group_wta_cert_llm, group_wtp_cert_llm)
t_wtp_cert_noncert_llm, p_wtp_cert_noncert_llm = run_ttest(group_wtp_cert_llm, group_wtp_noncert_llm)

ttest_table_llm = pd.DataFrame({
    "Comparison": ["WTA-certainty vs. WTP-certainty", "WTP-certainty vs. WTP-noncertainty"],
    "t-statistic": [t_wta_wtp_llm, t_wtp_cert_noncert_llm],
    "p-value": [p_wta_wtp_llm, p_wtp_cert_noncert_llm]
})

# Write LLM results to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet_name = sheet_name_new
    startrow = writer.sheets[sheet_name].max_row + 5
    writer.sheets[sheet_name].cell(row=startrow, column=1, value="LLM Descriptive Statistics and T-Tests")
    startrow += 2
    desc_table_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(desc_table_llm) + 3
    ttest_table_llm.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)


# ---------------------------
# Response Distributions
# ---------------------------
def get_response_distribution(series):
    return series.value_counts().sort_index().rename("Count")

distribution_tables = []

for wave_label, df, prefix in [("Wave 3", df_wave3, "W3"), ("LLM", df_llm, "LLM"),("Wave 4", df_wave4, "W4")]:
    for cond, colname in zip(["WTP-certainty", "WTA-certainty", "WTP-noncertainty"],
                              [wtp_cert_col, wta_cert_col, wtp_noncert_col]):
        mapped_col = {
            wtp_cert_col: "WTP_cert_mapped",
            wta_cert_col: "WTA_cert_mapped",
            wtp_noncert_col: "WTP_noncert_mapped"
        }[colname]
        series = df[mapped_col].dropna()
        dist = get_response_distribution(series).to_frame()
        dist.columns = [f"{prefix} {cond}"]
        distribution_tables.append(dist)

distributions_final = pd.concat(distribution_tables, axis=1)

# ---------------------------
# Write Wave 4 results into Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet = writer.sheets[sheet_name_new]
    startrow = sheet.max_row + 5

    # Write label
    sheet.cell(row=startrow, column=1, value="Wave 4 descriptive stats and t-test")
    startrow += 2

    # Write Wave 4 descriptive table
    desc_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(desc_table_w4) + 3

    # Write t-test table
    ttest_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(ttest_table_w4) + 5

    # Write distributions
    distributions_final.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow)


print(f"Excel file '{output_filename}' updated with Wave 4 analysis and response distributions.")

print("done")


# ## Probability Matching vs Maximizing (Stanovich & West, 2008)

# In[87]:


import pandas as pd
import numpy as np
import openpyxl
from statsmodels.stats.proportion import proportions_ztest
from tqdm import tqdm
from scipy.stats import pearsonr
from tqdm.notebook import tqdm
import random
from scipy.stats import pearsonr

#df_wave3 = df_wave3[df_wave3["PROLIFIC_PID"].isin(common_ids)]

# Column sets
card_cols = [f"Q198_{i}" for i in range(1, 11)]
dice_cols = [f"Q203_{i}" for i in range(1, 7)]

# ✅ Prioritize rows with more complete task data per participant
df_wave3["nonmissing_card"] = df_wave3[card_cols].notna().sum(axis=1)
df_wave3["nonmissing_dice"] = df_wave3[dice_cols].notna().sum(axis=1)

# Sort by completeness
df_wave3 = df_wave3.sort_values(by=["nonmissing_card", "nonmissing_dice"], ascending=False)

# Drop duplicates, keeping the most complete row per PROLIFIC_PID
df_wave3 = df_wave3.drop_duplicates(subset="PROLIFIC_PID", keep="first")

# Clean up helper columns
df_wave3 = df_wave3.drop(columns=["nonmissing_card", "nonmissing_dice"])

# ---------------------------
# We'll define two classification functions (MATCH, MAX, OTHER)
# ---------------------------
def classify_card_responses(responses):
    """
    responses: list of strings of length 10, each either '1' or '2' (or NaN).
    Return 'MATCH' if exactly 7 '1' and 3 '2'.
    Return 'MAX' if all 10 are '1'.
    Else 'OTHER'.
    """
    if len(responses) < 10 or any(pd.isna(x) for x in responses):
        return np.nan
    n1 = sum(x.strip() == "1" for x in responses)
    n2 = sum(x.strip() == "2" for x in responses)
    if n1 == 10:
        return "MAX"
    elif n1 == 7 and n2 == 3:
        return "MATCH"
    else:
        return "OTHER"

def classify_dice_responses(responses):
    """
    responses: list of strings of length 6, each either 'red' or 'green' (or NaN).
    Return 'MATCH' if exactly 4 'red' and 2 'green'.
    Return 'MAX' if all 6 are 'red'.
    Else 'OTHER'.
    """
    if len(responses) < 6 or any(pd.isna(x) for x in responses):
        return np.nan
    n_red = sum(x.strip().lower() == "red" for x in responses)
    n_green = sum(x.strip().lower() == "green" for x in responses)
    if n_red == 6:
        return "MAX"
    elif n_red == 4 and n_green == 2:
        return "MATCH"
    else:
        return "OTHER"

# ---------------------------
# We'll apply XOR logic to see who answered card vs. dice
# We'll store in "Task" ∈ {Card, Dice, BOTH, None} and "Strategy" ∈ {MATCH, MAX, OTHER, np.nan}
# ---------------------------
classifications = []
for idx, row in df_wave3.iterrows():
    card_answers = [row[c] for c in card_cols]
    dice_answers = [row[c] for c in dice_cols]
    n_card_nonmissing = sum(not pd.isna(x) for x in card_answers)
    n_dice_nonmissing = sum(not pd.isna(x) for x in dice_answers)
    
    if n_card_nonmissing > 0 and n_dice_nonmissing > 0:
        # Shouldn't happen if design is correct
        classifications.append(("BOTH", np.nan))
    elif n_card_nonmissing == 10:
        # Classify card
        strategy = classify_card_responses(card_answers)
        classifications.append(("Card", strategy))
    elif n_dice_nonmissing == 6:
        # Classify dice
        strategy = classify_dice_responses(dice_answers)
        classifications.append(("Dice", strategy))
    else:
        classifications.append(("None", np.nan))

df_wave3["Task"] = [c[0] for c in classifications]
df_wave3["Strategy"] = [c[1] for c in classifications]

# ---------------------------
# Summarize frequencies (Task x Strategy)
# We'll produce frequency tables for Card, Dice
# ---------------------------
card_strategies = df_wave3.loc[df_wave3["Task"]=="Card", "Strategy"].value_counts(dropna=True)
dice_strategies = df_wave3.loc[df_wave3["Task"]=="Dice", "Strategy"].value_counts(dropna=True)

# Reformat
card_strategies = card_strategies.rename_axis("Strategy").reset_index(name="Count")
dice_strategies = dice_strategies.rename_axis("Strategy").reset_index(name="Count")
card_strategies["Task"] = "Card Problem"
dice_strategies["Task"] = "Dice Problem"

def reorder_strategies(df):
    order = ["MATCH", "MAX", "OTHER"]
    df["SortOrder"] = df["Strategy"].apply(lambda x: order.index(x) if x in order else 999)
    df = df.sort_values("SortOrder")
    df = df.drop(columns=["SortOrder"])
    return df

card_strategies = reorder_strategies(card_strategies)
dice_strategies = reorder_strategies(dice_strategies)

final_counts = pd.concat([card_strategies, dice_strategies], ignore_index=True)
final_counts = final_counts[["Task", "Strategy", "Count"]]

# ---------------------------
# Compute 95% confidence intervals for MAX proportion
# ---------------------------
from statsmodels.stats.proportion import proportion_confint

# Pivot counts to extract MAX and total per task
counts_pivot = final_counts.pivot(index="Task", columns="Strategy", values="Count").fillna(0)

def ci_for_max(task_label):
    row = counts_pivot.loc[task_label]
    n_MAX = row.get("MAX", 0)
    total = row.sum()
    if total == 0:
        return np.nan, np.nan, np.nan
    prop = n_MAX / total
    ci_low, ci_high = proportion_confint(count=n_MAX, nobs=total, alpha=0.05, method="wilson")
    return round(prop, 3), round(ci_low, 3), round(ci_high, 3)

card_prop, card_ci_low, card_ci_high = ci_for_max("Card Problem")
dice_prop, dice_ci_low, dice_ci_high = ci_for_max("Dice Problem")

test_results = [
    {
        "Task": "Card Problem",
        "Test": "MAX % with 95% CI",
        "Proportion": card_prop,
        "95% CI Lower": card_ci_low,
        "95% CI Upper": card_ci_high
    },
    {
        "Task": "Dice Problem",
        "Test": "MAX % with 95% CI",
        "Proportion": dice_prop,
        "95% CI Lower": dice_ci_low,
        "95% CI Upper": dice_ci_high
    }
]
test_results_df = pd.DataFrame(test_results)


# We'll combine final_counts with a blank row, then test_results_df
blank_row_df = pd.DataFrame([["","",""]], columns=final_counts.columns)
combined_output = pd.concat([final_counts, blank_row_df, test_results_df], ignore_index=True)

# ---------------------------
# Prepare header note
# ---------------------------
header_note = (
    "Analysis: Probability Matching vs. Maximizing \n\n"
    "Card Problem (Q198_1..Q198_10):\n"
    "  - Distribution: 7 '1', 3 '2'. MATCH if exactly 7 '1' & 3 '2', MAX if all '1', OTHER otherwise.\n"
    "Dice Problem (Q203_1..Q203_6):\n"
    "  - Distribution: 4 'red', 2 'green'. MATCH if exactly 4 'red' & 2 'green', MAX if all 'red', OTHER otherwise.\n\n"
    "Frequencies are shown below for each strategy in each task. The 2-prop z-test compares\n"
    "the proportion using MAX vs. the proportion using MATCH+OTHER (although the counts seem to speak for themselves).\n"
    "Stanovich & West observed that, even though the MAX strategy is normative, a majority of subjects used MATCH or OTHER."
)

# ---------------------------
# Write everything in one pass
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W3-PMatchvsMax"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

print("Final chosen sheet name:", sheet_name_new)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    combined_output.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")



# =================================
# Additional code to analyze Wave 4
# =================================

# ---------------------------
# Classify Wave 4 responses
# ---------------------------
classifications_w4 = []
for idx, row in df_wave4.iterrows():
    card_answers = [row.get(c, np.nan) for c in card_cols]
    dice_answers = [row.get(c, np.nan) for c in dice_cols]
    n_card_nonmissing = sum(not pd.isna(x) for x in card_answers)
    n_dice_nonmissing = sum(not pd.isna(x) for x in dice_answers)

    if n_card_nonmissing > 0 and n_dice_nonmissing > 0:
        classifications_w4.append(("BOTH", np.nan))
    elif n_card_nonmissing == 10:
        strategy = classify_card_responses(card_answers)
        classifications_w4.append(("Card", strategy))
    elif n_dice_nonmissing == 6:
        strategy = classify_dice_responses(dice_answers)
        classifications_w4.append(("Dice", strategy))
    else:
        classifications_w4.append(("None", np.nan))

df_wave4["Task"] = [c[0] for c in classifications_w4]
df_wave4["Strategy"] = [c[1] for c in classifications_w4]

# ---------------------------
# Summarize Wave 4 Frequencies
# ---------------------------
card_strategies_w4 = df_wave4.loc[df_wave4["Task"]=="Card", "Strategy"].value_counts(dropna=True)
dice_strategies_w4 = df_wave4.loc[df_wave4["Task"]=="Dice", "Strategy"].value_counts(dropna=True)

card_strategies_w4 = card_strategies_w4.rename_axis("Strategy").reset_index(name="Count")
dice_strategies_w4 = dice_strategies_w4.rename_axis("Strategy").reset_index(name="Count")
card_strategies_w4["Task"] = "Card Problem (Wave 4)"
dice_strategies_w4["Task"] = "Dice Problem (Wave 4)"

card_strategies_w4 = reorder_strategies(card_strategies_w4)
dice_strategies_w4 = reorder_strategies(dice_strategies_w4)

final_counts_w4 = pd.concat([card_strategies_w4, dice_strategies_w4], ignore_index=True)
final_counts_w4 = final_counts_w4[["Task", "Strategy", "Count"]]

# ---------------------------
# Compute 95% confidence intervals for Wave 4
# ---------------------------
counts_pivot_w4 = final_counts_w4.pivot(index="Task", columns="Strategy", values="Count").fillna(0)

def ci_for_max(task_label, pivot_table):
    row = pivot_table.loc[task_label]
    n_MAX = row.get("MAX", 0)
    total = row.sum()
    if total == 0:
        return np.nan, np.nan, np.nan
    prop = n_MAX / total
    ci_low, ci_high = proportion_confint(count=n_MAX, nobs=total, alpha=0.05, method="wilson")
    return round(prop, 3), round(ci_low, 3), round(ci_high, 3)

card_prop_w4, card_ci_low_w4, card_ci_high_w4 = ci_for_max("Card Problem (Wave 4)", counts_pivot_w4)
dice_prop_w4, dice_ci_low_w4, dice_ci_high_w4 = ci_for_max("Dice Problem (Wave 4)", counts_pivot_w4)

test_results_w4 = [
    {
        "Task": "Card Problem (Wave 4)",
        "Test": "MAX % with 95% CI",
        "Proportion": card_prop_w4,
        "95% CI Lower": card_ci_low_w4,
        "95% CI Upper": card_ci_high_w4
    },
    {
        "Task": "Dice Problem (Wave 4)",
        "Test": "MAX % with 95% CI",
        "Proportion": dice_prop_w4,
        "95% CI Lower": dice_ci_low_w4,
        "95% CI Upper": dice_ci_high_w4
    }
]
test_results_df_w4 = pd.DataFrame(test_results_w4)


# =================================
# Additional code to analyze LLM data
# =================================

# Fix numeric-looking strings like '1.0' to clean strings like '1'
for col in [f"Q198_{i}" for i in range(1, 11)]:
    df_llm[col] = df_llm[col].apply(lambda x: str(int(float(x))) if pd.notna(x) else np.nan)


# --- Classify LLM responses ---
classifications_llm = []
for idx, row in df_llm.iterrows():
    card_answers = [row.get(c, np.nan) for c in card_cols]
    dice_answers = [row.get(c, np.nan) for c in dice_cols]
    n_card_nonmissing = sum(not pd.isna(x) for x in card_answers)
    n_dice_nonmissing = sum(not pd.isna(x) for x in dice_answers)

    if n_card_nonmissing > 0 and n_dice_nonmissing > 0:
        classifications_llm.append(("BOTH", np.nan))
    elif n_card_nonmissing == 10:
        strategy = classify_card_responses(card_answers)
        classifications_llm.append(("Card", strategy))
    elif n_dice_nonmissing == 6:
        strategy = classify_dice_responses(dice_answers)
        classifications_llm.append(("Dice", strategy))
    else:
        classifications_llm.append(("None", np.nan))

df_llm["Task"] = [c[0] for c in classifications_llm]
df_llm["Strategy"] = [c[1] for c in classifications_llm]

# --- Frequency summaries ---
card_strategies_llm = df_llm.loc[df_llm["Task"] == "Card", "Strategy"].value_counts(dropna=True).rename_axis("Strategy").reset_index(name="Count")
dice_strategies_llm = df_llm.loc[df_llm["Task"] == "Dice", "Strategy"].value_counts(dropna=True).rename_axis("Strategy").reset_index(name="Count")

card_strategies_llm["Task"] = "Card Problem (LLM)"
dice_strategies_llm["Task"] = "Dice Problem (LLM)"

card_strategies_llm = reorder_strategies(card_strategies_llm)
dice_strategies_llm = reorder_strategies(dice_strategies_llm)

final_counts_llm = pd.concat([card_strategies_llm, dice_strategies_llm], ignore_index=True)
final_counts_llm = final_counts_llm[["Task", "Strategy", "Count"]]

# --- Compute 95% confidence intervals ---
counts_pivot_llm = final_counts_llm.pivot(index="Task", columns="Strategy", values="Count").fillna(0)

card_prop_llm, card_ci_low_llm, card_ci_high_llm = ci_for_max("Card Problem (LLM)", counts_pivot_llm)
dice_prop_llm, dice_ci_low_llm, dice_ci_high_llm = ci_for_max("Dice Problem (LLM)", counts_pivot_llm)

test_results_llm = [
    {
        "Task": "Card Problem (LLM)",
        "Test": "MAX % with 95% CI",
        "Proportion": card_prop_llm,
        "95% CI Lower": card_ci_low_llm,
        "95% CI Upper": card_ci_high_llm
    },
    {
        "Task": "Dice Problem (LLM)",
        "Test": "MAX % with 95% CI",
        "Proportion": dice_prop_llm,
        "95% CI Lower": dice_ci_low_llm,
        "95% CI Upper": dice_ci_high_llm
    }
]
test_results_df_llm = pd.DataFrame(test_results_llm)

# --- Combine and write to Excel ---
blank_row_llm = pd.DataFrame([["", "", ""]], columns=final_counts_llm.columns)
combined_output_llm = pd.concat([final_counts_llm, blank_row_llm, test_results_df_llm], ignore_index=True)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5
    pd.DataFrame({"LLM Analysis": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    combined_output_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)

print("✅ LLM task classification, frequencies, and confidence intervals written to Excel.")


# ---------------------------
# Write everything to Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5

    # Write Wave 4 results
    pd.DataFrame({"Wave 4 Analysis": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    final_counts_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(final_counts_w4) + 3
    test_results_df_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(test_results_df_w4) + 5

print(f"Wave 4 results appended to '{output_filename}' under '{sheet_name_new}' tab.")

# =================================
# Write raw response distributions for each condition and wave
# =================================

def get_raw_response_distribution(df, wave_label):
    rows = []
    card_cols = [f"Q198_{i}" for i in range(1, 11)]
    dice_cols = [f"Q203_{i}" for i in range(1, 7)]

    # Card responses
    card_subset = df[df["Task"] == "Card"]
    card_raw = card_subset[card_cols].stack().dropna().str.strip().str.lower()
    card_counts = card_raw.value_counts().reset_index()
    card_counts.columns = ["Raw Response", "Count"]
    card_counts["Task"] = "Card"
    card_counts["Wave"] = wave_label

    # Dice responses
    dice_subset = df[df["Task"] == "Dice"]
    dice_raw = dice_subset[dice_cols].stack().dropna().str.strip().str.lower()
    dice_counts = dice_raw.value_counts().reset_index()
    dice_counts.columns = ["Raw Response", "Count"]
    dice_counts["Task"] = "Dice"
    dice_counts["Wave"] = wave_label

    return pd.concat([card_counts, dice_counts], ignore_index=True)[["Wave", "Task", "Raw Response", "Count"]]

# Apply to both waves
dist_w3_raw = get_raw_response_distribution(df_wave3, "Wave 3")
dist_w4_raw = get_raw_response_distribution(df_wave4, "Wave 4")
dist_llm_raw = get_raw_response_distribution(df_llm, "LLM")

dist_combined_raw = pd.concat([dist_w3_raw, dist_llm_raw,dist_w4_raw], ignore_index=True)

# Write to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5
    pd.DataFrame({"Raw Response Frequencies (actual strings)": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    dist_combined_raw.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)

print("✅ Raw response string distributions written to Excel.")


print("Done.")



# ## Denominator neglect (Stanovich & West, 2008)

# In[88]:


import pandas as pd
import numpy as np
import openpyxl


# ---------------------------
# The column is "Denominator neglect " (with trailing space)
# Answers: "the small tray" or "the large tray"
# ---------------------------
col_name = "Denominator neglect"

def classify_tray(resp):
    if pd.isna(resp):
        return np.nan
    text = str(resp).lower().strip()
    if "small" in text:
        return "Small tray"
    elif "large" in text:
        return "Large tray"
    else:
        return np.nan

#print("df_wave3 columns:", list(df_wave3.columns))
df_wave3["DenomNeg_tray"] = df_wave3[col_name].apply(classify_tray)

# ---------------------------
# Frequency table
# ---------------------------
freq_table = df_wave3["DenomNeg_tray"].value_counts(dropna=True).reset_index()
freq_table.columns = ["Response", "Count"]
freq_table["Percent"] = (freq_table["Count"] / freq_table["Count"].sum() * 100).round(1)

# ---------------------------
# Combine data with a blank row if desired,
# but here we only have this single frequency table.
# ---------------------------
combined_output = freq_table  # no test results

# ---------------------------
# Header note
# ---------------------------
header_note = (
    "Analysis: Denominator Neglect (Wave 3)\n\n"
    "Participants chose between:\n"
    "  - A small tray (1 black, 9 white => 10% chance of winning $2),\n"
    "  - A large tray (8 black, 92 white => 8% chance).\n\n"
    "The small tray is the normative choice, but following Stanovich & West \n"
    "we note a sizable minority still picks the large tray."
)

# ---------------------------
# Write everything in one pass
# ---------------------------
#output_filename = "experiment_analysis.xlsx"
base_sheet_name = "W3-DenomNeglect"
try:
    wb = openpyxl.load_workbook(output_filename)
    sheet_name_new = base_sheet_name
    suffix = 1
    while sheet_name_new in wb.sheetnames:
        sheet_name_new = f"{base_sheet_name}_v{suffix}"
        suffix += 1
    wb.close()
except FileNotFoundError:
    sheet_name_new = base_sheet_name

print("Final chosen sheet name:", sheet_name_new)

with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    startrow = 3
    combined_output.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    ws = writer.sheets[sheet_name_new]
    ws.cell(row=1, column=1, value=header_note)

print(f"Excel file '{output_filename}' updated with the '{sheet_name_new}' tab.")

# =================================
# Additional code: 95% CI for choosing the Small Tray
# =================================

from scipy.stats import norm

# Inputs
n_large = (df_wave3["DenomNeg_tray"] == "Large tray").sum()
n_total = df_wave3["DenomNeg_tray"].notna().sum()

p_hat = n_large / n_total
z_star = norm.ppf(0.975)  # 1.96 for 95% confidence

# Calculate margin of error
margin = z_star * np.sqrt(p_hat * (1 - p_hat) / n_total)

# Confidence interval
ci_low = round((p_hat - margin) * 100, 2)  # convert to %
ci_high = round((p_hat + margin) * 100, 2)

# Also report proportion
prop_large = round(p_hat * 100, 2)

# Create summary
summary_df = pd.DataFrame({
    "Measure": ["Proportion choosing Large Tray (%)", "95% CI Lower (%)", "95% CI Upper (%)"],
    "Value": [prop_large, ci_low, ci_high]
})

# ---------------------------
# Write to Excel under previous output
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5  # Leave some space
    pd.DataFrame({"Large Tray Choice Summary": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    summary_df.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)

print("Large tray choice proportion and 95% CI added to Excel.")

# =================================
# Additional code: Analyze Wave 4 
# =================================
#col_name = "Denominator neglect"
#print("df_wave4 columns:", list(df_wave4.columns))
df_wave4["DenomNeg_tray"] = df_wave4[col_name].apply(classify_tray)

# ---------------------------
# Frequency table for Wave 4
# ---------------------------
freq_table_w4 = df_wave4["DenomNeg_tray"].value_counts(dropna=True).reset_index()
freq_table_w4.columns = ["Response", "Count"]
freq_table_w4["Percent"] = (freq_table_w4["Count"] / freq_table_w4["Count"].sum() * 100).round(1)

# ---------------------------
# Compute 95% CI for choosing Small Tray in Wave 4
# ---------------------------
n_large_w4 = (df_wave4["DenomNeg_tray"] == "Large tray").sum()
n_total_w4 = df_wave4["DenomNeg_tray"].notna().sum()

p_hat_w4 = n_large_w4 / n_total_w4
z_star = norm.ppf(0.975)

margin_w4 = z_star * np.sqrt(p_hat_w4 * (1 - p_hat_w4) / n_total_w4)

ci_low_w4 = round((p_hat_w4 - margin_w4) * 100, 2)
ci_high_w4 = round((p_hat_w4 + margin_w4) * 100, 2)
prop_large_w4 = round(p_hat_w4 * 100, 2)

summary_df_w4 = pd.DataFrame({
    "Measure": ["Proportion choosing Large Tray (%)", "95% CI Lower (%)", "95% CI Upper (%)"],
    "Value": [prop_large_w4, ci_low_w4, ci_high_w4]
})


# =================================
# Additional code: Analyze LLM
# =================================

# Classify tray choice
df_llm["DenomNeg_tray"] = df_llm["Denominator neglect"].apply(classify_tray)

# Frequency table
freq_table_llm = df_llm["DenomNeg_tray"].value_counts(dropna=True).reset_index()
freq_table_llm.columns = ["Response", "Count"]
freq_table_llm["Percent"] = (freq_table_llm["Count"] / freq_table_llm["Count"].sum() * 100).round(1)

# 95% CI for choosing Large Tray
n_large_llm = (df_llm["DenomNeg_tray"] == "Large tray").sum()
n_total_llm = df_llm["DenomNeg_tray"].notna().sum()

p_hat_llm = n_large_llm / n_total_llm
z_star = norm.ppf(0.975)
margin_llm = z_star * np.sqrt(p_hat_llm * (1 - p_hat_llm) / n_total_llm)

ci_low_llm = round((p_hat_llm - margin_llm) * 100, 2)
ci_high_llm = round((p_hat_llm + margin_llm) * 100, 2)
prop_large_llm = round(p_hat_llm * 100, 2)

summary_df_llm = pd.DataFrame({
    "Measure": ["Proportion choosing Large Tray (%)", "95% CI Lower (%)", "95% CI Upper (%)"],
    "Value": [prop_large_llm, ci_low_llm, ci_high_llm]
})

# Write to Excel
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5

    pd.DataFrame({"LLM Analysis": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    freq_table_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(freq_table_llm) + 2
    summary_df_llm.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)


# ---------------------------
# Write Wave 4 results and correlation to Excel
# ---------------------------
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    ws = writer.sheets[sheet_name_new]
    startrow = ws.max_row + 5  # leave a gap

    # Write Wave 4 header
    pd.DataFrame({"Wave 4 Analysis": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2

    # Write Wave 4 frequency table
    freq_table_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(freq_table_w4) + 2

    # Write Wave 4 small tray summary
    summary_df_w4.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(summary_df_w4) + 5

    # Write distributions for both waves
    pd.DataFrame({"Raw Response Distributions": [""]}).to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += 2
    dist_combined.to_excel(writer, sheet_name=sheet_name_new, startrow=startrow, index=False)
    startrow += len(dist_combined) + 5

print("Wave 4 analysis, raw response distributions written to Excel.")

print("Done.")




# In[ ]:




